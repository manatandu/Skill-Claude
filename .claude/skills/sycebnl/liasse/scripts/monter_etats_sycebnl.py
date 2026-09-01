#!/usr/bin/env python3
"""
monter_etats_sycebnl.py — Monte le jeu COMPLET d'états financiers SYCEBNL des
ASSOCIATIONS ET ORDRES PROFESSIONNELS (Système normal) à partir d'une balance
générale : Bilan (actif et passif sur feuilles séparées), Compte de résultat,
Tableau des flux de trésorerie et LES 35 NOTES ANNEXES OFFICIELLES dans leur
présentation du Journal officiel (Partie 4, ch. 2), plus les feuilles d'audit
GARDE / BALANCE / BALANCE_N1 / CONTROLES / ANOMALIES.

    python monter_etats_sycebnl.py balance_N.xlsx [balance_N-1.xlsx] \
        --sortie etats.xlsx --entite "..." --identifiant "..." \
        --exercice "31/12/N" --duree 12

v3 — chaque poste des états et chaque ligne calculée des notes porte une
FORMULE Excel (SUMIF sur BALANCE / BALANCE_N1) : tout chiffre est retraçable
jusqu'aux comptes de la balance qui l'alimentent. La correspondance
compte → poste reste `references/correspondance-associations.tsv` (vérifiée
contre le Journal officiel, corrections documentées).

TFT : ZA (trésorerie d'ouverture), ZC (investissement), ZD (fonds propres),
ZE (fonds étrangers), ZF et ZG sont calculés en formules dès que la balance
N-1 est fournie ; ZB (activités opérationnelles) est un résidu garanti par
construction (ZB = ZF - ZC - ZD - ZE) — ses lignes de détail FA à FH restent
à saisir depuis le journal de trésorerie, une balance ne les porte pas.
"""

import argparse
import os
import sys

import openpyxl
from openpyxl.styles import Font, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib_mapping_sycebnl import (
    charger_maquette, normaliser_compte, compte_dans_expr, groupe_compte,
)
from formules_sycebnl import (
    formule_tokens, formule_expr, set_lignes_max, q, nom_feuille,
    ecrire_feuille_balance,
    F_TITRE, F_SOUS_TITRE, F_ENTETE, F_NORMAL, F_GRAS,
    R_TITRE, R_ENTETE, R_BANDE, R_TOTAL, BORD_FIN, AL_CENTRE, AL_GAUCHE,
    FMT_MONTANT, style_entetes, style_zone_donnees, style_ligne_total,
    largeurs, style_titre, retirer_tirets, construire_identification,
    construire_fiche_notes, ordonner_feuilles,
    ecrire_cartouche, titre_etat, entetes_bande, style_ligne, fusion, cadre,
    MOYEN, set_identite_etendue, construire_couverture,
    construire_garde_etafi, construire_fiche2, construire_controle_balance,
    construire_table_commentaires, construire_bilan_paysage,
    appliquer_police_arial, numeroter_pages, NOM_BALANCE, NOM_BALANCE_N1,
)
import notes_sycebnl

# Noms de feuilles du classeur produit (présentation du modèle ETAFI).
NOM_ACTIF = "Bilan-Actif"
NOM_PASSIF = "Bilan-Passif"
CR_NOM = "Résultat"

# --------------------------------------------------------------------------
# Lecture souple d'une balance (inchangée)
# --------------------------------------------------------------------------

ENTETES = {
    "compte":  ["cpte", "compte", "n° compte", "numero", "num compte", "code"],
    "libelle": ["intitul", "libell", "designation", "désignation"],
    # soldes d'ouverture (repérés avant les soldes de clôture : « solde
    # d'ouverture débit » ne doit pas être pris pour un « solde débit »)
    "od":      ["ouverture debit", "solde initial debit", "initial debit",
                "a nouveau debit", "a-nouveau debit", "an debit", "si debit",
                "debit ouverture", "debit initial"],
    "oc":      ["ouverture credit", "solde initial credit", "initial credit",
                "a nouveau credit", "a-nouveau credit", "an credit",
                "si credit", "credit ouverture", "credit initial"],
    "md":      ["mouvement debit", "mvt debit", "m.debit", "mouvements debit",
                "debit mouvement", "debit periode", "mouv. debit"],
    "mc":      ["mouvement credit", "mvt credit", "m.credit",
                "mouvements credit", "credit mouvement", "credit periode",
                "mouv. credit"],
    "sd":      ["s.f. debit", "sf debit", "solde final debit", "solde debit",
                "sf_d", "final debit", "debit final", "cloture debit",
                "debit cloture"],
    "sc":      ["s.f. credit", "sf credit", "solde final credit", "solde credit",
                "sf_c", "final credit", "credit final", "cloture credit",
                "credit cloture"],
}


def _norm(x):
    return str(x).strip().lower().replace("é", "e").replace("è", "e")


def _trouver_colonnes(entetes):
    idx = {}
    for i, cell in enumerate(entetes):
        h = _norm(cell)
        for cle, motifs in ENTETES.items():
            if cle in idx:
                continue
            if any(m in h for m in motifs):
                idx[cle] = i
    return idx


def lire_balance(chemin):
    ext = os.path.splitext(chemin)[1].lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(chemin)
        sh = wb.sheet_by_index(0)
        for s in wb.sheets():
            if "balance" in s.name.lower():
                sh = s
                break
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    elif ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(chemin, data_only=True)
        sh = wb.active
        for name in wb.sheetnames:
            if "balance" in name.lower():
                sh = wb[name]
                break
        rows = [[c for c in row] for row in sh.iter_rows(values_only=True)]
    elif ext in (".csv", ".tsv"):
        import csv
        delim = "\t" if ext == ".tsv" else None
        with open(chemin, encoding="utf-8-sig") as f:
            if delim is None:
                sample = f.read(2048)
                f.seek(0)
                delim = "\t" if sample.count("\t") > sample.count(",") else ","
            rows = list(csv.reader(f, delimiter=delim))
    else:
        raise ValueError(f"Extension non gérée : {ext}")

    idx, debut = {}, 0
    for r, row in enumerate(rows[:30]):
        cand = _trouver_colonnes(row)
        if "compte" in cand and ("sd" in cand or "sc" in cand):
            idx, debut = cand, r + 1
            break
    if "compte" not in idx:
        raise ValueError("Impossible de repérer la colonne des comptes.")

    lignes = []
    for row in rows[debut:]:
        if idx["compte"] >= len(row):
            continue
        compte = normaliser_compte(row[idx["compte"]])
        if not compte:
            continue

        def num(i):
            if i is None or i >= len(row):
                return 0.0
            try:
                return float(row[i] or 0)
            except (TypeError, ValueError):
                return 0.0

        lignes.append({
            "compte": compte,
            "libelle": str(row[idx["libelle"]]).strip() if "libelle" in idx
                       and idx["libelle"] < len(row) else "",
            "sd": num(idx.get("sd")),
            "sc": num(idx.get("sc")),
            "md": num(idx.get("md")), "mc": num(idx.get("mc")),
            "od": num(idx.get("od")), "oc": num(idx.get("oc")),
        })
    return lignes, idx


# --------------------------------------------------------------------------
# Anomalies (mêmes règles que la v2 — references/anomalies.md)
# --------------------------------------------------------------------------

def detecter_anomalies(bal, rubs, seuil=1.0):
    a = []
    sd_tot = sum(l["sd"] for l in bal)
    sc_tot = sum(l["sc"] for l in bal)
    if abs(sd_tot - sc_tot) > seuil:
        a.append({"gravite": "BLOQUANT", "compte": "", "libelle": "Balance entière",
                  "probleme": f"Balance déséquilibrée : débit {sd_tot:,.2f} ≠ crédit {sc_tot:,.2f}",
                  "solution": "Reprendre la saisie : les états ne peuvent pas boucler tant que la balance ne boucle pas."})
    elif abs(sd_tot - sc_tot) > 0.005:
        a.append({"gravite": "MINEUR", "compte": "", "libelle": "Balance entière",
                  "probleme": f"Écart d'arrondi de {sd_tot - sc_tot:,.5f}.",
                  "solution": "Sans effet, peut être ignoré."})

    all_exprs = []
    for r in rubs.values():
        all_exprs += [r.brut, r.amort]

    for l in bal:
        c, net = l["compte"], l["sd"] - l["sc"]
        if abs(net) <= seuil:
            continue
        cl = c[0] if c else ""

        if len(c) < 2 or cl not in "123456789":
            a.append({"gravite": "A_TRAITER", "compte": c, "libelle": l["libelle"],
                      "probleme": f"Compte non conforme au plan SYCEBNL (classe '{cl}').",
                      "solution": "Rapprocher du libellé et réaffecter au compte SYCEBNL équivalent avant montage."})
            continue

        if cl == "9":
            a.append({"gravite": "INFO", "compte": c, "libelle": l["libelle"],
                      "probleme": "Compte de classe 9 (contributions volontaires en nature / comptabilité analytique).",
                      "solution": "Hors bilan et hors compte de résultat par construction : repris seulement dans la NOTE 1 (contributions volontaires)."})
            continue

        affecte = any(compte_dans_expr(c, e) for e in all_exprs)
        if not affecte:
            a.append({"gravite": "A_TRAITER", "compte": c, "libelle": l["libelle"],
                      "probleme": "Solde non capté par aucune rubrique (classes 1 à 8). Il fera fuir l'équilibre.",
                      "solution": "Vérifier le préfixe SYCEBNL du numéro, ou compléter la maquette pour ce cas."})

        if cl == "6" and (l["sc"] - l["sd"]) > seuil:
            a.append({"gravite": "A_VERIFIER", "compte": c, "libelle": l["libelle"],
                      "probleme": "Compte de charge (classe 6) au solde créditeur.",
                      "solution": "Contrôler : transfert de charge, RRR obtenus ou erreur d'imputation."})
        if cl == "7" and net > seuil:
            a.append({"gravite": "A_VERIFIER", "compte": c, "libelle": l["libelle"],
                      "probleme": "Compte de produit (classe 7) au solde débiteur.",
                      "solution": "Contrôler : RRR accordés, annulation de produit ou erreur d'imputation."})
        if cl == "2" and c[:2] not in ("28", "29") and (l["sc"] - l["sd"]) > seuil:
            a.append({"gravite": "A_VERIFIER", "compte": c, "libelle": l["libelle"],
                      "probleme": "Immobilisation (classe 2 hors amort/déprec) au solde créditeur.",
                      "solution": "Vérifier : cession non soldée, avoir sur immo ou mauvaise ventilation."})

    cl6_7_8 = any(l["compte"][0] in "678" and abs(l["sd"] - l["sc"]) > seuil for l in bal)
    cl13 = any(l["compte"].startswith("13") and abs(l["sd"] - l["sc"]) > seuil for l in bal)
    if cl6_7_8 and cl13:
        a.append({"gravite": "A_VERIFIER", "compte": "", "libelle": "Résultat",
                  "probleme": "Classes 6/7/8 ouvertes ET compte 13 mouvementé : risque de double résultat.",
                  "solution": "Fournir soit une balance avant clôture (6/7/8 ouverts), soit après (13 seul)."})
    if not cl6_7_8:
        a.append({"gravite": "INFO", "compte": "", "libelle": "Résultat",
                  "probleme": "Classes 6/7/8 non mouvementées : le compte de résultat ressortira à zéro.",
                  "solution": "Pour un compte de résultat renseigné, fournir la balance avant affectation."})
    return a


# --------------------------------------------------------------------------
# Construction des états (formules)
# --------------------------------------------------------------------------

TOTAUX = {
    "AA": ["AB", "AC"],
    "AD": ["AE", "AF", "AG"],
    "AH": ["AI", "AJ", "AK", "AL", "AM", "AN"],
    "AO": ["AX", "AY"],
    "AZ": ["AA", "AD", "AH", "AO"],
    "BT": ["BA", "BB", "BC", "BD", "BE"],
    "BX": ["BU", "BV", "BW"],
    "BZ": ["AZ", "BT", "BX", "BY"],
    "CK": ["CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ"],
    "CY": ["CW", "CX"],
    "CZ": ["CK", "CY"],
    "DD": ["DA", "DB", "DC"],
    "DE": ["CZ", "DD"],
    "DV": ["DF", "DG", "DH", "DI"],
    "DX": ["DW"],
    "DZ": ["DE", "DV", "DX", "DY"],
    "XA": ["RA", "RB", "RC", "RD", "RE", "RF", "RG", "RH"],
    "XB": ["TA", "TB", "TC", "TD", "TE", "TF", "TG", "TH", "TI", "TJ", "TK", "TL"],
    "XC": ["XA", "XB"],
    "XD": ["TM", "TN"],
    "XE": ["XC", "XD"],
}


def _c(f):
    return f[1:] if f and f.startswith("=") else (f or "0")


# Niveaux de présentation des lignes remarquables (charte ETAFI) :
# rubrique = fond jaune pâle, inter = total intermédiaire gris,
# section = total de section vert, general = TOTAL GENERAL bleu nuit.
NIVEAUX_ETAT = {
    "AA": "rubrique", "AD": "rubrique", "AH": "rubrique", "AO": "inter",
    "AZ": "section", "BT": "section", "BX": "section", "BZ": "general",
    "CK": "inter", "CY": "inter", "CZ": "inter", "DD": "inter",
    "DE": "section", "DV": "section", "DX": "section", "DZ": "general",
    "XA": "section", "XB": "section", "XC": "inter", "XD": "inter",
    "XE": "section",
}


def construire_etat(wb, nom, titre, rubs_etat, rubs, avec_n1, ident, actif,
                    page_ref=None, libelle_col=None):
    ws = wb.create_sheet(nom)
    ncols = 7 if actif else 5
    ecrire_cartouche(ws, ident, page_ref or titre, ncols)
    titre_etat(ws, titre, 2, ncols - 1, row=7,
               taille=16 if "BILAN" in titre.upper() else 14)
    # bandeau d'en-têtes sur deux lignes (8 et 9), à la façon du modèle
    lib = libelle_col or ("ACTIF" if actif else "LIBELLES")
    ws.cell(8, 1, "REF")
    ws.cell(8, 2, lib)
    ws.cell(8, 3, "NOTE")
    for c in (1, 2, 3):
        fusion(ws, 8, c, 9, c)
    if actif:
        fusion(ws, 8, 4, 8, 6)
        ws.cell(8, 4, "EXERCICE AU 31/12/N")
        ws.cell(9, 4, "BRUT")
        ws.cell(9, 5, "AMORT et DEPREC.")
        ws.cell(9, 6, "NET")
        ws.cell(8, 7, "EXERCICE AU 31/12/N-1")
        ws.cell(9, 7, "NET")
    else:
        ws.cell(8, 4, "EXERCICE AU 31/12/N")
        ws.cell(9, 4, "NET")
        ws.cell(8, 5, "EXERCICE AU 31/12/N-1")
        ws.cell(9, 5, "NET")
    entetes_bande(ws, 8, 9, 1, ncols)
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 22
    ws.cell(8, 2).font = Font(name="Arial Black", size=11, bold=True)

    r = 9
    debut = r + 1
    cols_montant = tuple(range(4, ncols + 1))
    ref_row = {}
    for ref, rub in rubs_etat:
        r += 1
        ref_row[ref] = r
        ws.cell(r, 1, ref)
        ws.cell(r, 2, rub.libelle)
        note = rub.note.split(";")[0].replace("note officielle", "note").strip() \
            if rub.note else ""
        ws.cell(r, 3, note if len(note) <= 24 else "")
        est_total = bool(rub.formule) or ref in TOTAUX
        if not est_total:
            if actif:
                ws.cell(r, 4).value = formule_expr(rub.brut, "nd", "BALANCE") or 0
                ws.cell(r, 5).value = formule_expr(rub.amort, "nc", "BALANCE") or 0
                ws.cell(r, 6).value = f"=D{r}-E{r}"
                if avec_n1:
                    fb = formule_expr(rub.brut, "nd", "BALANCE_N1")
                    fa = formule_expr(rub.amort, "nc", "BALANCE_N1")
                    ws.cell(r, 7).value = (f"=({_c(fb)})-({_c(fa)})" if (fb or fa)
                                           else 0)
            else:
                ws.cell(r, 4).value = formule_expr(rub.brut, "nc", "BALANCE") or 0
                if avec_n1:
                    ws.cell(r, 5).value = formule_expr(rub.brut, "nc", "BALANCE_N1") or 0
        niveau = NIVEAUX_ETAT.get(ref, "inter" if est_total else "normal")
        style_ligne(ws, r, 1, ncols, niveau, cols_montant=cols_montant,
                    col_ref=1)
        c3 = ws.cell(r, 3)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22

    # formules de totalisation
    for ref, composants in TOTAUX.items():
        if ref not in ref_row:
            continue
        row = ref_row[ref]
        refs_rows = [ref_row[c] for c in composants if c in ref_row]
        if not refs_rows:
            continue
        cols = (4, 5, 6, 7) if actif else ((4, 5) if avec_n1 else (4,))
        for col in cols:
            lettre = chr(64 + col)
            ws.cell(row, col).value = "=" + "+".join(f"{lettre}{rr}" for rr in refs_rows)

    cadre(ws, 8, 1, r, ncols, MOYEN)
    largeurs(ws, {"A": 5.5, "B": 52, "C": 6.5, "D": 15.7, "E": 15.7,
                  "F": 15.7, "G": 15.7})
    ws.freeze_panes = f"A{debut}"
    return ref_row


TFT_LIGNES = [
    ("ZA", "Trésorerie nette au 1er janvier (Trésorerie actif N-1 - Trésorerie passif N-1)", "A"),
    ("__", "Flux de trésorerie provenant des activités opérationnelles", ""),
    ("FA", "+ Encaissement des cotisations", ""),
    ("FB", "+ Encaissement des subventions d'exploitation et d'équilibre", ""),
    ("FC", "+ Encaissement des revenus liés à la générosité", ""),
    ("FD", "+ Encaissement des revenus des manifestations", ""),
    ("FE", "+ Encaissement des autres revenus", ""),
    ("FF", "- Décaissement des sommes versées aux fournisseurs (1)", ""),
    ("FG", "- Décaissement des sommes versées au personnel", ""),
    ("FH", "- Autres décaissements", ""),
    ("ZB", "FLUX DE TRÉSORERIE PROVENANT DES ACTIVITÉS OPÉRATIONNELLES — calculé en résidu (voir README)", "B"),
    ("__", "Flux de trésorerie provenant des activités d'investissement", ""),
    ("FI", "- Décaissements liés aux acquisitions d'immobilisations incorporelles et corporelles", ""),
    ("FJ", "- Décaissements liés aux acquisitions d'immobilisations financières", ""),
    ("FK", "+ Encaissements liés aux cessions d'immobilisations incorporelles et corporelles", ""),
    ("FL", "+ Encaissements liés aux cessions d'immobilisations financières", ""),
    ("ZC", "FLUX DE TRÉSORERIE PROVENANT DES ACTIVITÉS D'INVESTISSEMENT", "C"),
    ("__", "Flux de trésorerie provenant du financement par les fonds propres", ""),
    ("FM", "+ Encaissement des dotations et autres fonds propres", ""),
    ("FN", "+ Subventions d'investissement reçues", ""),
    ("FO", "- Décaissement des dotations et autres fonds propres", ""),
    ("ZD", "FLUX DE TRÉSORERIE PROVENANT DES FONDS PROPRES", "D"),
    ("__", "Trésorerie provenant du financement par les fonds étrangers", ""),
    ("FP", "+ Encaissement provenant des emprunts et des autres dettes financières", ""),
    ("FQ", "- Remboursements des emprunts et autres dettes financières", ""),
    ("ZE", "TRÉSORERIE PROVENANT DES FONDS ÉTRANGERS", "E"),
    ("ZF", "VARIATION DE LA TRÉSORERIE NETTE DE LA PÉRIODE (B+C+D+E)", "G"),
    ("ZG", "TRÉSORERIE NETTE AU 31 DÉCEMBRE (G+A) — contrôle : trésorerie actif N - trésorerie passif N", "H"),
]

IMMOS_BRUT_REFS = ["AB", "AC", "AE", "AF", "AG", "AI", "AJ", "AK", "AL",
                   "AM", "AN", "AX", "AY"]
FONDS_PROPRES_CASH_REFS = ["CA", "CB", "CC", "CD", "CI", "CW", "CX"]
DETTES_FIN_REFS = ["DA", "DB", "DC"]


NIVEAUX_TFT = {"ZA": "cle", "ZB": "section", "ZC": "section",
               "ZD": "inter", "ZE": "inter", "ZF": "cle", "ZG": "cle"}


def construire_tft(wb, rubs, avec_n1, ident):
    ws = wb.create_sheet("TFT")
    ecrire_cartouche(ws, ident, "TABLEAU DES FLUX\nDE TRESORERIE", 5)
    titre_etat(ws, "TABLEAU DES FLUX DE TRESORERIE", 1, 5, row=7, taille=14)
    r = 8
    for i, h in enumerate(["REF", "LIBELLES", "Rep.", "EXERCICE N",
                           "EXERCICE N-1"], start=1):
        ws.cell(r, i, h)
    entetes_bande(ws, r, r, 1, 5)
    ws.row_dimensions[r].height = 22
    ref_row = {}
    debut = r + 1
    for ref, lib, rep in TFT_LIGNES:
        r += 1
        ws.row_dimensions[r].height = 22
        if ref == "__":
            ws.cell(r, 2, lib)
            style_ligne(ws, r, 2, 5, "bande", cols_montant=(4, 5))
            style_ligne(ws, r, 1, 1, "normal")
            continue
        ws.cell(r, 1, ref)
        ws.cell(r, 2, lib)
        ws.cell(r, 3, rep)
        ref_row[ref] = r
        style_ligne(ws, r, 1, 5, NIVEAUX_TFT.get(ref, "normal"),
                    cols_montant=(4, 5), col_ref=1)
        ws.cell(r, 3).alignment = Alignment(horizontal="center",
                                            vertical="center")

    def tresorerie(bal):
        actif = "+".join(
            f"(({_c(formule_expr(rubs[x].brut, 'nd', bal))})-"
            f"({_c(formule_expr(rubs[x].amort, 'nc', bal))}))"
            for x in ("BU", "BV", "BW"))
        passif = _c(formule_expr(rubs["DW"].brut, "nc", bal))
        return f"({actif})-({passif})"

    def somme_brute(refs, bal):
        return "+".join(f"({_c(formule_expr(rubs[x].brut, 'nd', bal))})"
                        for x in refs)

    def somme_passif(refs, bal):
        return "+".join(f"({_c(formule_expr(rubs[x].brut, 'nc', bal))})"
                        for x in refs)

    C = {k: f"D{v}" for k, v in ref_row.items()}
    if avec_n1:
        ws[C["ZA"]] = f"={tresorerie('BALANCE_N1')}"
        ws[C["ZC"]] = (f"=-(({somme_brute(IMMOS_BRUT_REFS, 'BALANCE')})"
                       f"-({somme_brute(IMMOS_BRUT_REFS, 'BALANCE_N1')}))")
        ws[C["ZD"]] = (f"=({somme_passif(FONDS_PROPRES_CASH_REFS, 'BALANCE')})"
                       f"-({somme_passif(FONDS_PROPRES_CASH_REFS, 'BALANCE_N1')})")
        ws[C["ZE"]] = (f"=({somme_passif(DETTES_FIN_REFS, 'BALANCE')})"
                       f"-({somme_passif(DETTES_FIN_REFS, 'BALANCE_N1')})")
        ws[C["ZF"]] = f"=({tresorerie('BALANCE')})-D{ref_row['ZA']}"
        ws[C["ZB"]] = (f"=D{ref_row['ZF']}-D{ref_row['ZC']}"
                       f"-D{ref_row['ZD']}-D{ref_row['ZE']}")
        ws[C["ZG"]] = f"=D{ref_row['ZF']}+D{ref_row['ZA']}"
    else:
        ws.cell(ref_row["ZA"], 4, "Balance N-1 non fournie : TFT non calculé.")
    cadre(ws, 8, 1, r, 5, MOYEN)
    r += 2
    ws.cell(r, 1, "(1) à l'exclusion des fournisseurs d'investissements. "
                  "Les lignes FA à FH (ventilation par nature des encaissements/"
                  "décaissements) se saisissent depuis le journal de trésorerie : "
                  "une balance de clôture ne porte pas cette information.")
    ws.cell(r, 1).font = F_NORMAL
    largeurs(ws, {"A": 5.5, "B": 72, "C": 6, "D": 15.7, "E": 15.7})
    return ref_row


# --------------------------------------------------------------------------
# Feuilles d'audit
# --------------------------------------------------------------------------

def ecrire_balance(wb, nom, bal, rubs=None):
    """Feuille de balance à la présentation attendue (comptes croissants,
    ouverture / mouvements / clôture en débit-crédit, totaux généraux)."""
    return ecrire_feuille_balance(wb, nom, bal)


def construire_garde(wb, ident, avec_n1):
    construire_garde_etafi(
        wb, ident,
        bandeau="ETATS FINANCIERS NORMALISES\nDU SYSTEME COMPTABLE DES "
                "ENTITES A BUT NON LUCRATIF (SYCEBNL)",
        sous_bandeau="Associations, Ordres Professionnels, Fondations "
                     "et Assimilées",
        systeme="SYSTEME NORMAL",
        documents=["Fiche d'identification et renseignements divers",
                   "Bilan (actif et passif)",
                   "Compte de résultat",
                   "Tableau des flux de trésorerie",
                   "Notes annexes"])


def construire_controles(wb, bal, refs, controles_notes, avec_n1):
    ctl = wb.create_sheet("CONTROLES")
    n = max(len(bal), 1)
    ctl.append(["Contrôle", "Valeur", "Attendu"])
    style_entetes(ctl, 1, 1, 3)
    ac, pa, cr = refs["ACTIF"], refs["PASSIF"], refs["CR"]
    A, P, R = q(NOM_ACTIF), q(NOM_PASSIF), q(CR_NOM)
    B = q(NOM_BALANCE)
    lignes = [
        ("Total solde de clôture débit balance",
         f"=SUM({B}!G2:G{n+1})", ""),
        ("Total solde de clôture crédit balance",
         f"=SUM({B}!H2:H{n+1})", ""),
        ("Écart balance (doit être 0)", "=B2-B3", 0),
        ("Total général actif net (BZ)", f"={A}!F{ac['BZ']}", ""),
        ("Total général passif (DZ)", f"={P}!D{pa['DZ']}", ""),
        ("Écart bilan actif - passif (doit être 0)", "=B5-B6", 0),
        ("Résultat net (compte de résultat, XE)", f"={R}!D{cr['XE']}", ""),
        ("Résultat net (bilan, CH = résultat CR + solde compte 13)",
         f"={P}!D{pa['CH']}", ""),
        ("Écart résultat CR / bilan (doit être 0 si balance avant affectation)",
         "=B8-B9", 0),
    ]
    if avec_n1 and refs.get("TFT"):
        tft = refs["TFT"]
        lignes += [
            ("Trésorerie nette au 31/12 (TFT, ZG)", f"=TFT!D{tft['ZG']}", ""),
            ("Trésorerie nette au 31/12 (bilan, BX - DX)",
             f"={A}!F{ac['BX']}-{P}!D{pa['DX']}", ""),
            ("Écart trésorerie TFT / bilan (doit être 0)", "=B11-B12", 0),
        ]
    lignes.append(("- Recoupements notes annexes / états -", "", ""))
    lignes += controles_notes
    for lab, f, att in lignes:
        ctl.append([lab, f, att])
    style_zone_donnees(ctl, 2, ctl.max_row, 1, 3, cols_montant=(2,))
    largeurs(ctl, {"A": 62, "B": 20, "C": 10})


def construire_anomalies(wb, anomalies):
    an = wb.create_sheet("ANOMALIES")
    an.append(["Gravité", "Compte", "Intitulé", "Problème", "Solution proposée"])
    style_entetes(an, 1, 1, 5)
    ordre = {"BLOQUANT": 0, "A_TRAITER": 1, "A_VERIFIER": 2, "MINEUR": 3, "INFO": 4}
    for x in sorted(anomalies, key=lambda z: ordre.get(z["gravite"], 9)):
        an.append([x["gravite"], x["compte"], x["libelle"], x["probleme"], x["solution"]])
    style_zone_donnees(an, 2, max(an.max_row, 2), 1, 5)
    largeurs(an, {"A": 12, "B": 12, "C": 26, "D": 62, "E": 62})
    an.freeze_panes = "A2"


# --------------------------------------------------------------------------
# Programme principal
# --------------------------------------------------------------------------

def main():
    ici = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser()
    ap.add_argument("balance_N")
    ap.add_argument("balance_N1", nargs="?")
    ap.add_argument("--correspondance",
                    default=os.path.join(ici, "..", "references",
                                         "correspondance-associations.tsv"))
    ap.add_argument("--sortie", default="etats-sycebnl.xlsx")
    ap.add_argument("--entite", default="")
    ap.add_argument("--identifiant", default="")
    ap.add_argument("--exercice", default="")
    ap.add_argument("--duree", default="12")
    ap.add_argument("--adresse", default="")
    ap.add_argument("--sigle", default="")
    ap.add_argument("--ntd", default="")
    args = ap.parse_args()
    set_identite_etendue(args.adresse, args.sigle, args.ntd)

    rubs = charger_maquette(args.correspondance)
    bal, idx = lire_balance(args.balance_N)
    print(f"Balance N : {len(bal)} comptes. Colonnes repérées : {idx}")
    bal1 = None
    if args.balance_N1:
        bal1, _ = lire_balance(args.balance_N1)
        print(f"Balance N-1 : {len(bal1)} comptes.")
    avec_n1 = bal1 is not None
    set_lignes_max(max(len(bal), len(bal1 or [])) + 20)

    anomalies = detecter_anomalies(bal, rubs)
    ident = (args.entite, args.identifiant, args.exercice, args.duree)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    actif_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "BILAN-ACTIF"]
    passif_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "BILAN-PASSIF"]
    cr_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "COMPTE-DE-RESULTAT"]

    refs = {"NOM_ACTIF": NOM_ACTIF, "NOM_PASSIF": NOM_PASSIF,
            "CR_NOM": CR_NOM, "NOM_TFT": "TFT"}
    refs["ACTIF"] = construire_etat(
        wb, NOM_ACTIF, "BILAN", actif_rubs, rubs, avec_n1, ident, True,
        page_ref="BILAN SYSTEME NORMAL\nPAGE 1/2", libelle_col="ACTIF")
    refs["PASSIF"] = construire_etat(
        wb, NOM_PASSIF, "BILAN", passif_rubs, rubs, avec_n1, ident, False,
        page_ref="BILAN SYSTEME NORMAL\nPAGE 2/2", libelle_col="PASSIF")
    refs["CR"] = construire_etat(
        wb, CR_NOM, "COMPTE DE RESULTAT", cr_rubs, rubs, avec_n1, ident,
        False, page_ref="COMPTE DE RESULTAT\nSYSTEME NORMAL",
        libelle_col="LIBELLES")
    refs["TFT"] = construire_tft(wb, rubs, avec_n1, ident)

    # Résultat du bilan (CH) = résultat du compte de résultat + solde éventuel
    # du compte 13 : la liasse boucle que la balance soit avant ou après
    # affectation (même principe que le moteur SYSCOHADA).
    P = wb[NOM_PASSIF]
    row_ch, row_xe = refs["PASSIF"]["CH"], refs["CR"]["XE"]
    f13 = _c(formule_tokens(["13"], "nc", "BALANCE"))
    P.cell(row_ch, 4).value = f"={q(CR_NOM)}!D{row_xe}+({f13})"
    if avec_n1:
        f13b = _c(formule_tokens(["13"], "nc", "BALANCE_N1"))
        P.cell(row_ch, 5).value = f"={q(CR_NOM)}!E{row_xe}+({f13b})"

    controles_notes = notes_sycebnl.construire_notes(wb, avec_n1, ident, refs)

    ecrire_balance(wb, "BALANCE", bal, rubs)
    if avec_n1:
        ecrire_balance(wb, "BALANCE_N1", bal1, rubs)
    construire_controle_balance(wb, avec_n1, len(bal), len(bal1 or []))
    construire_controles(wb, bal, refs, controles_notes, avec_n1)
    construire_anomalies(wb, anomalies)
    construire_couverture(wb, ident, "LIASSE SYSTEME NORMAL")
    construire_garde(wb, ident, avec_n1)
    construire_identification(wb, ident, "SYCEBNL",
                              "Associations et ordres professionnels - "
                              "Système normal")
    construire_fiche2(wb, ident, "EQUIPE DE L'ENTITE A BUT NON LUCRATIF")
    ac, pa = refs["ACTIF"], refs["PASSIF"]
    construire_bilan_paysage(
        wb, ident,
        {"feuille": NOM_ACTIF, "lig_debut": min(ac.values()),
         "lig_fin": max(ac.values()), "col_note": "C", "libelle": "ACTIF",
         "cols": [("BRUT", "D"), ("AMORT et DEPREC.", "E"), ("NET", "F"),
                  ("NET N-1", "G")]},
        {"feuille": NOM_PASSIF, "lig_debut": min(pa.values()),
         "lig_fin": max(pa.values()), "col_note": "C", "libelle": "PASSIF",
         "cols": [("NET", "D"), ("NET N-1", "E")]},
        titre="BILAN", page_ref="BILAN SYSTEME NORMAL\nPAGE 1/1")
    parties = notes_sycebnl.parties_depuis_specs(
        notes_sycebnl.NOTES_ASSOCIATIONS,
        [("Partie 1 : Informations générales", 1, 4),
         ("Partie 2 : Notes sur le bilan", 5, 22),
         ("Partie 3 : Notes sur le compte de résultat", 23, 32),
         ("Partie 4 : Autres informations", 33, 35)])
    construire_fiche_notes(wb, parties, ident,
                           "SYCEBNL - Associations et ordres professionnels "
                           "(Partie 4, ch. 2, section 4)")
    construire_table_commentaires(wb, parties, ident)
    ordonner_feuilles(wb, [NOM_BALANCE, NOM_BALANCE_N1, "CONTROLE BALANCE",
                           "Couverture", "Garde", "Fiche 1", "Fiche 2",
                           "Bilan paysage", NOM_ACTIF, NOM_PASSIF,
                           CR_NOM, "TFT", "NOTES ANNEXES"]
                      + [spec["feuille"] for spec in
                         notes_sycebnl.NOTES_ASSOCIATIONS]
                      + ["TABLE COMMENTAIRE", "CONTROLES", "ANOMALIES"])
    retirer_tirets(wb)
    appliquer_police_arial(wb)
    numeroter_pages(wb)

    wb.save(args.sortie)
    print(f"États écrits : {args.sortie}")
    if not avec_n1:
        print("TFT non calculé : fournir balance_N1 pour l'obtenir.")
    bloquants = [a for a in anomalies if a["gravite"] in ("BLOQUANT", "A_TRAITER")]
    print(f"Anomalies : {len(anomalies)} dont {len(bloquants)} à traiter avant remise.")
    for a in anomalies[:12]:
        print(f"  [{a['gravite']}] {a['compte']} {a['probleme']}")


if __name__ == "__main__":
    main()
