#!/usr/bin/env python3
"""
monter_projets.py — Monte le jeu COMPLET d'états financiers SYCEBNL des
PROJETS DE DÉVELOPPEMENT ET ASSIMILÉS (Partie 4, ch. 3 du Journal officiel)
à partir d'une balance générale :

  - Tableau emplois-ressources (FA→GZ) ;
  - Tableau d'exécution budgétaire ;
  - Tableau de réconciliation de trésorerie (A→I) ;
  - Bilan (actif et passif sur feuilles séparées, présentation en net) ;
  - Compte d'exploitation (RA→XC) ;
  - les 24 NOTES ANNEXES officielles (fiche récapitulative, section 6) ;
  - feuilles d'audit GARDE / BALANCE / BALANCE_N1 / CONTROLES / ANOMALIES.

    python monter_projets.py balance_N.xlsx [balance_N-1.xlsx] \
        --sortie etats-projet.xlsx --entite "..." --identifiant "..." \
        --exercice "31/12/N" --duree 12

Chaque montant est une FORMULE Excel (SUMIF sur BALANCE / BALANCE_N1).
La correspondance compte → poste est references/correspondance-projets.tsv
(table officielle, corrections documentées ligne à ligne : présentation en
net, exclusion 479, provisions 499/599, ligne RC réintégrée, dotations 68).

Limites documentées (voir references/notes-projets.md) :
  - Tableau emplois-ressources : les colonnes de l'exercice sont calculées
    (immobilisations et charges depuis la balance ; fonds reçus approchés
    par variation de soldes + quote-part consommée) ; les cumuls de début
    de projet ne vivent pas dans une balance annuelle — colonne « solde
    cumulé début » à compléter, la colonne fin = début + exercice.
  - Tableau d'exécution budgétaire : budget et engagements à saisir
    (nomenclature budgétaire du projet), totaux et % en formules.
"""

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib_mapping_sycebnl import charger_maquette, compte_dans_expr
from formules_sycebnl import (
    formule_tokens, formule_expr, set_lignes_max, q, nom_feuille,
    appliquer_filigranes,
    F_TITRE, F_SOUS_TITRE, F_ENTETE, F_NORMAL, F_GRAS,
    R_TITRE, R_ENTETE, R_BANDE, R_TOTAL, BORD_FIN, AL_CENTRE, AL_GAUCHE,
    FMT_MONTANT, style_entetes, style_zone_donnees, style_ligne_total,
    largeurs, style_titre, retirer_tirets, construire_identification,
    construire_fiche_notes, ordonner_feuilles,
    ecrire_cartouche, titre_etat, entetes_bande, style_ligne, cadre, MOYEN,
    set_identite_etendue, construire_couverture, construire_garde_etafi,
    construire_fiche2, construire_controle_balance,
    construire_table_commentaires, construire_bilan_paysage,
    appliquer_police_arial, numeroter_pages, NOM_BALANCE, NOM_BALANCE_N1,
    anomalies_ouverture,
)
import notes_sycebnl
from notes_projets import NOTES_PROJETS
from monter_etats_sycebnl import (
    lire_balance, detecter_anomalies, ecrire_balance,
    construire_anomalies,
)
from openpyxl.styles import Alignment

NOM_ACTIF = "Bilan-Actif"
NOM_PASSIF = "Bilan-Passif"
CR_NOM = "Compte Exploitation"
PAGE_SYS = "PROJETS DE\nDEVELOPPEMENT"


def entete_pd(ws, titre, page_ref, ident, ncols, taille=14):
    """Cartouche ETAFI + titre : rend la ligne des en-têtes de colonnes."""
    ecrire_cartouche(ws, ident, page_ref, ncols)
    titre_etat(ws, titre, 1, ncols, row=7, taille=taille)
    return 8


def _c(f):
    return f[1:] if f and f.startswith("=") else (f or "0")


def _f(expr, mode, feuille):
    if "!" in expr:
        inc, exc = expr.split("!", 1)
        return formule_tokens([t.strip() for t in inc.split(",") if t.strip()],
                              mode, feuille,
                              exclude=[t.strip() for t in exc.split(",") if t.strip()])
    return formule_tokens([t.strip() for t in expr.split(",") if t.strip()],
                          mode, feuille)


TOTAUX_PD = {
    "AZ": ["AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH"],
    "BF": ["BA", "BB", "BC", "BD", "BE"],
    "BX": ["BV", "BW"],
    "BZ": ["AZ", "BF", "BX", "BY"],
    "CZ": ["CA", "CB", "CC", "CD"],
    "DC": ["DA", "DB"],
    "DD": ["CZ", "DC"],
    "DJ": ["DE", "DF", "DG", "DH", "DI"],
    "DX": ["DW"],
    "DZ": ["DD", "DJ", "DX", "DY"],
    "XA": ["RA", "RB", "RC", "RD", "RE"],
    "XB": ["TA", "TB", "TC", "TD", "TG", "TH", "TI", "TJ", "TK", "TJ2", "TK2", "TL"],
    "XC": ["XA", "XB"],
}

AFFICHAGE_REF = {"TJ2": "TJ", "TK2": "TK"}


NIVEAUX_PD = {
    "AZ": "section", "BF": "section", "BX": "section", "BZ": "general",
    "CZ": "inter", "DC": "inter", "DD": "section", "DJ": "section",
    "DX": "section", "DZ": "general",
    "XA": "section", "XB": "section", "XC": "section",
}


def construire_etat_pd(wb, nom, titre, rubs_etat, avec_n1, ident,
                       page_ref=None, libelle_col="LIBELLES"):
    """États des projets : une colonne de montant (net) par exercice."""
    ws = wb.create_sheet(nom)
    r = entete_pd(ws, titre, page_ref or titre, ident, 5,
                  taille=16 if "BILAN" in titre.upper() else 14)
    ws.cell(r, 1, "REF")
    ws.cell(r, 2, libelle_col)
    ws.cell(r, 3, "NOTE")
    ws.cell(r, 4, "EXERCICE AU 31/12/N")
    ws.cell(r, 5, "EXERCICE AU 31/12/N-1")
    entetes_bande(ws, r, r, 1, 5)
    ws.row_dimensions[r].height = 30
    debut = r + 1
    ref_row = {}
    for ref, rub in rubs_etat:
        r += 1
        ref_row[ref] = r
        ws.cell(r, 1, AFFICHAGE_REF.get(ref, ref))
        ws.cell(r, 2, rub.libelle)
        note = rub.note.split(";")[0].replace("note", "").strip() if rub.note else ""
        ws.cell(r, 3, note if len(note) <= 8 else "")
        est_total = bool(rub.formule) or ref in TOTAUX_PD
        if not est_total:
            if rub.etat == "BILAN-ACTIF":
                fb = formule_expr(rub.brut, "nd", "BALANCE")
                fa = formule_expr(rub.amort, "nc", "BALANCE")
                ws.cell(r, 4).value = (f"=({_c(fb)})-({_c(fa)})" if fa else (fb or 0))
                if avec_n1:
                    fb1 = formule_expr(rub.brut, "nd", "BALANCE_N1")
                    fa1 = formule_expr(rub.amort, "nc", "BALANCE_N1")
                    ws.cell(r, 5).value = (f"=({_c(fb1)})-({_c(fa1)})" if fa1
                                           else (fb1 or 0))
            else:  # PASSIF et COMPTE-DE-RESULTAT : net créditeur signé
                ws.cell(r, 4).value = formule_expr(rub.brut, "nc", "BALANCE") or 0
                if avec_n1:
                    ws.cell(r, 5).value = formule_expr(rub.brut, "nc", "BALANCE_N1") or 0
        niveau = NIVEAUX_PD.get(ref, "inter" if est_total else "normal")
        style_ligne(ws, r, 1, 5, niveau, cols_montant=(4, 5), col_ref=1)
        ws.cell(r, 3).alignment = Alignment(horizontal="center",
                                            vertical="center")
        ws.row_dimensions[r].height = 22

    for ref, composants in TOTAUX_PD.items():
        if ref not in ref_row:
            continue
        row = ref_row[ref]
        rows = [ref_row[c] for c in composants if c in ref_row]
        if not rows:
            continue
        for col in (4, 5) if avec_n1 else (4,):
            lettre = chr(64 + col)
            ws.cell(row, col).value = "=" + "+".join(f"{lettre}{rr}" for rr in rows)
    cadre(ws, debut - 1, 1, r, 5, MOYEN)
    largeurs(ws, {"A": 5.5, "B": 56, "C": 6.5, "D": 15.7, "E": 15.7})
    ws.freeze_panes = f"A{debut}"
    return ref_row


# --------------------------------------------------------------------------
# Tableau emplois-ressources
# --------------------------------------------------------------------------

TER_LIGNES = [
    ("FA", "Fonds reçus, Bailleurs", "162,462", "delta+702"),
    ("FB", "Fonds reçus, Bailleurs (autres)", None, "manuel"),
    ("FC", "Fonds contrepartie État", "163,463", "delta"),
    ("FD", "Autres fonds reçus", "161,164,165,168,464", "delta"),
    ("GR", "I. RESSOURCES", None, "total:FA,FB,FC,FD"),
    ("FE", "Immobilisations incorporelles", "21,251", "immo"),
    ("FF", "Terrains", "22", "immo"),
    ("FG", "Bâtiments", "231,232,233,2391,2392,2393,2396", "immo"),
    ("FH", "Aménagements, agencements et installations",
     "234,235,238,2394,2395,2398", "immo"),
    ("FI", "Matériel, mobilier et actifs biologiques", "24!245,2495", "immo"),
    ("FJ", "Matériel de transport", "245,2495", "immo"),
    ("FK", "Avances et acomptes sur immobilisations", "252", "immo"),
    ("FL", "Immobilisations financières", "26,27", "immo"),
    ("GS", "A - TOTAL DES IMMOBILISATIONS", None,
     "total:FE,FF,FG,FH,FI,FJ,FK,FL"),
    ("FM", "Achats de biens et services", "60", "charge"),
    ("FN", "Transports", "61", "charge"),
    ("FO", "Services extérieurs", "62,63", "charge"),
    ("FP", "Impôts et taxes", "64", "charge"),
    ("FQ", "Autres charges", "65", "charge"),
    ("FR", "Charges de personnel", "66", "charge"),
    ("FS", "Charges financières", "67", "charge"),
    ("FT", "Avances sur charges (à justifier)", "421", "delta_d"),
    ("GT", "B - TOTAL DES CHARGES DE FONCTIONNEMENT", None,
     "total:FM,FN,FO,FP,FQ,FR,FS,FT"),
    ("GU", "II. EMPLOIS (A + B)", None, "total:GS,GT"),
    ("GV", "III. EXCÉDENT / DÉFICIT DES FONDS REÇUS SUR LES EMPLOIS (I - II)",
     None, "diff:GR,GU"),
    ("FU", "Fonds Bailleur en début d'exercice N", None, "manuel"),
    ("FV", "Fonds de contrepartie État en début d'exercice N", None, "manuel"),
    ("FW", "Autres fonds en début d'exercice N", None, "manuel"),
    ("GW", "IV. FONDS DISPONIBLE EN DÉBUT D'EXERCICE", None, "total:FU,FV,FW"),
    ("GX", "V. MONTANT NET DE L'ENCAISSE DISPONIBLE (III + IV)", None,
     "somme:GV,GW"),
    ("FX", "Fonds Bailleur en fin d'exercice N", None, "manuel"),
    ("FY", "Fonds de contrepartie État en fin d'exercice N", None, "manuel"),
    ("FZ", "Autres fonds en fin d'exercice N", None, "manuel"),
    ("GY", "VI. FONDS DISPONIBLE EN FIN D'EXERCICE", None, "total:FX,FY,FZ"),
    ("GZ", "VII. CONTRÔLE : TOTAL V = TOTAL VI (écart)", None, "diff:GX,GY"),
]


NIVEAUX_TER = {"GR": "section", "GS": "inter", "GT": "inter",
               "GU": "section", "GV": "cle", "GW": "section",
               "GX": "cle", "GY": "section", "GZ": "general"}


def construire_ter(wb, avec_n1, ident):
    ws = wb.create_sheet("Emplois-Ressources")
    r = entete_pd(ws, "TABLEAU EMPLOIS-RESSOURCES",
                  f"EMPLOIS-RESSOURCES\n{PAGE_SYS}", ident, 5)
    ws.cell(r, 1, "REF")
    ws.cell(r, 2, "DESIGNATION")
    ws.cell(r, 3, "SOLDE CUMULE DEBUT EXERCICE N")
    ws.cell(r, 4, "EXERCICE N")
    ws.cell(r, 5, "SOLDE CUMULE FIN EXERCICE N")
    entetes_bande(ws, r, r, 1, 5)
    ws.row_dimensions[r].height = 34
    debut = r + 1
    rows = {}
    for ref, lib, expr, mode in TER_LIGNES:
        r += 1
        rows[ref] = r
        ws.cell(r, 1, ref)
        ws.cell(r, 2, lib)
        est_total = ref.startswith("G")
        if mode == "immo":
            # cumul début = brut N-1 ; exercice = variation ; fin = brut N
            if avec_n1:
                ws.cell(r, 3).value = _f(expr, "nd", "BALANCE_N1")
                ws.cell(r, 4).value = f"=E{r}-C{r}"
            ws.cell(r, 5).value = _f(expr, "nd", "BALANCE")
        elif mode == "charge":
            ws.cell(r, 4).value = _f(expr, "nd", "BALANCE")
            ws.cell(r, 5).value = f"=C{r}+D{r}"
        elif mode == "delta":
            if avec_n1:
                fn = _c(_f(expr, "nc", "BALANCE"))
                fn1 = _c(_f(expr, "nc", "BALANCE_N1"))
                ws.cell(r, 4).value = f"=({fn})-({fn1})"
            else:
                ws.cell(r, 4).value = _f(expr, "nc", "BALANCE")
            ws.cell(r, 5).value = f"=C{r}+D{r}"
        elif mode == "delta+702":
            f702 = _c(_f("702", "nc", "BALANCE"))
            if avec_n1:
                fn = _c(_f(expr, "nc", "BALANCE"))
                fn1 = _c(_f(expr, "nc", "BALANCE_N1"))
                ws.cell(r, 4).value = f"=({fn})-({fn1})+({f702})"
            else:
                fn = _c(_f(expr, "nc", "BALANCE"))
                ws.cell(r, 4).value = f"=({fn})+({f702})"
            ws.cell(r, 5).value = f"=C{r}+D{r}"
        elif mode == "delta_d":
            if avec_n1:
                fn = _c(_f(expr, "d", "BALANCE"))
                fn1 = _c(_f(expr, "d", "BALANCE_N1"))
                ws.cell(r, 4).value = f"=({fn})-({fn1})"
            else:
                ws.cell(r, 4).value = _f(expr, "d", "BALANCE")
            ws.cell(r, 5).value = f"=C{r}+D{r}"
        elif mode == "manuel":
            pass
        elif mode.startswith("total:"):
            comp = mode.split(":", 1)[1].split(",")
            for col in "CDE":
                ws[f"{col}{r}"] = "=" + "+".join(f"{col}{rows[c]}" for c in comp)
        elif mode.startswith("somme:"):
            comp = mode.split(":", 1)[1].split(",")
            for col in "CDE":
                ws[f"{col}{r}"] = "=" + "+".join(f"{col}{rows[c]}" for c in comp)
        elif mode.startswith("diff:"):
            a, b = mode.split(":", 1)[1].split(",")
            for col in "CDE":
                ws[f"{col}{r}"] = f"={col}{rows[a]}-{col}{rows[b]}"
        style_ligne(ws, r, 1, 5, NIVEAUX_TER.get(ref, "normal"),
                    cols_montant=(3, 4, 5), col_ref=1)
        ws.row_dimensions[r].height = 22
    cadre(ws, debut - 1, 1, r, 5, MOYEN)
    r += 2
    ws.cell(r, 1, "Colonnes calculées depuis la balance : immobilisations "
                  "(cumul = classe 2 brute), charges de l'exercice (classes "
                  "6), fonds reçus (variation des soldes 16x/46x + quote-part "
                  "702 consommée — approximation à ajuster depuis les avis de "
                  "décaissement des bailleurs). Fonds disponibles début/fin "
                  "(FU à FZ) : à ventiler par source ; rappel balance — "
                  "trésorerie totale : voir CONTROLES.")
    ws.merge_cells(f"A{r}:E{r+2}")
    ws.cell(r, 1).alignment = AL_GAUCHE
    largeurs(ws, {"A": 7, "B": 56, "C": 19, "D": 17, "E": 19})
    return rows


def construire_teb(wb, ident):
    ws = wb.create_sheet("Execution budgetaire")
    r = entete_pd(ws, "TABLEAU DE SUIVI D'EXECUTION DU BUDGET",
                  f"EXECUTION BUDGETAIRE\n{PAGE_SYS}", ident, 8)
    for i, h in enumerate(["Code", "Libellé", "Budget de l'exercice (1)",
                           "Décaissement (2)", "Engagement (3)",
                           "Réalisation (4 = 2 + 3)",
                           "Crédit disponible (5 = 1 - 4)",
                           "Exécution budget % (4/1)"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 8)
    d0 = r + 1
    for _ in range(14):
        r += 1
        ws[f"F{r}"] = f"=D{r}+E{r}"
        ws[f"G{r}"] = f"=C{r}-F{r}"
        ws[f"H{r}"] = f"=IF(C{r}=0,\"\",F{r}/C{r})"
        ws[f"H{r}"].number_format = "0.0%"
    r += 1
    ws.cell(r, 2, "TOTAL")
    for col in "CDEFG":
        ws[f"{col}{r}"] = f"=SUM({col}{d0}:{col}{r-1})"
    ws[f"H{r}"] = f"=IF(C{r}=0,\"\",F{r}/C{r})"
    ws[f"H{r}"].number_format = "0.0%"
    style_zone_donnees(ws, d0, r - 1, 1, 8, cols_montant=(3, 4, 5, 6, 7))
    style_ligne_total(ws, r, 1, 8, cols_montant=(3, 4, 5, 6, 7))
    cadre(ws, d0 - 1, 1, r, 8, MOYEN)
    r += 2
    ws.cell(r, 1, "Remplir code et libellé suivant la nomenclature budgétaire "
                  "du projet.")
    largeurs(ws, {"A": 10, "B": 40, "C": 16, "D": 15, "E": 15, "F": 16,
                  "G": 17, "H": 14})


def construire_reconciliation(wb, avec_n1, ident, ter_rows):
    ws = wb.create_sheet("Reconciliation tresorerie")
    r = entete_pd(ws, "TABLEAU DE RECONCILIATION DE LA TRESORERIE",
                  f"RECONCILIATION\n{PAGE_SYS}", ident, 3)
    ws.cell(r, 1, "LIBELLE")
    ws.cell(r, 2, "REP.")
    ws.cell(r, 3, "MONTANT")
    entetes_bande(ws, r, r, 1, 3)
    ws.row_dimensions[r].height = 22
    debut = r + 1
    tres_n = _c(_f("5!59", "nd", "BALANCE"))
    dep59 = _c(_f("59!599", "c", "BALANCE"))
    lignes = [
        ("TRÉSORERIE EN DÉBUT D'EXERCICE N", "A",
         _f("5!59", "nd", "BALANCE_N1") if avec_n1 else None),
        ("FONDS REÇUS DES BAILLEURS AU COURS DE L'EXERCICE N", "B",
         f"='Emplois-Ressources'!D{ter_rows['FA']}+"
         f"'Emplois-Ressources'!D{ter_rows['FB']}+"
         f"'Emplois-Ressources'!D{ter_rows['FC']}"),
        ("INTÉRÊTS REÇUS AU COURS DE L'EXERCICE N", "C",
         _f("77", "nc", "BALANCE")),
        ("AUTRES FONDS REÇUS AU COURS DE L'EXERCICE N", "D",
         f"='Emplois-Ressources'!D{ter_rows['FD']}"),
        ("VIREMENTS SUR COMPTES OPÉRATIONNELS", "E", None),
        ("DÉPENSES DE L'EXERCICE N", "F",
         f"='Emplois-Ressources'!D{ter_rows['GU']}"),
        ("TRÉSORERIE EN FIN D'EXERCICE N (A+B+C+D-E-F)", "G", "calc"),
        ("PAIEMENTS EN INSTANCE", "H", None),
        ("TRÉSORERIE NETTE DES PAIEMENTS EN INSTANCE (G-H)", "I", "calc2"),
    ]
    rows = {}
    for lib, rep, val in lignes:
        r += 1
        rows[rep] = r
        ws.cell(r, 1, lib)
        ws.cell(r, 2, rep)
        if val == "calc":
            ws.cell(r, 3).value = (f"=C{rows['A']}+C{rows['B']}+C{rows['C']}"
                                   f"+C{rows['D']}-C{rows['E']}-C{rows['F']}")
        elif val == "calc2":
            ws.cell(r, 3).value = f"=C{rows['G']}-C{rows['H']}"
        elif val:
            ws.cell(r, 3).value = val
    for rep, rr in rows.items():
        niveau = {"G": "section", "I": "general"}.get(rep, "normal")
        style_ligne(ws, rr, 1, 3, niveau, cols_montant=(3,))
        ws.row_dimensions[rr].height = 22
    cadre(ws, debut - 1, 1, r, 3, MOYEN)
    r += 2
    ws.cell(r, 1, "Rappel balance — trésorerie de clôture (classe 5 nette) :")
    ws.cell(r, 3).value = f"=({tres_n})-({dep59})"
    ws.cell(r, 3).number_format = FMT_MONTANT
    r += 1
    ws.cell(r, 1, "Écart avec la ligne G (dépenses non décaissées, créances : "
                  "à expliquer en NOTE 2) :")
    ws.cell(r, 3).value = f"=C{rows['G']}-C{r-1}"
    ws.cell(r, 3).number_format = FMT_MONTANT
    largeurs(ws, {"A": 66, "B": 7, "C": 20})
    return rows


def construire_garde_pd(wb, ident, avec_n1):
    construire_garde_etafi(
        wb, ident,
        bandeau="ETATS FINANCIERS NORMALISES\nDU SYSTEME COMPTABLE DES "
                "ENTITES A BUT NON LUCRATIF (SYCEBNL)",
        sous_bandeau="Projets de développement et assimilés",
        systeme="PROJETS DE DEVELOPPEMENT",
        documents=["Fiche d'identification et renseignements divers",
                   "Tableau emplois-ressources",
                   "Tableau d'exécution budgétaire",
                   "Tableau de réconciliation de trésorerie",
                   "Bilan (actif et passif)",
                   "Compte d'exploitation",
                   "Notes annexes"])


def construire_controles_pd(wb, bal, refs, controles_notes, avec_n1):
    ctl = wb.create_sheet("CONTROLES")
    n = max(len(bal), 1)
    ctl.append(["Contrôle", "Valeur", "Attendu"])
    style_entetes(ctl, 1, 1, 3)
    ac, pa, cr = refs["ACTIF"], refs["PASSIF"], refs["CR"]
    A, P, R = q(NOM_ACTIF), q(NOM_PASSIF), q(CR_NOM)
    B = q(NOM_BALANCE)
    lignes = [
        ("Total solde de clôture débit balance",
         f"=SUM({B}!G2:G{n+1})", ""),
        ("Total solde de clôture crédit balance",
         f"=SUM({B}!H2:H{n+1})", ""),
        ("Écart balance (doit être 0)", "=B2-B3", 0),
        ("Total général actif (BZ)", f"={A}!D{ac['BZ']}", ""),
        ("Total général passif (DZ)", f"={P}!D{pa['DZ']}", ""),
        ("Écart bilan actif - passif (doit être 0)", "=B5-B6", 0),
        ("Solde des opérations (compte d'exploitation, XC)",
         f"={R}!D{cr['XC']}", ""),
        ("Solde des opérations (bilan, CC = XC + solde 13)",
         f"={P}!D{pa['CC']}", ""),
        ("- Recoupements notes annexes / états -", "", ""),
    ]
    lignes += controles_notes
    for lab, f, att in lignes:
        ctl.append([lab, f, att])
    style_zone_donnees(ctl, 2, ctl.max_row, 1, 3, cols_montant=(2,))
    largeurs(ctl, {"A": 62, "B": 20, "C": 10})


def main():
    ici = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser()
    ap.add_argument("balance_N")
    ap.add_argument("balance_N1", nargs="?")
    ap.add_argument("--correspondance",
                    default=os.path.join(ici, "..", "references",
                                         "correspondance-projets.tsv"))
    ap.add_argument("--sortie", default="etats-projet.xlsx")
    ap.add_argument("--entite", default="")
    ap.add_argument("--identifiant", default="")
    ap.add_argument("--exercice", default="")
    ap.add_argument("--duree", default="12")
    ap.add_argument("--adresse", default="")
    ap.add_argument("--sigle", default="")
    ap.add_argument("--ntd", default="")
    args = ap.parse_args()
    set_identite_etendue(args.adresse, args.sigle, args.ntd)
    notes_sycebnl.set_suffixe_page("PROJETS DE DEVELOPPEMENT")

    rubs = charger_maquette(args.correspondance)
    bal, idx = lire_balance(args.balance_N)
    print(f"Balance N : {len(bal)} comptes. Colonnes repérées : {idx}")
    bal1 = None
    if args.balance_N1:
        bal1, _ = lire_balance(args.balance_N1)
        print(f"Balance N-1 : {len(bal1)} comptes.")
    avec_n1 = bal1 is not None
    set_lignes_max(max(len(bal), len(bal1 or [])) + 20)

    anomalies = detecter_anomalies(bal, rubs)
    anomalies += anomalies_ouverture(bal)
    ident = (args.entite, args.identifiant, args.exercice, args.duree)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ter_rows = construire_ter(wb, avec_n1, ident)
    construire_teb(wb, ident)
    reco = construire_reconciliation(wb, avec_n1, ident, ter_rows)

    actif_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "BILAN-ACTIF"]
    passif_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "BILAN-PASSIF"]
    cr_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "COMPTE-DE-RESULTAT"]

    refs = {"CR_NOM": CR_NOM, "NOM_ACTIF": NOM_ACTIF,
            "NOM_PASSIF": NOM_PASSIF}
    refs["ACTIF"] = construire_etat_pd(
        wb, NOM_ACTIF, "BILAN (EN NET)", actif_rubs, avec_n1, ident,
        page_ref=f"BILAN PAGE 1/2\n{PAGE_SYS}", libelle_col="ACTIF")
    refs["PASSIF"] = construire_etat_pd(
        wb, NOM_PASSIF, "BILAN (EN NET)", passif_rubs, avec_n1, ident,
        page_ref=f"BILAN PAGE 2/2\n{PAGE_SYS}", libelle_col="PASSIF")
    refs["CR"] = construire_etat_pd(
        wb, CR_NOM, "COMPTE D'EXPLOITATION", cr_rubs, avec_n1, ident,
        page_ref=f"COMPTE D'EXPLOITATION\n{PAGE_SYS}",
        libelle_col="LIBELLES")

    # Solde des opérations au bilan (CC) = solde du compte d'exploitation
    # + solde éventuel du compte 13 : boucle avant/après affectation.
    P = wb[NOM_PASSIF]
    row_cc, row_xc = refs["PASSIF"]["CC"], refs["CR"]["XC"]
    f13 = _c(formule_tokens(["13"], "nc", "BALANCE"))
    P.cell(row_cc, 4).value = f"={q(CR_NOM)}!D{row_xc}+({f13})"
    if avec_n1:
        f13b = _c(formule_tokens(["13"], "nc", "BALANCE_N1"))
        P.cell(row_cc, 5).value = f"={q(CR_NOM)}!E{row_xc}+({f13b})"

    controles_notes = notes_sycebnl.construire_notes(
        wb, avec_n1, ident, refs, notes=NOTES_PROJETS)

    ecrire_balance(wb, "BALANCE", bal, rubs)
    if avec_n1:
        ecrire_balance(wb, "BALANCE_N1", bal1, rubs)
    construire_controle_balance(wb, avec_n1, len(bal), len(bal1 or []))
    construire_controles_pd(wb, bal, refs, controles_notes, avec_n1)
    construire_anomalies(wb, anomalies)
    construire_couverture(wb, ident, "LIASSE PROJETS DE DEVELOPPEMENT")
    construire_garde_pd(wb, ident, avec_n1)
    construire_identification(wb, ident, "SYCEBNL",
                              "Projets de développement et assimilés")
    construire_fiche2(wb, ident, "EQUIPE DU PROJET")
    ac, pa = refs["ACTIF"], refs["PASSIF"]
    construire_bilan_paysage(
        wb, ident,
        {"feuille": NOM_ACTIF, "lig_debut": min(ac.values()),
         "lig_fin": max(ac.values()), "col_note": "C", "libelle": "ACTIF",
         "cols": [("NET", "D"), ("NET N-1", "E")]},
        {"feuille": NOM_PASSIF, "lig_debut": min(pa.values()),
         "lig_fin": max(pa.values()), "col_note": "C", "libelle": "PASSIF",
         "cols": [("NET", "D"), ("NET N-1", "E")]},
        titre="BILAN", page_ref=f"BILAN PAGE 1/1\n{PAGE_SYS}")
    parties = notes_sycebnl.parties_depuis_specs(
        NOTES_PROJETS,
        [("Partie 1 : Informations générales", 1, 1),
         ("Partie 2 : Notes sur le tableau emplois-ressources, le tableau "
          "d'exécution budgétaire et la réconciliation de trésorerie", 2, 2),
         ("Partie 3 : Notes sur le bilan", 3, 13),
         ("Partie 4 : Notes sur le compte d'exploitation", 14, 24)])
    construire_fiche_notes(wb, parties, ident)
    construire_table_commentaires(wb, parties, ident)
    ordonner_feuilles(wb, [NOM_BALANCE, NOM_BALANCE_N1, "CONTROLE BALANCE",
                           "Couverture", "Garde", "Fiche 1", "Fiche 2",
                           "Emplois-Ressources", "Execution budgetaire",
                           "Reconciliation tresorerie", "Bilan paysage",
                           NOM_ACTIF, NOM_PASSIF, CR_NOM, "NOTES ANNEXES"]
                      + [spec["feuille"] for spec in NOTES_PROJETS]
                      + ["TABLE COMMENTAIRE", "CONTROLES", "ANOMALIES"])
    appliquer_filigranes(wb)
    retirer_tirets(wb)
    appliquer_police_arial(wb)
    numeroter_pages(wb)

    wb.save(args.sortie)
    print(f"États écrits : {args.sortie}")
    bloquants = [a for a in anomalies if a["gravite"] in ("BLOQUANT", "A_TRAITER")]
    print(f"Anomalies : {len(anomalies)} dont {len(bloquants)} à traiter avant remise.")
    for a in anomalies[:10]:
        print(f"  [{a['gravite']}] {a['compte']} {a['probleme']}")


if __name__ == "__main__":
    main()
