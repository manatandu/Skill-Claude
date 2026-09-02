#!/usr/bin/env python3
"""
build_gabarit_sycebnl.py — (Re)génère le gabarit Excel VIERGE du jeu d'états
des associations et ordres professionnels (assets/gabarit-etats-associations
.xlsx) : mêmes feuilles, mêmes formules et même présentation (charte ETAFI)
que le classeur produit par monter_etats_sycebnl.py, mais avec des feuilles
BALANCE N / BALANCE N-1 vides — utile comme modèle à consulter ou à remplir
à la main.

    python build_gabarit_sycebnl.py

Il n'existe pas de fichier Excel officiel du SYCEBNL (seuls des tableaux
scannés ont été publiés au Journal officiel) : ce gabarit est une
construction, bâtie strictement sur les libellés et codes REF officiels
(Partie 4, ch. 2) et sur la maquette references/correspondance-associations
.tsv. Relancer ce script écrase le fichier existant.
"""

import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib_mapping_sycebnl import charger_maquette
from formules_sycebnl import (formule_tokens, retirer_tirets, q,
                              construire_identification, construire_fiche2,
                              construire_couverture, construire_fiche_notes,
                              construire_controle_balance,
                              construire_table_commentaires,
                              construire_bilan_paysage, ordonner_feuilles,
                              appliquer_police_arial, numeroter_pages,
                              appliquer_filigranes,
                              NOM_BALANCE, NOM_BALANCE_N1)
import monter_etats_sycebnl as m
import notes_sycebnl

ICI = os.path.dirname(os.path.abspath(__file__))
CORRESPONDANCE = os.path.join(ICI, "..", "references",
                              "correspondance-associations.tsv")
SORTIE = os.path.join(ICI, "..", "assets", "gabarit-etats-associations.xlsx")


def main():
    rubs = charger_maquette(CORRESPONDANCE)
    ident = ("", "", "", "12")
    avec_n1 = True

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    actif_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "BILAN-ACTIF"]
    passif_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "BILAN-PASSIF"]
    cr_rubs = [(ref, r) for ref, r in rubs.items() if r.etat == "COMPTE-DE-RESULTAT"]

    refs = {"NOM_ACTIF": m.NOM_ACTIF, "NOM_PASSIF": m.NOM_PASSIF,
            "CR_NOM": m.CR_NOM, "NOM_TFT": "TFT"}
    refs["ACTIF"] = m.construire_etat(
        wb, m.NOM_ACTIF, "BILAN", actif_rubs, rubs, avec_n1, ident, True,
        page_ref="BILAN SYSTEME NORMAL\nPAGE 1/2", libelle_col="ACTIF")
    refs["PASSIF"] = m.construire_etat(
        wb, m.NOM_PASSIF, "BILAN", passif_rubs, rubs, avec_n1, ident, False,
        page_ref="BILAN SYSTEME NORMAL\nPAGE 2/2", libelle_col="PASSIF")
    refs["CR"] = m.construire_etat(
        wb, m.CR_NOM, "COMPTE DE RESULTAT", cr_rubs, rubs, avec_n1, ident,
        False, page_ref="COMPTE DE RESULTAT\nSYSTEME NORMAL",
        libelle_col="LIBELLES")
    refs["TFT"] = m.construire_tft(wb, rubs, avec_n1, ident)

    P = wb[m.NOM_PASSIF]
    row_ch, row_xe = refs["PASSIF"]["CH"], refs["CR"]["XE"]
    f13 = formule_tokens(["13"], "nc", "BALANCE")[1:]
    P.cell(row_ch, 4).value = f"={q(m.CR_NOM)}!D{row_xe}+({f13})"
    f13b = formule_tokens(["13"], "nc", "BALANCE_N1")[1:]
    P.cell(row_ch, 5).value = f"={q(m.CR_NOM)}!E{row_xe}+({f13b})"

    controles = notes_sycebnl.construire_notes(wb, avec_n1, ident, refs)
    m.ecrire_balance(wb, "BALANCE", [], rubs)
    m.ecrire_balance(wb, "BALANCE_N1", [], rubs)
    construire_controle_balance(wb, avec_n1, 1, 1)
    m.construire_controles(wb, [], refs, controles, avec_n1)
    construire_couverture(wb, ident, "LIASSE SYSTEME NORMAL")
    m.construire_garde(wb, ident, avec_n1)
    construire_identification(wb, ident, "SYCEBNL",
                              "Associations et ordres professionnels - "
                              "Système normal")
    construire_fiche2(wb, ident, "EQUIPE DE L'ENTITE A BUT NON LUCRATIF")
    ac, pa = refs["ACTIF"], refs["PASSIF"]
    construire_bilan_paysage(
        wb, ident,
        {"feuille": m.NOM_ACTIF, "lig_debut": min(ac.values()),
         "lig_fin": max(ac.values()), "col_note": "C", "libelle": "ACTIF",
         "cols": [("BRUT", "D"), ("AMORT et DEPREC.", "E"), ("NET", "F"),
                  ("NET N-1", "G")]},
        {"feuille": m.NOM_PASSIF, "lig_debut": min(pa.values()),
         "lig_fin": max(pa.values()), "col_note": "C", "libelle": "PASSIF",
         "cols": [("NET", "D"), ("NET N-1", "E")]},
        titre="BILAN", page_ref="BILAN SYSTEME NORMAL\nPAGE 1/1")
    parties = notes_sycebnl.parties_depuis_specs(
        notes_sycebnl.NOTES_ASSOCIATIONS,
        [("Partie 1 : Informations générales", 1, 4),
         ("Partie 2 : Notes sur le bilan", 5, 22),
         ("Partie 3 : Notes sur le compte de résultat", 23, 32),
         ("Partie 4 : Autres informations", 33, 35)])
    construire_fiche_notes(wb, parties, ident)
    construire_table_commentaires(wb, parties, ident)
    ordonner_feuilles(wb, [NOM_BALANCE, NOM_BALANCE_N1, "CONTROLE BALANCE",
                           "Couverture", "Garde", "Fiche 1", "Fiche 2",
                           "Bilan paysage", m.NOM_ACTIF, m.NOM_PASSIF,
                           m.CR_NOM, "TFT", "NOTES ANNEXES"]
                      + [spec["feuille"] for spec in
                         notes_sycebnl.NOTES_ASSOCIATIONS]
                      + ["TABLE COMMENTAIRE", "CONTROLES", "ANOMALIES"])
    appliquer_filigranes(wb)
    retirer_tirets(wb)
    appliquer_police_arial(wb)
    numeroter_pages(wb)

    os.makedirs(os.path.dirname(SORTIE), exist_ok=True)
    wb.save(SORTIE)
    print(f"Gabarit vierge écrit : {SORTIE} ({len(wb.sheetnames)} feuilles)")


if __name__ == "__main__":
    main()
