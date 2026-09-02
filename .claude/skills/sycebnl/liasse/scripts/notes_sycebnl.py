#!/usr/bin/env python3
"""
notes_sycebnl.py — Les 35 notes annexes officielles des ASSOCIATIONS ET
ORDRES PROFESSIONNELS (SYCEBNL, Système normal), construites feuille par
feuille dans leur présentation officielle (Partie 4, ch. 2, section 4 du
Journal officiel OHADA du 22/02/2023) et alimentées par FORMULES Excel
(SUMIF sur les feuilles BALANCE / BALANCE_N1) : chaque chiffre se retrace
jusqu'au compte de la balance.

Chaque affectation ligne → comptes a été vérifiée contre le plan des comptes
SYCEBNL (references/partie2-ch2-plan-comptes.md). Les lignes purement
déclaratives (identité, engagements, effectifs...) restent à saisir : le
gabarit est construit, l'en-tête pré-rempli, mais aucun chiffre n'est inventé.

Micro-langage des lignes :
  ("h", libellé)                 : tête de section (gras, sans montants)
  ("d", libellé, expr, mode)     : ligne calculée — expr "31" / "24!245,2495",
                                   mode nd/nc/d/c (cf. formules_sycebnl)
  ("man", libellé)               : ligne à saisir (bordures, pas de formule)
  ("s", libellé[, id])           : sous-total des lignes d/man depuis le
                                   dernier sous-total
  ("x", libellé, gabarit[, id])  : formule libre — {id} y est remplacé par la
                                   référence de cellule de la ligne id, {C}
                                   par la colonne de montants courante
  ("b",)                         : ligne vide
Une ligne "d"/"s" peut porter un identifiant en dernier élément pour être
référencée par une ligne "x" ou par un contrôle.
"""

from formules_sycebnl import (
    formule_tokens, FMT_MONTANT, q,
    F_TITRE, F_SOUS_TITRE, F_ENTETE, F_NORMAL, F_GRAS, F_DONNEE,
    R_TITRE, R_ENTETE, R_BANDE, R_TOTAL, BORD_FIN, AL_CENTRE, AL_GAUCHE,
    style_entetes, style_zone_donnees, style_ligne_total, largeurs, style_titre,
    ecrire_cartouche, titre_note, fusion, C_ENTETE, B_FIN,
)
from openpyxl.styles import PatternFill

# --------------------------------------------------------------------------
# Jeux de colonnes
# --------------------------------------------------------------------------
# n / n1 : colonnes des exercices ; extra : colonnes manuelles en plus.

COLS = {
    "std5": {"entetes": ["Libellés", "Année N", "Année N-1",
                         "Variation en valeur", "Variation en %"],
             "n": "B", "n1": "C", "var": True},
    "ech8": {"entetes": ["Libellés", "Année N", "Année N-1",
                         "Variation en valeur", "Variation en %",
                         "À un an au plus", "À plus d'un an et deux ans au plus",
                         "À plus de deux ans"],
             "n": "B", "n1": "C", "var": True},
    "subv8": {"entetes": ["Libellés", "Année N", "Année N-1",
                          "Variation en valeur", "Variation en %",
                          "Régime fiscal", "Échéances"],
              "n": "B", "n1": "C", "var": True},
}


def _tok(expr):
    if "!" in expr:
        inc, exc = expr.split("!", 1)
        return ([t.strip() for t in inc.split(",") if t.strip()],
                [t.strip() for t in exc.split(",") if t.strip()])
    return ([t.strip() for t in expr.split(",") if t.strip()], [])


def _f(expr, mode, feuille):
    inc, exc = _tok(expr)
    return formule_tokens(inc, mode, feuille, exclude=exc)


# --------------------------------------------------------------------------
# Définition des notes — associations et ordres professionnels
# --------------------------------------------------------------------------

NOTES_ASSOCIATIONS = [
    # ---------------------------------------------------------- NOTE 1
    {"feuille": "NOTE 1", "cols": "std5", "type": "special1",
     "titre": "DETTES GARANTIES PAR DES SÛRETÉS RÉELLES, ENGAGEMENTS "
              "FINANCIERS ET CONTRIBUTIONS VOLONTAIRES EN NATURE"},
    # ---------------------------------------------------------- NOTES 2-4 (déclaratives)
    {"feuille": "NOTE 2", "type": "texte",
     "titre": "INFORMATIONS OBLIGATOIRES",
     "blocs": ["A - IDENTITÉ, ORGANISATION : décrire brièvement l'identité de "
               "l'entité (date de création, nature du projet associatif, mission) "
               "et l'organisation avec des données chiffrées (effectifs, nombre "
               "de bénévoles, total des revenus, moyens utilisés).",
               "B - DÉCLARATION DE CONFORMITÉ AU SYSTÈME COMPTABLE DES ENTITÉS "
               "À BUT NON LUCRATIF ET FAITS MARQUANTS DE L'EXERCICE.",
               "C - RÈGLES, MÉTHODES COMPTABLES ET DÉROGATION AUX PRINCIPES "
               "COMPTABLES.",
               "D - INFORMATIONS COMPLÉMENTAIRES RELATIVES AU BILAN, AU COMPTE "
               "DE RÉSULTAT ET AU TABLEAU DES FLUX DE TRÉSORERIE."]},
    {"feuille": "NOTE 3", "type": "texte",
     "titre": "ÉVÈNEMENTS POSTÉRIEURS À LA CLÔTURE DE L'EXERCICE",
     "blocs": ["Date d'arrêté des états financiers : ",
               "Organe ayant autorisé la publication des comptes : ",
               "A - ÉVÈNEMENTS POSTÉRIEURS DONNANT LIEU À AJUSTEMENTS : nature "
               "des évènements et précisions sur les comptes ajustés.",
               "B - ÉVÈNEMENTS POSTÉRIEURS NE DONNANT PAS LIEU À AJUSTEMENTS : "
               "estimation de l'impact financier ou indication que l'estimation "
               "ne peut être fournie.",
               "C - ÉVÈNEMENTS REMETTANT EN CAUSE LA CONTINUITÉ DE "
               "L'EXPLOITATION : nature de l'évènement, valeurs liquidatives "
               "retenues."]},
    {"feuille": "NOTE 4", "type": "texte",
     "titre": "CHANGEMENTS DE MÉTHODES COMPTABLES, D'ESTIMATIONS ET "
              "CORRECTIONS D'ERREURS",
     "blocs": ["A - CHANGEMENTS DE MÉTHODES COMPTABLES : 1. changement de "
               "réglementation comptable ; 2. changement à l'initiative de "
               "l'entité (impact à l'ouverture, retraitement rétrospectif ou "
               "application prospective).",
               "B - CHANGEMENTS D'ESTIMATIONS : indiquer et justifier le "
               "changement d'estimation.",
               "C - CORRECTIONS D'ERREURS : nature des erreurs corrigées et "
               "principaux postes retraités."]},
    # ---------------------------------------------------------- NOTE 5A-5G (mouvements)
    {"feuille": "NOTE 5A", "type": "mouvements_brut",
     "titre": "DONS ET LEGS D'IMMOBILISATIONS NON REÇUS DESTINÉS À LA VENTE "
              "ET USUFRUIT TEMPORAIRE",
     "lignes": [
         ("h", "IMMOBILISATIONS INCORPORELLES"),
         ("d", "Usufruit", "2011", "nd"),
         ("d", "Brevets, licences, logiciels et droits similaires", "2012,2013", "nd"),
         ("d", "Autres immobilisations incorporelles", "2014,2017", "nd"),
         ("h", "IMMOBILISATIONS CORPORELLES"),
         ("d", "Terrains", "202", "nd"),
         ("d", "Bâtiments", "203", "nd"),
         ("d", "Matériels et mobiliers", "204", "nd"),
         ("h", "IMMOBILISATIONS FINANCIÈRES"),
         ("d", "Titres de participation", "205", "nd"),
     ]},
    {"feuille": "NOTE 5B", "type": "mouvements_brut",
     "titre": "IMMOBILISATIONS BRUTES",
     "lignes": [
         ("h", "IMMOBILISATIONS INCORPORELLES"),
         ("d", "Brevets, licences et droits similaires", "212,214,218", "nd"),
         ("d", "Logiciels et sites internet", "213,2193", "nd"),
         ("d", "Avances et acomptes sur immobilisations incorporelles", "251", "nd"),
         ("d", "Autres immobilisations incorporelles", "2198", "nd"),
         ("h", "IMMOBILISATIONS CORPORELLES"),
         ("d", "Terrains hors immeuble de placement", "22!2281", "nd"),
         ("d", "Terrains - immeuble de placement", "2281", "nd"),
         ("d", "Bâtiments hors immeuble de placement",
          "231,232,233,2391,2392,2393!2315,2325", "nd"),
         ("d", "Bâtiments - immeuble de placement", "2315,2325,2396", "nd"),
         ("d", "Aménagements, agencements et installations",
          "234,235,238,2394,2395,2398", "nd"),
         ("d", "Matériel, mobilier et actifs biologiques", "24!245,2495", "nd"),
         ("d", "Matériel de transport", "245,2495", "nd"),
         ("d", "Avances et acomptes sur immobilisations corporelles", "252", "nd"),
         ("h", "IMMOBILISATIONS FINANCIÈRES"),
         ("d", "Titres de participation", "26", "nd"),
         ("d", "Autres immobilisations financières", "27", "nd"),
     ]},
    {"feuille": "NOTE 5C", "type": "mouvements_brut",
     "titre": "BIENS PRIS EN LOCATION-ACQUISITION",
     "sous_titre": "I : crédit-bail immobilier ; M : crédit-bail mobilier ; "
                   "A : autres contrats",
     "lignes": [
         ("d", "Terrains", "2286", "nd"),
         ("d", "Bâtiments", "2316", "nd"),
         ("d", "Matériel, mobilier", "2416,2426,2446", "nd"),
         ("d", "Matériel de transport", "2456", "nd"),
     ]},
    {"feuille": "NOTE 5D", "type": "mouvements_amort",
     "titre": "DONS ET LEGS D'IMMOBILISATIONS NON REÇUS DESTINÉS À LA VENTE ET "
              "USUFRUIT TEMPORAIRE (AMORTISSEMENTS ET DÉPRÉCIATIONS)",
     "lignes": [
         ("d", "Usufruit", "280,2901", "nc"),
         ("d", "Immobilisations destinées à la vente (dons et legs)", "2902", "nc"),
     ]},
    {"feuille": "NOTE 5E", "type": "mouvements_amort",
     "titre": "IMMOBILISATIONS : AMORTISSEMENTS",
     "lignes": [
         ("h", "IMMOBILISATIONS INCORPORELLES"),
         ("d", "Brevets, licences et droits similaires", "2812,2814,2817", "nc"),
         ("d", "Logiciels et sites internet", "2813", "nc"),
         ("d", "Autres immobilisations incorporelles", "2818", "nc"),
         ("h", "IMMOBILISATIONS CORPORELLES"),
         ("d", "Terrains", "282", "nc"),
         ("d", "Bâtiments et ouvrages d'infrastructure", "2831,2832,2833", "nc"),
         ("d", "Aménagements, agencements et installations", "2834,2835,2838", "nc"),
         ("d", "Matériel, mobilier et actifs biologiques", "284!2845", "nc"),
         ("d", "Matériel de transport", "2845", "nc"),
     ]},
    {"feuille": "NOTE 5F", "type": "mouvements_deprec",
     "titre": "IMMOBILISATIONS : DÉPRÉCIATIONS",
     "lignes": [
         ("h", "IMMOBILISATIONS INCORPORELLES"),
         ("d", "Immobilisations incorporelles", "291", "nc"),
         ("h", "IMMOBILISATIONS CORPORELLES"),
         ("d", "Terrains", "292", "nc"),
         ("d", "Bâtiments et installations", "293", "nc"),
         ("d", "Matériel, mobilier et actifs biologiques", "294!2945", "nc"),
         ("d", "Matériel de transport", "2945", "nc"),
         ("d", "Avances et acomptes versés sur immobilisations", "295", "nc"),
         ("h", "IMMOBILISATIONS FINANCIÈRES"),
         ("d", "Titres de participation", "296", "nc"),
         ("d", "Autres immobilisations financières", "297", "nc"),
     ]},
    {"feuille": "NOTE 5G", "type": "plus_values",
     "titre": "IMMOBILISATIONS : PLUS-VALUES ET MOINS-VALUES DE CESSION",
     "familles": [
         ("Immobilisations incorporelles", "811", "821"),
         ("Immobilisations corporelles", "812", "822"),
         ("Immobilisations financières", "816", "826"),
         ("Immobilisations reçues destinées à la vente (dons et legs)", "818", "828"),
     ]},
    {"feuille": "NOTE 5H", "type": "texte",
     "titre": "INFORMATIONS SUR LES RÉÉVALUATIONS EFFECTUÉES PAR L'ENTITÉ",
     "blocs": ["Nature et date des réévaluations : ",
               "Éléments réévalués par postes du bilan | Montants en coûts "
               "historiques | Montants réévalués | Écarts et provisions "
               "spéciales de réévaluation.",
               "Méthode de réévaluation utilisée : ",
               "Traitement fiscal de l'écart de réévaluation et des "
               "amortissements supplémentaires : ",
               "Montant de l'écart incorporé à la dotation : "]},
    # ---------------------------------------------------------- NOTE 6
    {"feuille": "NOTE 6", "cols": "ech8", "type": "table",
     "titre": "IMMOBILISATIONS FINANCIÈRES",
     "controle": ("ACTIF-net", "AO", "net"),
     "lignes": [
         ("d", "Titres de participation", "26", "nd"),
         ("d", "Prêts et créances", "271,278", "nd"),
         ("d", "Prêt au personnel", "272", "nd"),
         ("d", "Créances sur l'État", "273", "nd"),
         ("d", "Titres immobilisés", "274", "nd"),
         ("d", "Dépôts et cautionnements", "275", "nd"),
         ("d", "Intérêts courus", "276", "nd"),
         ("s", "TOTAL BRUT", "brut"),
         ("d", "Dépréciations des titres de participation", "296", "nc"),
         ("d", "Dépréciations des autres immobilisations financières", "297", "nc"),
         ("s", "", "deprec_hidden"),
         ("x", "TOTAL NET DE DÉPRÉCIATIONS", "={brut}-{deprec_hidden}", "net"),
     ],
     "commentaires": ["Justifier toute variation significative ; commenter les "
                      "créances anciennes ; indiquer le nombre et la date "
                      "d'acquisition des actions ou parts ; pour les "
                      "dépréciations, indiquer les évènements et circonstances."]},
    # ---------------------------------------------------------- NOTE 7
    {"feuille": "NOTE 7", "cols": "std5", "type": "table",
     "titre": "ACTIF CIRCULANT ET DETTES CIRCULANTES HAO",
     "lignes": [
         ("h", "ACTIF CIRCULANT HAO"),
         ("d", "Créances sur cessions d'immobilisations", "485", "nd"),
         ("d", "Créances reçues par dons et legs d'immobilisations", "4865", "nd"),
         ("s", "TOTAL BRUT", "brut"),
         ("d", "Dépréciations des créances HAO", "498", "nc"),
         ("s", "", "dep"),
         ("x", "TOTAL NET DE DÉPRÉCIATIONS", "={brut}-{dep}", "net_actif"),
         ("b",),
         ("h", "DETTES CIRCULANTES HAO"),
         ("d", "Fournisseurs d'investissements", "481", "nc"),
         ("d", "Dettes des dons et legs d'immobilisations", "4861", "nc"),
         ("d", "Créditeurs, dons nature HAO non consommés", "4881", "nc"),
         ("d", "Autres dettes hors activités ordinaires", "484,4998,488!4881", "nc"),
         ("s", "TOTAL DETTES CIRCULANTES HAO", "net_passif"),
     ],
     "commentaires": ["Commenter toute variation significative ; indiquer la "
                      "date et la nature de l'immobilisation achetée et/ou cédée."]},
    # ---------------------------------------------------------- NOTE 8
    {"feuille": "NOTE 8", "cols": "std5", "type": "table",
     "titre": "STOCKS ET ENCOURS",
     "controle": ("ACTIF-net", "BB", "net"),
     "lignes": [
         ("d", "Biens et services liés à l'activité", "31", "nd"),
         ("d", "Marchandises, matières premières", "32", "nd"),
         ("d", "Autres approvisionnements", "33", "nd"),
         ("d", "Dons en nature", "34", "nd"),
         ("d", "Produits finis et en cours", "35,36", "nd"),
         ("d", "Dons en nature HAO", "38", "nd"),
         ("d", "Stocks en cours de route, en consignation ou en dépôt", "37", "nd"),
         ("s", "TOTAL STOCKS ET ENCOURS", "brut"),
         ("d", "Dépréciations des stocks", "39", "nc"),
         ("s", "", "dep"),
         ("x", "TOTAL NET DE DÉPRÉCIATIONS", "={brut}-{dep}", "net"),
     ],
     "commentaires": ["Indiquer la date de prise d'inventaire, la procédure et "
                      "les méthodes d'évaluation ; commenter toute variation "
                      "significative ; détailler les stocks dépréciés."]},
    # ---------------------------------------------------------- NOTE 9
    {"feuille": "NOTE 9", "cols": "ech8", "type": "table",
     "titre": "ADHÉRENTS, CLIENTS-USAGERS",
     "controle": ("ACTIF-net", "BD", "net"),
     "lignes": [
         ("d", "Adhérents", "411", "nd"),
         ("d", "Clients-usagers", "412", "nd"),
         ("d", "Adhérents, clients-usagers, chèques, effets et autres valeurs "
               "impayés", "413", "nd"),
         ("d", "Adhérents, créances litigieuses ou douteuses", "416", "nd"),
         ("d", "Adhérents, clients-usagers, produits à recevoir", "418", "nd"),
         ("s", "TOTAL BRUT ADHÉRENTS, CLIENTS-USAGERS", "brut"),
         ("d", "Dépréciations des comptes adhérents et clients-usagers", "491", "nc"),
         ("s", "", "dep"),
         ("x", "TOTAL NET DE DÉPRÉCIATIONS", "={brut}-{dep}", "net"),
         ("b",),
         ("d", "Adhérents, avances reçues", "4191", "c"),
         ("d", "Clients-usagers, avances et acomptes reçus", "4192", "c"),
         ("d", "Autres clients créditeurs", "419!4191,4192", "c"),
         ("s", "TOTAL CLIENTS CRÉDITEURS", "crediteurs"),
     ],
     "commentaires": ["Commenter toute variation significative et les créances "
                      "anciennes ; indiquer les évènements motivant "
                      "dépréciation et reprise."]},
    # ---------------------------------------------------------- NOTE 10
    {"feuille": "NOTE 10", "cols": "ech8", "type": "table",
     "titre": "AUTRES CRÉANCES",
     "controle": ("ACTIF-net", "BE", "net"),
     "lignes": [
         ("d", "Personnel", "42", "d"),
         ("d", "Organismes sociaux", "43", "d"),
         ("d", "État et collectivités publiques", "44", "d"),
         ("d", "Fondateurs, apporteurs et comptes courants", "45", "d"),
         ("d", "Bailleurs, État et autres organismes, fonds d'administration",
          "46", "d"),
         ("d", "Débiteurs divers", "471,472,473,475", "d"),
         ("d", "Autres débiteurs divers", "474,476,477", "d"),
         ("s", "TOTAL BRUT", "brut"),
         ("d", "Dépréciations des autres créances", "492,493,494,497", "nc"),
         ("s", "", "dep"),
         ("x", "TOTAL NET DE DÉPRÉCIATIONS", "={brut}-{dep}", "net"),
     ],
     "commentaires": ["Justifier toute variation significative ; détailler et "
                      "justifier les créances significatives ou anciennes."]},
    # ---------------------------------------------------------- NOTE 11
    {"feuille": "NOTE 11", "cols": "std5", "type": "table",
     "titre": "TITRES DE PLACEMENT",
     "controle": ("ACTIF-net", "BU", "net"),
     "lignes": [
         ("d", "Titres de trésor et bons de caisse à court terme", "501", "nd"),
         ("d", "Actions", "502", "nd"),
         ("d", "Obligations", "503", "nd"),
         ("d", "Bons de souscription", "504", "nd"),
         ("d", "Titres négociables hors région", "505", "nd"),
         ("d", "Intérêts courus", "506", "nd"),
         ("d", "Autres valeurs assimilées", "508", "nd"),
         ("s", "TOTAL BRUT", "brut"),
         ("d", "Dépréciations des titres de placement", "590", "nc"),
         ("s", "", "dep"),
         ("x", "TOTAL NET DE DÉPRÉCIATIONS", "={brut}-{dep}", "net"),
     ]},
    # ---------------------------------------------------------- NOTE 12
    {"feuille": "NOTE 12", "cols": "std5", "type": "table",
     "titre": "VALEURS À ENCAISSER",
     "controle": ("ACTIF-net", "BV", "net"),
     "lignes": [
         ("d", "Chèques à encaisser", "513", "nd"),
         ("d", "Chèques à l'encaissement", "514", "nd"),
         ("d", "Cartes de crédit à encaisser", "515", "nd"),
         ("d", "Autres valeurs à encaisser", "518", "nd"),
         ("s", "TOTAL BRUT", "brut"),
         ("d", "Dépréciations des valeurs à encaisser", "591", "nc"),
         ("s", "", "dep"),
         ("x", "TOTAL NET DE DÉPRÉCIATIONS", "={brut}-{dep}", "net"),
     ]},
    # ---------------------------------------------------------- NOTE 13
    {"feuille": "NOTE 13", "cols": "std5", "type": "table",
     "titre": "DISPONIBILITÉS",
     "controle": ("ACTIF-net", "BW", "net"),
     "lignes": [
         ("d", "Banques locales", "521", "d"),
         ("d", "Banques autres États région", "522", "d"),
         ("d", "Banques, dépôt à terme et assimilés", "525", "d"),
         ("d", "Autres banques", "523,524", "d"),
         ("d", "Banques intérêts courus", "526", "nd"),
         ("d", "Banques postales", "531", "d"),
         ("d", "Autres établissements financiers", "532,533,538", "d"),
         ("d", "Établissements financiers intérêts courus", "536", "nd"),
         ("d", "Instruments de monnaie électronique", "55", "nd"),
         ("d", "Caisse", "57", "nd"),
         ("s", "TOTAL BRUT", "brut"),
         ("d", "Dépréciations", "592,593,595", "nc"),
         ("s", "", "dep"),
         ("x", "TOTAL NET DE DÉPRÉCIATIONS", "={brut}-{dep}", "net"),
     ],
     "commentaires": ["Indiquer la date de rapprochement des comptes bancaires "
                      "et la date d'inventaire de la caisse et des instruments "
                      "de monnaie électronique. NB : les intérêts courus "
                      "figurent ici en négatif si le compte principal attaché "
                      "est débiteur."]},
    # ---------------------------------------------------------- NOTE 14
    {"feuille": "NOTE 14", "cols": "std5", "type": "table",
     "titre": "ÉCARTS DE CONVERSION",
     "lignes": [
         ("d", "Écarts de conversion-actif (détailler créances et dettes "
               "concernées)", "478", "nd"),
         ("man", "…"),
         ("d", "Écart de conversion-passif (détailler créances et dettes "
               "concernées)", "479", "nc"),
         ("man", "…"),
     ],
     "commentaires": ["Détailler par devise : montant en devises, cours UML à "
                      "l'année d'acquisition, cours UML au 31/12."]},
    # ---------------------------------------------------------- NOTE 15
    {"feuille": "NOTE 15", "cols": "std5", "type": "table",
     "titre": "DOTATION",
     "lignes": [
         ("d", "Dotation non consomptible (sans et avec droit de reprise)",
          "101,102", "nc"),
         ("d", "Droit d'entrée", "103", "nc"),
         ("d", "Dotation consomptible", "104", "nc"),
         ("s", "TOTAL", "net"),
         ("b",),
         ("man", "Détail par membre : Nom et prénoms | Nationalité | Montant | "
                 "Avec ou sans droit de reprise"),
         ("man", "…"), ("man", "…"), ("man", "…"),
     ]},
    # ---------------------------------------------------------- NOTE 16
    {"feuille": "NOTE 16", "cols": "std5", "type": "table",
     "titre": "RÉSERVES",
     "lignes": [
         ("d", "Réserves statutaires ou contractuelles", "112", "nc"),
         ("d", "Autres réserves", "118", "nc"),
         ("s", "TOTAL RÉSERVES", "net"),
         ("d", "Report à nouveau (+ ou -)", "12", "nc"),
     ],
     "commentaires": ["Indiquer la date des délibérations ou dispositions "
                      "statutaires justifiant la variation des réserves et du "
                      "report à nouveau."]},
    # ---------------------------------------------------------- NOTE 17A
    {"feuille": "NOTE 17A", "cols": "subv8", "type": "table",
     "titre": "SUBVENTIONS ET PROVISIONS RÉGLEMENTÉES",
     "lignes": [
         ("d", "État", "1411", "nc"),
         ("d", "Régions", "1412", "nc"),
         ("d", "Départements", "1413", "nc"),
         ("d", "Communes et collectivités publiques décentralisées", "1414", "nc"),
         ("d", "Entités publiques ou mixtes", "1415", "nc"),
         ("d", "Entités et organismes privés", "1416", "nc"),
         ("d", "Organismes internationaux", "1417", "nc"),
         ("d", "Autres subventions d'investissement", "1418,148", "nc"),
         ("s", "TOTAL SUBVENTIONS", "subv"),
         ("d", "Provisions réglementées", "15", "nc"),
         ("s", "", "prov"),
         ("x", "TOTAL SUBVENTIONS ET PROVISIONS RÉGLEMENTÉES",
          "={subv}+{prov}", "net"),
     ],
     "commentaires": ["Indiquer pour chaque subvention la date d'octroi, la "
                      "nature, les obligations éventuelles ; pour les provisions "
                      "réglementées, le texte de référence."]},
    # ---------------------------------------------------------- NOTE 17B
    {"feuille": "NOTE 17B", "cols": "std5", "type": "table",
     "titre": "FONDS AFFECTÉS ET REPORTÉS",
     "lignes": [
         ("d", "Fonds affectés aux investissements", "161,162,163,164", "nc"),
         ("d", "Fonds non consommés en fin d'exercice destinés à un projet "
               "spécifique", "165", "nc"),
         ("d", "Fonds provenant des dons et legs d'immobilisations", "167", "nc"),
         ("d", "Autres fonds affectés", "168,169", "nc"),
         ("s", "TOTAL FONDS AFFECTÉS", "aff"),
         ("d", "Donations et legs non encore reçus d'immobilisations destinés "
               "à la vente", "172", "nc"),
         ("d", "Donation temporaire d'usufruit", "171", "nc"),
         ("d", "Autres fonds reportés", "178", "nc"),
         ("s", "TOTAL FONDS REPORTÉS", "rep"),
         ("x", "TOTAL FONDS AFFECTÉS ET REPORTÉS", "={aff}+{rep}", "net"),
     ],
     "commentaires": ["Indiquer la date d'affectation des fonds et leur mode de "
                      "reprise ; la date des actes de donation et legs, la "
                      "nature et la durée de jouissance de l'usufruit."]},
    # ---------------------------------------------------------- NOTE 18A
    {"feuille": "NOTE 18A", "cols": "ech8", "type": "table",
     "titre": "DETTES FINANCIÈRES ET RESSOURCES ASSIMILÉES",
     "controle": ("PASSIF", "DD", "net"),
     "lignes": [
         ("d", "Emprunts obligataires", "181", "nc"),
         ("d", "Emprunts et dettes auprès des établissements de crédit", "182", "nc"),
         ("d", "Avances reçues de l'État", "183", "nc"),
         ("d", "Dépôts et cautionnements reçus", "185", "nc"),
         ("d", "Intérêts courus", "186", "nc"),
         ("d", "Autres emprunts et dettes", "188", "nc"),
         ("s", "TOTAL EMPRUNTS ET DETTES FINANCIÈRES", "emprunts"),
         ("d", "Crédit-bail immobilier", "1871", "nc"),
         ("d", "Crédit-bail mobilier", "1872", "nc"),
         ("d", "Location-vente", "1873", "nc"),
         ("d", "Intérêts courus (location-acquisition)", "1876", "nc"),
         ("d", "Autres dettes de location-acquisition",
          "187!1871,1872,1873,1876", "nc"),
         ("s", "TOTAL DETTES DE LOCATION-ACQUISITION", "loc"),
         ("d", "Provisions pour litiges", "191", "nc"),
         ("d", "Provisions pour charges sur donations et legs", "192", "nc"),
         ("d", "Provisions pour pertes de change", "194", "nc"),
         ("d", "Provisions pour pensions et obligations similaires", "196", "nc"),
         ("d", "Autres provisions pour risques et charges", "198", "nc"),
         ("s", "TOTAL PROVISIONS FINANCIÈRES POUR RISQUES ET CHARGES", "prov"),
         ("x", "TOTAL DETTES FINANCIÈRES ET RESSOURCES ASSIMILÉES",
          "={emprunts}+{loc}+{prov}", "net"),
     ],
     "commentaires": ["Pour chaque emprunt et dette de location-acquisition : "
                      "date d'octroi, organisme, montant initial, durée, "
                      "garanties données ; pour les pensions, méthode "
                      "d'évaluation et descriptif de la convention."]},
    {"feuille": "NOTE 18B", "type": "texte",
     "titre": "ACTIFS ET PASSIFS ÉVENTUELS",
     "blocs": ["Actif éventuel (litiges, ...) : année N / année N-1.",
               "Passif éventuel (litiges, ...) : année N / année N-1.",
               "Décrire les principales caractéristiques, l'horizon des "
               "encaissements/décaissements attendus et les éventuels "
               "remboursements à percevoir."]},
    # ---------------------------------------------------------- NOTE 19
    {"feuille": "NOTE 19", "cols": "ech8", "type": "table",
     "titre": "FOURNISSEURS D'EXPLOITATION",
     "controle": ("PASSIF", "DH", "net"),
     "lignes": [
         ("d", "Fournisseurs, dettes en compte", "4011,4013", "nc"),
         ("d", "Fournisseurs, réserve de propriété", "4016", "nc"),
         ("d", "Fournisseurs, retenue de garantie", "4017", "nc"),
         ("d", "Fournisseurs, effets à payer", "402", "nc"),
         ("d", "Fournisseurs, factures non parvenues", "408", "nc"),
         ("s", "TOTAL FOURNISSEURS", "net"),
         ("b",),
         ("d", "Fournisseurs, avances et acomptes", "4091", "d"),
         ("d", "Fournisseurs sous-traitants, avances et acomptes", "4093", "d"),
         ("d", "Autres fournisseurs débiteurs", "4094,4098", "d"),
         ("s", "TOTAL FOURNISSEURS DÉBITEURS", "debiteurs"),
     ]},
    # ---------------------------------------------------------- NOTE 20
    {"feuille": "NOTE 20", "cols": "ech8", "type": "table",
     "titre": "DETTES FISCALES ET SOCIALES",
     "lignes": [
         ("d", "Personnel, rémunérations dues", "422", "c"),
         ("d", "Personnel, congés à payer", "4281", "c"),
         ("d", "Charges sociales sur congés à payer", "4382", "c"),
         ("d", "Autres personnel", "421,423,424,425,427,428!4281", "c"),
         ("d", "Caisse de sécurité sociale", "431", "c"),
         ("d", "Caisse de retraite", "432", "c"),
         ("d", "Mutuelle de santé", "4331", "c"),
         ("d", "Assurance retraite", "4332", "c"),
         ("d", "Autres charges sociales à payer et cotisations",
          "4333,438!4382", "c"),
         ("s", "TOTAL DETTES SOCIALES", "soc"),
         ("d", "État, autres impôts et taxes", "442,446", "c"),
         ("d", "État, TVA", "443,444,445", "c"),
         ("d", "État, impôts retenus à la source", "447", "c"),
         ("d", "Autres dettes État", "448,449", "c"),
         ("s", "TOTAL DETTES FISCALES", "fisc"),
         ("x", "TOTAL DETTES SOCIALES ET FISCALES", "={soc}+{fisc}", "net"),
     ],
     "commentaires": ["Commenter toute variation significative et les dettes "
                      "anciennes."]},
    # ---------------------------------------------------------- NOTE 21
    {"feuille": "NOTE 21", "cols": "ech8", "type": "table",
     "titre": "AUTRES DETTES ET PROVISIONS POUR RISQUES ET CHARGES À COURT TERME",
     "lignes": [
         ("h", "Fonds d'administration des projets"),
         ("d", "Bailleurs de fonds", "462", "c"),
         ("d", "État", "463", "c"),
         ("d", "Autres organismes de financement assimilés", "464", "c"),
         ("s", "TOTAL BAILLEURS, FONDS D'ADMINISTRATION", "bailleurs"),
         ("d", "Créditeurs divers", "4711,4712,4717,4719", "c"),
         ("d", "Créditeurs, dons en nature courants non consommés", "4713", "c"),
         ("d", "Versements restant à effectuer sur titres de placement non "
               "libérés", "4726", "c"),
         ("d", "Générosités financières à recevoir", "475", "c"),
         ("d", "Autres créditeurs divers",
          "45,4721,473,474,476,477", "c"),
         ("s", "TOTAL CRÉDITEURS DIVERS", "crediteurs"),
         ("x", "TOTAL AUTRES DETTES", "={bailleurs}+{crediteurs}", "net"),
         ("b",),
         ("d", "Provisions pour risques et charges à court terme (voir note 30)",
          "499!4998,599", "nc"),
     ],
     "commentaires": ["Commenter toute variation significative et les dettes "
                      "anciennes."]},
    # ---------------------------------------------------------- NOTE 22
    {"feuille": "NOTE 22", "cols": "std5", "type": "table",
     "titre": "BANQUES, CRÉDIT D'ESCOMPTE ET DE TRÉSORERIE",
     "controle": ("PASSIF", "DX", "net"),
     "lignes": [
         ("d", "Banques locales", "521", "c"),
         ("d", "Banques autres États région", "522", "c"),
         ("d", "Autres banques", "523,524,525", "c"),
         ("d", "Banques, intérêts courus", "526,566", "c"),
         ("d", "Crédit de trésorerie et d'escompte", "561,565", "c"),
         ("s", "TOTAL : BANQUES, CRÉDITS DE TRÉSORERIE", "net"),
     ],
     "commentaires": ["Indiquer l'organisme, les conditions de crédit, le taux "
                      "d'intérêt, la durée. NB : « Banques, intérêts courus » "
                      "figure ici si le compte principal attaché est créditeur."]},
    # ---------------------------------------------------------- NOTE 23
    {"feuille": "NOTE 23", "cols": "std5", "type": "table",
     "titre": "REVENUS ET AUTRES PRODUITS",
     "lignes": [
         ("d", "Cotisations des adhérents", "701", "nc"),
         ("d", "Quote-part de dotation consomptible transférée", "703", "nc"),
         ("d", "Revenus liés à la générosité", "704", "nc"),
         ("d", "Ventes de marchandises, services et produits finis", "705", "nc"),
         ("d", "Revenus des manifestations", "706", "nc"),
         ("d", "Autres revenus", "702,707,708", "nc"),
         ("s", "TOTAL : REVENUS", "revenus"),
         ("d", "Subventions d'exploitation", "71", "nc"),
         ("d", "Autres produits et transferts de charges d'exploitation",
          "72,73,75,77,78", "nc"),
         ("s", "TOTAL : SUBVENTIONS D'EXPLOITATION ET AUTRES PRODUITS", "subv"),
         ("x", "TOTAL", "={revenus}+{subv}", "net"),
     ],
     "commentaires": ["Justifier toute variation significative ; détailler les "
                      "revenus liés à la générosité (dons, legs, deniers du "
                      "culte, zakat, célébrations, mécénat, parrainage)."]},
    # ---------------------------------------------------------- NOTE 24
    {"feuille": "NOTE 24", "cols": "std5", "type": "table",
     "titre": "ACHATS",
     "lignes": [
         ("d", "Achats de biens et services liés à l'activité", "601", "nd"),
         ("s", "TOTAL : ACHATS DE BIENS ET SERVICES LIÉS À L'ACTIVITÉ", "bs"),
         ("d", "Achats de marchandises et matières premières", "602", "nd"),
         ("s", "TOTAL : ACHATS MARCHANDISES ET MATIÈRES PREMIÈRES", "mp"),
         ("d", "Matières et fournitures consommables", "604", "nd"),
         ("d", "Eau, électricité, énergies et fournitures non stockées", "605", "nd"),
         ("d", "Achats autres activités (billetteries, tombolas, voyages...)",
          "606", "nd"),
         ("d", "Achats d'emballages", "608", "nd"),
         ("d", "Rabais, remises et ristournes obtenus (-)", "609", "nd"),
         ("s", "TOTAL AUTRES ACHATS", "autres"),
         ("x", "TOTAL ACHATS", "={bs}+{mp}+{autres}", "net"),
     ]},
    # ---------------------------------------------------------- NOTE 25
    {"feuille": "NOTE 25", "cols": "std5", "type": "table",
     "titre": "TRANSPORTS",
     "controle": ("CR", "TF", "net"),
     "lignes": [
         ("d", "Transports sur ventes", "612", "nd"),
         ("d", "Transports pour le compte de tiers", "613", "nd"),
         ("d", "Transport du personnel", "614", "nd"),
         ("d", "Transports de plis", "616", "nd"),
         ("d", "Voyages, déplacements et autres frais de transport", "618,619", "nd"),
         ("s", "TOTAL", "net"),
     ]},
    # ---------------------------------------------------------- NOTE 26
    {"feuille": "NOTE 26", "cols": "std5", "type": "table",
     "titre": "SERVICES EXTÉRIEURS",
     "controle": ("CR", "TG", "net"),
     "lignes": [
         ("d", "Sous-traitance générale", "621", "nd"),
         ("d", "Locations et charges locatives", "622", "nd"),
         ("d", "Redevances de location-acquisition", "623", "nd"),
         ("d", "Entretien, réparations et maintenance", "624", "nd"),
         ("d", "Primes d'assurance", "625", "nd"),
         ("d", "Études, recherches et documentation", "626", "nd"),
         ("d", "Publicité, publications, relations publiques", "627", "nd"),
         ("d", "Frais de télécommunications", "628", "nd"),
         ("d", "Frais bancaires", "631", "nd"),
         ("d", "Rémunérations d'intermédiaires et de conseils", "632", "nd"),
         ("d", "Frais de formation du personnel", "633", "nd"),
         ("d", "Redevances pour brevets, licences, logiciels et droits "
               "similaires", "634", "nd"),
         ("d", "Cotisations", "635", "nd"),
         ("d", "Frais de recherche de fonds", "636", "nd"),
         ("d", "Rémunérations de personnel extérieur à l'entité", "637", "nd"),
         ("d", "Autres charges externes", "638", "nd"),
         ("s", "TOTAL", "net"),
     ]},
    # ---------------------------------------------------------- NOTE 27
    {"feuille": "NOTE 27", "cols": "std5", "type": "table",
     "titre": "IMPÔTS ET TAXES",
     "controle": ("CR", "TH", "net"),
     "lignes": [
         ("d", "Impôts et taxes directs", "641", "nd"),
         ("d", "Impôts et taxes indirects", "645", "nd"),
         ("d", "Droits d'enregistrement", "646", "nd"),
         ("d", "Pénalités et amendes fiscales", "647", "nd"),
         ("d", "Autres impôts et taxes", "648", "nd"),
         ("d", "Dégrèvements et annulations des impôts et taxes (-)", "649", "nd"),
         ("s", "TOTAL", "net"),
     ],
     "commentaires": ["Détailler pénalités et amendes et en indiquer la cause."]},
    # ---------------------------------------------------------- NOTE 28
    {"feuille": "NOTE 28", "cols": "std5", "type": "table",
     "titre": "AUTRES CHARGES",
     "controle": ("CR", "TI", "net"),
     "lignes": [
         ("d", "Pertes sur créances adhérents, clients et autres débiteurs",
          "651", "nd"),
         ("d", "Subventions versées par l'entité", "652", "nd"),
         ("d", "Dons en nature courants à distribuer", "654", "nd"),
         ("d", "Pénalités et amendes pénales", "657", "nd"),
         ("d", "Autres charges diverses", "658", "nd"),
         ("d", "Charges pour dépréciations et provisions pour risques à court "
               "terme d'exploitation (voir note 30)", "659", "nd"),
         ("s", "TOTAL", "net"),
     ],
     "commentaires": ["Indiquer les bénéficiaires des subventions versées."]},
    # ---------------------------------------------------------- NOTE 29A
    {"feuille": "NOTE 29A", "cols": "std5", "type": "table",
     "titre": "CHARGES DE PERSONNEL",
     "controle": ("CR", "TJ", "net"),
     "lignes": [
         ("d", "Rémunérations directes versées au personnel national", "661", "nd"),
         ("d", "Rémunérations directes versées au personnel non national",
          "662", "nd"),
         ("d", "Indemnités forfaitaires versées au personnel", "663", "nd"),
         ("d", "Charges sociales", "664", "nd"),
         ("d", "Habillement et équipement du personnel", "665", "nd"),
         ("d", "Rémunération transférée de personnel extérieur", "667", "nd"),
         ("d", "Autres charges sociales", "668", "nd"),
         ("d", "Dégrèvements et annulations des charges sociales (-)", "669", "nd"),
         ("s", "TOTAL", "net"),
     ],
     "commentaires": ["Indiquer la nature et la durée du contrat du personnel "
                      "extérieur."]},
    {"feuille": "NOTE 29B", "type": "effectifs",
     "titre": "EFFECTIFS, MASSE SALARIALE ET PERSONNEL EXTÉRIEUR"},
    # ---------------------------------------------------------- NOTE 30
    {"feuille": "NOTE 30", "type": "mouvements_prov",
     "titre": "DOTATIONS ET CHARGES POUR PROVISIONS ET DÉPRÉCIATIONS",
     "lignes": [
         ("d", "Provisions réglementées", "15", "nc", "H"),
         ("d", "Provisions pour risques et charges", "19", "nc", "F"),
         ("d", "Dépréciations des immobilisations", "29!290", "nc", "E"),
         ("d", "Dépréciations des dons et legs, usufruit et immobilisations "
               "reçues destinées à la vente", "290", "nc", "E"),
         ("s", "TOTAL : DOTATIONS", None),
         ("d", "Dépréciations des stocks et en-cours", "39", "nc", "E"),
         ("d", "Dépréciations des comptes fournisseurs", "490", "nc", "E"),
         ("d", "Dépréciations des comptes adhérents et clients", "491", "nc", "E"),
         ("d", "Dépréciations autres créances d'exploitation", "492,493,494,497",
          "nc", "E"),
         ("d", "Dépréciations des comptes de créances HAO", "498", "nc", "H"),
         ("d", "Dépréciations des titres de placement", "590", "nc", "F"),
         ("d", "Dépréciations des titres et valeurs à encaisser", "591", "nc", "F"),
         ("d", "Dépréciations des comptes banques", "592", "nc", "F"),
         ("d", "Dépréciations des comptes établissements financiers", "593",
          "nc", "F"),
         ("d", "Dépréciations des instruments de monnaie électronique", "595",
          "nc", "F"),
         ("d", "Provisions pour risques à court terme d'exploitation", "4991",
          "nc", "E"),
         ("d", "Provisions pour risques à court terme HAO", "4998", "nc", "H"),
         ("d", "Provisions pour risques à court terme à caractère financier",
          "599", "nc", "F"),
         ("s", "TOTAL : CHARGES POUR DÉPRÉCIATIONS ET PROVISIONS À COURT TERME",
          None),
     ]},
    # ---------------------------------------------------------- NOTE 31
    {"feuille": "NOTE 31", "cols": "std5", "type": "table",
     "titre": "CHARGES ET REVENUS FINANCIERS",
     "controle": ("CR", "TK", "net_frais"),
     "lignes": [
         ("d", "Intérêts des emprunts", "671", "nd"),
         ("d", "Intérêts dans loyers de location-acquisition", "672", "nd"),
         ("d", "Escomptes accordés", "673", "nd"),
         ("d", "Autres intérêts", "674", "nd"),
         ("d", "Pertes de change financières", "676", "nd"),
         ("d", "Pertes sur cessions de titres de placement", "677", "nd"),
         ("d", "Pertes et charges sur risques financiers", "678", "nd"),
         ("d", "Charges pour dépréciations et provisions à court terme à "
               "caractère financier (voir note 30)", "679", "nd"),
         ("s", "TOTAL : FRAIS FINANCIERS", "net_frais"),
         ("d", "Intérêts de prêts et créances diverses", "771", "nc"),
         ("d", "Revenus de participations et autres titres immobilisés", "772", "nc"),
         ("d", "Escomptes obtenus", "773", "nc"),
         ("d", "Revenus de placement", "774", "nc"),
         ("d", "Gains de change financiers", "776", "nc"),
         ("d", "Gains sur cessions de titres de placement", "777", "nc"),
         ("d", "Gains sur risques financiers", "778", "nc"),
         ("d", "Transferts de charges financières", "787", "nc"),
         ("d", "Reprises de charges pour dépréciations et provisions à court "
               "terme à caractère financier (voir note 30)", "779", "nc"),
         ("s", "TOTAL : REVENUS FINANCIERS", "revenus"),
         ("x", "TOTAL (revenus - frais)", "={revenus}-{net_frais}", "net"),
     ]},
    # ---------------------------------------------------------- NOTE 32
    {"feuille": "NOTE 32", "cols": "std5", "type": "table",
     "titre": "AUTRES CHARGES ET PRODUITS HAO",
     "lignes": [
         ("d", "Charges H.A.O. constatées (compte 831)", "831", "nd"),
         ("d", "Dons en nature H.A.O. à distribuer (compte 832)", "832", "nd"),
         ("d", "Pertes sur créances HAO", "834", "nd"),
         ("d", "Abandons de créances consentis", "836", "nd"),
         ("d", "Transferts de charges HAO (-)", "838", "nd"),
         ("d", "Charges pour dépréciations et provisions pour risques à court "
               "terme HAO", "839", "nd"),
         ("d", "Dotations hors activités ordinaires", "85", "nd"),
         ("d", "Valeurs comptables des cessions d'immobilisations", "81", "nd"),
         ("d", "Variations de stocks de dons en nature HAO (+/-)", "87", "nd"),
         ("s", "TOTAL : AUTRES CHARGES HAO", "charges"),
         ("d", "Produits H.A.O. constatés (compte 841)", "841", "nc"),
         ("d", "Contributions volontaires en nature (compte 842)", "842", "nc"),
         ("d", "Contributions volontaires en numéraire", "843", "nc"),
         ("d", "Abandons de créances obtenus", "846", "nc"),
         ("d", "Transferts de produits H.A.O.", "848", "nc"),
         ("d", "Reprises des charges pour dépréciations et provisions à court "
               "terme HAO", "849", "nc"),
         ("d", "Reprises d'amortissements, provisions et dépréciations H.A.O.",
          "86", "nc"),
         ("d", "Produits des cessions d'immobilisations", "82", "nc"),
         ("d", "Subventions d'équilibre", "88", "nc"),
         ("s", "TOTAL : AUTRES PRODUITS HAO", "produits"),
         ("x", "TOTAL (produits - charges) = RÉSULTAT HAO",
          "={produits}-{charges}", "net"),
     ]},
    # ---------------------------------------------------------- NOTES 33-35
    {"feuille": "NOTE 33", "type": "synthese",
     "titre": "FICHE DE SYNTHÈSE DES PRINCIPAUX INDICATEURS FINANCIERS"},
    {"feuille": "NOTE 34", "type": "texte",
     "titre": "LISTE DES INFORMATIONS SOCIALES, ENVIRONNEMENTALES ET SOCIÉTALES",
     "blocs": ["(Note obligatoire pour les entités de plus de 250 personnes, "
               "bénévoles compris.)",
               "INFORMATIONS SOCIALES : emploi (effectif total, répartition par "
               "sexe, âge et zone géographique ; embauches et licenciements ; "
               "rémunérations et leur évolution) ; relations sociales ; santé "
               "et sécurité ; formation ; égalité de traitement.",
               "INFORMATIONS ENVIRONNEMENTALES : politique générale ; pollution "
               "et gestion des déchets ; utilisation durable des ressources ; "
               "changement climatique ; protection de la biodiversité.",
               "ENGAGEMENTS SOCIÉTAUX EN FAVEUR DU DÉVELOPPEMENT DURABLE : "
               "impact territorial, économique et social ; relations avec les "
               "parties prenantes ; sous-traitance et fournisseurs."]},
    {"feuille": "NOTE 35", "type": "budget",
     "titre": "TABLEAU D'EXÉCUTION BUDGÉTAIRE"},
]


# ==========================================================================
# MOTEUR DE RENDU
# ==========================================================================

# Mention portée sous le numéro de note dans la réf. de page (en haut à
# droite) : « SYSTEME NORMAL » pour les associations, ajustée par les autres
# moteurs (projets, SMT) via set_suffixe_page.
SUFFIXE_PAGE = "SYSTEME NORMAL"


def set_suffixe_page(s):
    global SUFFIXE_PAGE
    SUFFIXE_PAGE = s


def entete_note(ws, titre, ident, ncols, sous_titre=""):
    """Cartouche ETAFI + titre de note (« NOTE X : LIBELLE » en Arial Black
    bleu nuit). Rend la ligne où écrire les en-têtes de colonnes."""
    ecrire_cartouche(ws, ident, f"{ws.title}\n{SUFFIXE_PAGE}", max(ncols, 5))
    titre_note(ws, f"{ws.title} : {titre}", max(ncols, 5), row=7)
    r = 8
    if sous_titre:
        fusion(ws, r, 1, r, max(ncols, 5))
        c = ws.cell(r, 1, sous_titre)
        c.font = F_DONNEE
        c.alignment = AL_CENTRE
        r += 1
    return r


def bande_note(ws, r, ncols, texte):
    """Bande de section de note : fond CCFFFF, gras, bordée (modèle)."""
    ws.cell(r, 1, texte)
    for c in range(1, ncols + 1):
        cell = ws.cell(r, c)
        cell.fill = PatternFill("solid", fgColor=C_ENTETE)
        cell.font = F_SOUS_TITRE
        cell.border = B_FIN
    ws.cell(r, 1).alignment = AL_GAUCHE


def _commentaires(ws, r, textes, ncols):
    r += 1
    ws.cell(r, 1, "Commentaires :")
    ws.cell(r, 1).font = F_GRAS
    for t in textes:
        r += 1
        ws.cell(r, 1, "• " + t)
        ws.cell(r, 1).alignment = AL_GAUCHE
    return r


def _render_table(ws, spec, avec_n1, ident):
    cols = COLS[spec.get("cols", "std5")]
    entetes = cols["entetes"]
    n_col, n1_col = cols["n"], cols["n1"]
    r = entete_note(ws, spec["titre"], ident, len(entetes))
    for i, h in enumerate(entetes, start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, len(entetes))
    debut_donnees = r + 1
    bloc_debut = r + 1
    ids = {}
    lignes_num = []          # lignes portant des montants (pour bordures)

    for ligne in spec["lignes"]:
        kind = ligne[0]
        r += 1
        if kind == "b":
            bloc_debut = r + 1
            continue
        if kind == "h":
            bande_note(ws, r, len(entetes), ligne[1])
            bloc_debut = r + 1
            continue
        if kind == "man":
            ws.cell(r, 1, ligne[1])
            lignes_num.append(r)
            continue
        if kind == "d":
            _, label, expr, mode = ligne[:4]
            ws.cell(r, 1, label)
            ws[f"{n_col}{r}"] = _f(expr, mode, "BALANCE")
            if avec_n1:
                ws[f"{n1_col}{r}"] = _f(expr, mode, "BALANCE_N1")
            lignes_num.append(r)
            if len(ligne) > 4:
                ids[ligne[4]] = r
            continue
        if kind == "s":
            label = ligne[1]
            ws.cell(r, 1, label)
            for col in (n_col, n1_col) if avec_n1 else (n_col,):
                ws[f"{col}{r}"] = f"=SUM({col}{bloc_debut}:{col}{r-1})"
            if len(ligne) > 2 and ligne[2]:
                ids[ligne[2]] = r
            style_ligne_total(ws, r, 1, len(entetes),
                              cols_montant=tuple(range(2, len(entetes) + 1)))
            lignes_num.append(r)
            bloc_debut = r + 1
            continue
        if kind == "x":
            _, label, gabarit = ligne[:3]
            ws.cell(r, 1, label)
            for col in (n_col, n1_col) if avec_n1 else (n_col,):
                f = gabarit
                for k, rr in ids.items():
                    f = f.replace("{" + k + "}", f"{col}{rr}")
                ws[f"{col}{r}"] = f
            if len(ligne) > 3:
                ids[ligne[3]] = r
            style_ligne_total(ws, r, 1, len(entetes),
                              cols_montant=tuple(range(2, len(entetes) + 1)))
            lignes_num.append(r)
            bloc_debut = r + 1
            continue

    # colonnes de variation
    if cols.get("var") and avec_n1:
        for rr in lignes_num:
            if ws[f"{n_col}{rr}"].value is not None:
                ws[f"D{rr}"] = f"=B{rr}-C{rr}"
                ws[f"E{rr}"] = f"=IF(C{rr}=0,\"\",(B{rr}-C{rr})/C{rr})"
                ws[f"E{rr}"].number_format = "0.0%"
    style_zone_donnees(ws, debut_donnees, r, 1, len(entetes),
                       cols_montant=(2, 3, 4))
    if spec.get("commentaires"):
        r = _commentaires(ws, r + 1, spec["commentaires"], len(entetes))
    largeurs(ws, {"A": 58, "B": 16, "C": 16, "D": 16, "E": 11,
                  "F": 14, "G": 16, "H": 14})
    return ids


def _render_mouvements(ws, spec, avec_n1, ident, mode_defaut, col_titres):
    """Notes 5A/5B/5C (brut) et 5D/5E/5F (amortissements, dépréciations)."""
    r = entete_note(ws, spec["titre"], ident, 5, spec.get("sous_titre", ""))
    for i, h in enumerate(col_titres, start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 5)
    debut = r + 1
    data_rows = []
    for ligne in spec["lignes"]:
        r += 1
        if ligne[0] == "h":
            bande_note(ws, r, 5, ligne[1])
            continue
        _, label, expr, mode = ligne[:4]
        ws.cell(r, 1, label)
        if avec_n1:
            ws[f"B{r}"] = _f(expr, mode, "BALANCE_N1")
            fn = _f(expr, mode, "BALANCE")[1:]
            fn1 = _f(expr, mode, "BALANCE_N1")[1:]
            ws[f"C{r}"] = f"=MAX(0,({fn})-({fn1}))"
            ws[f"D{r}"] = f"=MAX(0,({fn1})-({fn}))"
            ws[f"E{r}"] = f"=B{r}+C{r}-D{r}"
        else:
            ws[f"E{r}"] = _f(expr, mode, "BALANCE")
        data_rows.append(r)
    r += 1
    ws.cell(r, 1, "TOTAL GÉNÉRAL")
    for col in "BCDE":
        ws[f"{col}{r}"] = "=" + "+".join(f"{col}{d}" for d in data_rows)
    style_zone_donnees(ws, debut, r - 1, 1, 5, cols_montant=(2, 3, 4, 5))
    style_ligne_total(ws, r, 1, 5, cols_montant=(2, 3, 4, 5))
    r += 2
    ws.cell(r, 1, "Sans balance N-1, seule la colonne de clôture est servie ; "
                  "avec N-1, les mouvements sont posés en variation nette "
                  "(augmentation ou diminution) faute d'état des mouvements — "
                  "à ajuster depuis l'inventaire des immobilisations.")
    if spec.get("commentaires"):
        r = _commentaires(ws, r + 1, spec["commentaires"], 5)
    largeurs(ws, {"A": 56, "B": 17, "C": 17, "D": 17, "E": 17})
    return {"total": r}


def _render_plus_values(ws, spec, ident):
    r = entete_note(ws, spec["titre"], ident, 6)
    for i, h in enumerate(["Libellés", "Montant brut (A)",
                           "Amortissements pratiqués (B)",
                           "Valeur comptable nette (C = A - B)",
                           "Prix de cession (D)",
                           "Plus ou moins-value (E = D - C)"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 6)
    debut = r + 1
    for label, vnc, prix in spec["familles"]:
        r += 1
        ws.cell(r, 1, label)
        ws[f"D{r}"] = _f(vnc, "nd", "BALANCE")
        ws[f"E{r}"] = _f(prix, "nc", "BALANCE")
        ws[f"F{r}"] = f"=E{r}-D{r}"
    r += 1
    ws.cell(r, 1, "TOTAL GÉNÉRAL")
    for col in "BCDEF":
        ws[f"{col}{r}"] = f"=SUM({col}{debut}:{col}{r-1})"
    style_zone_donnees(ws, debut, r - 1, 1, 6, cols_montant=(2, 3, 4, 5, 6))
    style_ligne_total(ws, r, 1, 6, cols_montant=(2, 3, 4, 5, 6))
    r += 2
    ws.cell(r, 1, "Colonnes A et B à compléter depuis l'inventaire ; la valeur "
                  "comptable (81x) et le prix de cession (82x) sortent de la "
                  "balance. Mentionner la justification de la cession, la date "
                  "d'acquisition et la date de sortie.")
    largeurs(ws, {"A": 48, "B": 15, "C": 17, "D": 17, "E": 15, "F": 17})
    return {}


def _render_mouvements_prov(ws, spec, avec_n1, ident):
    r = entete_note(ws, spec["titre"], ident, 9)
    ws.cell(r, 1, "NATURE")
    ws.cell(r, 2, "Provisions à l'ouverture (A)")
    ws.cell(r, 3, "Dotations exploitation")
    ws.cell(r, 4, "Dotations financières")
    ws.cell(r, 5, "Dotations HAO")
    ws.cell(r, 6, "Reprises exploitation")
    ws.cell(r, 7, "Reprises financières")
    ws.cell(r, 8, "Reprises HAO")
    ws.cell(r, 9, "Provisions à la clôture (D = A + B - C)")
    style_entetes(ws, r, 1, 9)
    fam_cols = {"E": ("C", "F"), "F": ("D", "G"), "H": ("E", "H")}
    debut = r + 1
    bloc = r + 1
    totaux = []
    for ligne in spec["lignes"]:
        r += 1
        if ligne[0] == "s":
            ws.cell(r, 1, ligne[1])
            for col in "BCDEFGHI":
                ws[f"{col}{r}"] = f"=SUM({col}{bloc}:{col}{r-1})"
            style_ligne_total(ws, r, 1, 9,
                              cols_montant=(2, 3, 4, 5, 6, 7, 8, 9))
            totaux.append(r)
            bloc = r + 1
            continue
        _, label, expr, mode, fam = ligne
        ws.cell(r, 1, label)
        dot, rep = fam_cols[fam]
        if avec_n1:
            ws[f"B{r}"] = _f(expr, mode, "BALANCE_N1")
            fn = _f(expr, mode, "BALANCE")[1:]
            fn1 = _f(expr, mode, "BALANCE_N1")[1:]
            ws[f"{dot}{r}"] = f"=MAX(0,({fn})-({fn1}))"
            ws[f"{rep}{r}"] = f"=MAX(0,({fn1})-({fn}))"
        ws[f"I{r}"] = (f"=B{r}+C{r}+D{r}+E{r}-F{r}-G{r}-H{r}" if avec_n1
                       else _f(expr, mode, "BALANCE"))
    r += 1
    ws.cell(r, 1, "TOTAL")
    for col in "BCDEFGHI":
        ws[f"{col}{r}"] = "=" + "+".join(f"{col}{t}" for t in totaux)
    style_zone_donnees(ws, debut, r - 1, 1, 9,
                       cols_montant=(2, 3, 4, 5, 6, 7, 8, 9))
    style_ligne_total(ws, r, 1, 9, cols_montant=(2, 3, 4, 5, 6, 7, 8, 9))
    r += 2
    ws.cell(r, 1, "À défaut d'échéancier, dotations et reprises sont posées en "
                  "variation nette N/N-1 par nature — indiquer les évènements "
                  "et circonstances de chaque constitution et reprise.")
    largeurs(ws, {"A": 52, "B": 15, "C": 13, "D": 13, "E": 13, "F": 13,
                  "G": 13, "H": 13, "I": 16})
    return {}


def _render_texte(ws, spec, ident):
    r = entete_note(ws, spec["titre"], ident, 6)
    for bloc in spec["blocs"]:
        r += 1
        ws.cell(r, 1, bloc)
        ws.cell(r, 1).alignment = AL_GAUCHE
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=6)
        r += 2
    largeurs(ws, {"A": 20, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18})
    return {}


def _render_special1(ws, spec, avec_n1, ident):
    r = entete_note(ws, spec["titre"], ident, 6)
    bande_note(ws, r, 6, "DETTES GARANTIES PAR DES SÛRETÉS RÉELLES")
    r += 1
    for i, h in enumerate(["Libellés", "Note", "Montant brut",
                           "Hypothèques", "Nantissements", "Gages / autres"],
                          start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 6)
    debut = r + 1
    sections = [
        ("Emprunts et dettes financières diverses :", "18A", None, None),
        ("Emprunts obligataires", "", "181", "nc"),
        ("Emprunts et dettes des établissements de crédit", "", "182", "nc"),
        ("Autres dettes financières", "", "183,185,186,188", "nc"),
        ("Dettes de location-acquisition :", "18A", None, None),
        ("Dettes de crédit-bail immobilier", "", "1871", "nc"),
        ("Dettes de crédit-bail mobilier", "", "1872", "nc"),
        ("Dettes sur contrats de location-vente et autres", "",
         "1873,1876,187!1871,1872,1873,1876", "nc"),
        ("Dettes du passif circulant :", "", None, None),
        ("Fournisseurs et comptes rattachés", "19", "40!409", "c"),
        ("Adhérents, clients-usagers créditeurs", "9", "419", "c"),
        ("Personnel", "20", "42", "c"),
        ("Organismes sociaux", "20", "43", "c"),
        ("État et collectivités", "20", "44", "c"),
        ("Fondateurs, apporteurs et comptes rattachés", "21", "45", "c"),
        ("Bailleurs, État et autres organismes, fonds d'administration",
         "21", "46", "c"),
        ("Créditeurs divers", "21", "47!478,479", "c"),
    ]
    data = []
    for label, note, expr, mode in sections:
        r += 1
        ws.cell(r, 1, label)
        ws.cell(r, 2, note)
        if expr is None:
            ws.cell(r, 1).font = F_SOUS_TITRE
        else:
            ws[f"C{r}"] = _f(expr, mode, "BALANCE")
            data.append(r)
    r += 1
    ws.cell(r, 1, "TOTAL")
    ws[f"C{r}"] = "=" + "+".join(f"C{d}" for d in data)
    style_zone_donnees(ws, debut, r - 1, 1, 6, cols_montant=(3, 4, 5, 6))
    style_ligne_total(ws, r, 1, 6, cols_montant=(3,))
    r += 2
    ws.cell(r, 1, "Colonnes hypothèques / nantissements / gages : à servir "
                  "depuis les actes de sûretés — la balance ne les porte pas. "
                  "Indiquer la raison d'être des sûretés.")
    r += 2
    ws.cell(r, 1, "ENGAGEMENTS FINANCIERS")
    ws.cell(r, 1).font = F_SOUS_TITRE
    r += 1
    for i, h in enumerate(["Libellés", "Engagements réciproques",
                           "Engagements donnés", "Engagements reçus"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 4)
    d0 = r + 1
    for lab in ["Avals, cautions, garanties",
                "Hypothèques, nantissements, gages, autres",
                "Effets escomptés non échus"]:
        r += 1
        ws.cell(r, 1, lab)
    r += 1
    ws.cell(r, 1, "TOTAL")
    for col in "BCD":
        ws[f"{col}{r}"] = f"=SUM({col}{d0}:{col}{r-1})"
    style_zone_donnees(ws, d0, r - 1, 1, 4, cols_montant=(2, 3, 4))
    style_ligne_total(ws, r, 1, 4, cols_montant=(2, 3, 4))
    r += 2
    ws.cell(r, 1, "CONTRIBUTIONS VOLONTAIRES EN NATURE (classe 9)")
    ws.cell(r, 1).font = F_SOUS_TITRE
    r += 1
    for i, h in enumerate(["Libellés", "Ressources (91)", "Emplois (90)"],
                          start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 3)
    d0 = r + 1
    for lab, res, emp in [
            ("Dons en nature", "910", "900"),
            ("Mises à disposition gratuite des biens", None, "901"),
            ("Prestations en nature", "911", "902"),
            ("Personnel bénévole", "914", "904")]:
        r += 1
        ws.cell(r, 1, lab)
        if res:
            ws[f"B{r}"] = _f(res, "nc", "BALANCE")
        ws[f"C{r}"] = _f(emp, "nd", "BALANCE")
    r += 1
    ws.cell(r, 1, "TOTAL")
    for col in "BC":
        ws[f"{col}{r}"] = f"=SUM({col}{d0}:{col}{r-1})"
    style_zone_donnees(ws, d0, r - 1, 1, 3, cols_montant=(2, 3))
    style_ligne_total(ws, r, 1, 3, cols_montant=(2, 3))
    r += 2
    ws.cell(r, 1, "Évaluer les contributions volontaires à la valeur actuelle "
                  "(cadre conceptuel SYCEBNL).")
    largeurs(ws, {"A": 56, "B": 16, "C": 16, "D": 14, "E": 14, "F": 14})
    return {}


def _render_effectifs(ws, spec, ident):
    r = entete_note(ws, spec["titre"], ident, 9)
    ws.cell(r, 1, "1. PERSONNEL PROPRE")
    ws.cell(r, 1).font = F_SOUS_TITRE
    r += 1
    entetes = ["QUALIFICATIONS",
               "Effectifs Nationaux (M/F)", "Effectifs Autres États Région (M/F)",
               "Effectifs Hors Région (M/F)", "Effectifs Total",
               "Masse salariale Nationaux", "Masse salariale Autres États",
               "Masse salariale Hors Région", "Masse salariale Total"]
    for i, h in enumerate(entetes, start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 9)
    d0 = r + 1
    quals = ["YA. 1. Cadres supérieurs",
             "YB. 2. Techniciens supérieurs et cadres moyens",
             "YC. 3. Techniciens, agents de maîtrise et ouvriers qualifiés",
             "YD. 4. Employés, manœuvres, ouvriers et apprentis"]
    for q in quals:
        r += 1
        ws.cell(r, 1, q)
    r += 1
    ws.cell(r, 1, "YE. TOTAL (1)")
    for col in "BCDEFGHI":
        ws[f"{col}{r}"] = f"=SUM({col}{d0}:{col}{r-1})"
    style_ligne_total(ws, r, 1, 9, cols_montant=(6, 7, 8, 9))
    style_zone_donnees(ws, d0, r - 1, 1, 9, cols_montant=(6, 7, 8, 9))
    r += 1
    ws.cell(r, 1, "YF. Permanents")
    r += 1
    ws.cell(r, 1, "YG. Saisonniers")
    r += 2
    ws.cell(r, 1, "2. PERSONNEL EXTÉRIEUR ET BÉNÉVOLE — facturation à l'entité")
    ws.cell(r, 1).font = F_SOUS_TITRE
    r += 1
    d0 = r + 1
    for q in ["YH. 1. Cadres supérieurs",
              "YI. 2. Techniciens supérieurs et cadres moyens",
              "YJ. 3. Techniciens, agents de maîtrise et ouvriers qualifiés",
              "YK. 4. Employés, manœuvres, ouvriers et apprentis"]:
        r += 1
        ws.cell(r, 1, q)
    r += 1
    ws.cell(r, 1, "YL. TOTAL (2)")
    ws[f"B{r}"] = f"=SUM(B{d0}:B{r-1})"
    style_ligne_total(ws, r, 1, 2, cols_montant=(2,))
    r += 1
    ws.cell(r, 1, "YO. TOTAL (1 + 2) : à rapprocher de la note 29A et du "
                  "compte 637")
    largeurs(ws, {"A": 52, "B": 14, "C": 14, "D": 14, "E": 12, "F": 14,
                  "G": 14, "H": 14, "I": 14})
    return {}


def _render_budget(ws, spec, ident):
    r = entete_note(ws, spec["titre"], ident, 8)
    for i, h in enumerate(["Code", "Libellé", "Budget de l'exercice (1)",
                           "Décaissement (2)", "Engagement (3)",
                           "Réalisation (4 = 2 + 3)",
                           "Crédit disponible (5 = 1 - 4)",
                           "Exécution budget % (4/1)"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 8)
    d0 = r + 1
    for _ in range(12):
        r += 1
        ws[f"F{r}"] = f"=D{r}+E{r}"
        ws[f"G{r}"] = f"=C{r}-F{r}"
        ws[f"H{r}"] = f"=IF(C{r}=0,\"\",F{r}/C{r})"
        ws[f"H{r}"].number_format = "0.0%"
    r += 1
    ws.cell(r, 2, "TOTAL")
    for col in "CDEFG":
        ws[f"{col}{r}"] = f"=SUM({col}{d0}:{col}{r-1})"
    ws[f"H{r}"] = f"=IF(C{r}=0,\"\",F{r}/C{r})"
    ws[f"H{r}"].number_format = "0.0%"
    style_zone_donnees(ws, d0, r - 1, 1, 8, cols_montant=(3, 4, 5, 6, 7))
    style_ligne_total(ws, r, 1, 8, cols_montant=(3, 4, 5, 6, 7))
    r += 2
    ws.cell(r, 1, "Remplir code et libellé suivant la nomenclature budgétaire "
                  "de l'entité ; expliquer les écarts significatifs entre "
                  "budget et réalisation.")
    largeurs(ws, {"A": 10, "B": 40, "C": 16, "D": 15, "E": 15, "F": 16,
                  "G": 17, "H": 14})
    return {}


def _render_synthese(ws, spec, avec_n1, ident, refs):
    """NOTE 33 — formules croisées vers ACTIF/PASSIF/CR/TFT."""
    r = entete_note(ws, spec["titre"], ident, 5)
    for i, h in enumerate(["(en unités monétaires légales)", "Année N",
                           "Année N-1", "Variation en valeur",
                           "Variation en %"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 5)
    na = q(refs.get("NOM_ACTIF", "ACTIF"))
    npf = q(refs.get("NOM_PASSIF", "PASSIF"))
    ncr = q(refs.get("CR_NOM", "Compte de Resultat"))
    ntf = q(refs.get("NOM_TFT", "TFT"))
    CRn, CRn1 = f"{ncr}!D", f"{ncr}!E"
    A_n, A_n1 = f"{na}!F", f"{na}!G"
    P_n, P_n1 = f"{npf}!D", f"{npf}!E"
    cr, ac, pa, tft = refs["CR"], refs["ACTIF"], refs["PASSIF"], refs.get("TFT", {})

    def s(expr, mode, feuille="BALANCE"):
        return _f(expr, mode, feuille)[1:]

    def paire(fN, fN1):
        return (fN, fN1 if avec_n1 else None)

    lignes = [
        ("h", "ANALYSE DE L'ACTIVITÉ — SOLDES INTERMÉDIAIRES DE GESTION"),
        ("v", "Résultat des activités ordinaires",
         paire(f"={CRn}{cr['XC']}", f"={CRn1}{cr['XC']}")),
        ("v", "Résultat hors activités ordinaires",
         paire(f"={CRn}{cr['XD']}", f"={CRn1}{cr['XD']}")),
        ("v", "Résultat net",
         paire(f"={CRn}{cr['XE']}", f"={CRn1}{cr['XE']}")),
        ("v", "Capacité d'autofinancement globale (CAFG) (a)",
         paire(f"={CRn}{cr['XE']}+{s('68,69,85', 'nd')}-{s('79,86', 'nc')}"
               f"+{s('81', 'nd')}-{s('82', 'nc')}",
               f"={CRn1}{cr['XE']}+{s('68,69,85', 'nd', 'BALANCE_N1')}"
               f"-{s('79,86', 'nc', 'BALANCE_N1')}+{s('81', 'nd', 'BALANCE_N1')}"
               f"-{s('82', 'nc', 'BALANCE_N1')}")),
        ("v", "Ratio de cotisations acquises = cotisations / charges de "
              "l'exercice (b)",
         paire(f"=IF({CRn}{cr['XB']}=0,\"\",{s('701', 'nc')}/ABS({CRn}{cr['XB']}))",
               f"=IF({CRn1}{cr['XB']}=0,\"\",{s('701', 'nc', 'BALANCE_N1')}"
               f"/ABS({CRn1}{cr['XB']}))"), "pct"),
        ("man", "Ratio d'utilisation des dons = sommes versées aux "
                "bénéficiaires / sommes collectées brutes"),
        ("h", "ANALYSE DE LA STRUCTURE FINANCIÈRE"),
        ("v", "Ressources stables (c)",
         paire(f"={P_n}{pa['DE']}", f"={P_n1}{pa['DE']}")),
        ("v", "-  Actif immobilisé (c)",
         paire(f"={A_n}{ac['AZ']}", f"={A_n1}{ac['AZ']}")),
        ("vx", "FONDS DE ROULEMENT (1) = ressources stables - actif immobilisé", "-", 2),
        ("v", "Actif circulant d'exploitation (c)",
         paire(f"={A_n}{ac['BB']}+{A_n}{ac['BC']}+{A_n}{ac['BD']}+{A_n}{ac['BE']}",
               f"={A_n1}{ac['BB']}+{A_n1}{ac['BC']}+{A_n1}{ac['BD']}+{A_n1}{ac['BE']}")),
        ("v", "-  Passif circulant d'exploitation (c)",
         paire(f"={P_n}{pa['DG']}+{P_n}{pa['DH']}+{P_n}{pa['DI']}",
               f"={P_n1}{pa['DG']}+{P_n1}{pa['DH']}+{P_n1}{pa['DI']}")),
        ("vx", "BESOIN DE FINANCEMENT D'EXPLOITATION (2)", "-", 2),
        ("v", "Actif circulant HAO (c)",
         paire(f"={A_n}{ac['BA']}", f"={A_n1}{ac['BA']}")),
        ("v", "-  Passif circulant HAO (c)",
         paire(f"={P_n}{pa['DF']}", f"={P_n1}{pa['DF']}")),
        ("vx", "BESOIN DE FINANCEMENT HAO (3)", "-", 2),
        ("vsum", "BESOIN DE FINANCEMENT GLOBAL (4) = (2) + (3)", ("BFE", "BFH")),
        ("vfr", "TRÉSORERIE NETTE (5) = (1) - (4)", None),
        ("v", "CONTRÔLE : trésorerie nette = trésorerie actif - trésorerie passif",
         paire(f"={A_n}{ac['BX']}-{P_n}{pa['DX']}",
               f"={A_n1}{ac['BX']}-{P_n1}{pa['DX']}")),
        ("h", "ANALYSE DE LA SOLVABILITÉ"),
        ("v", "Ratio de liquidité générale = (créances + trésorerie actif) / "
              "passif circulant",
         paire(f"=IF({P_n}{pa['DV']}=0,\"\",({A_n}{ac['BC']}+{A_n}{ac['BD']}"
               f"+{A_n}{ac['BE']}+{A_n}{ac['BX']})/{P_n}{pa['DV']})",
               f"=IF({P_n1}{pa['DV']}=0,\"\",({A_n1}{ac['BC']}+{A_n1}{ac['BD']}"
               f"+{A_n1}{ac['BE']}+{A_n1}{ac['BX']})/{P_n1}{pa['DV']})"), "ratio"),
    ]
    if tft:
        lignes += [
            ("h", "ANALYSE DE LA VARIATION DE LA TRÉSORERIE"),
            ("v", "Flux de trésorerie des activités opérationnelles",
             paire(f"={ntf}!D{tft['ZB']}", f"={ntf}!E{tft['ZB']}")),
            ("v", "Flux de trésorerie des activités d'investissement",
             paire(f"={ntf}!D{tft['ZC']}", f"={ntf}!E{tft['ZC']}")),
            ("v", "Flux de trésorerie des activités de financement",
             paire(f"={ntf}!D{tft['ZD']}+{ntf}!D{tft['ZE']}",
                   f"={ntf}!E{tft['ZD']}+{ntf}!E{tft['ZE']}")),
            ("v", "VARIATION DE LA TRÉSORERIE NETTE DE LA PÉRIODE",
             paire(f"={ntf}!D{tft['ZF']}", f"={ntf}!E{tft['ZF']}")),
        ]

    marqueurs = {}
    pile_vx = []
    for ligne in lignes:
        r += 1
        kind = ligne[0]
        if kind == "h":
            bande_note(ws, r, 5, ligne[1])
            pile_vx = []
            continue
        if kind == "man":
            ws.cell(r, 1, ligne[1])
            continue
        if kind == "v":
            label, (fN, fN1) = ligne[1], ligne[2]
            ws.cell(r, 1, label)
            ws[f"B{r}"] = fN
            if fN1:
                ws[f"C{r}"] = fN1
            fmt = ligne[3] if len(ligne) > 3 else None
            if fmt in ("pct", "ratio"):
                ws[f"B{r}"].number_format = "0.00"
                ws[f"C{r}"].number_format = "0.00"
            pile_vx.append(r)
            continue
        if kind == "vx":       # total = premier - somme des suivants
            label, op, npil = ligne[1], ligne[2], ligne[3]
            src = pile_vx[-npil:]
            ws.cell(r, 1, label)
            for col in ("B", "C") if avec_n1 else ("B",):
                ws[f"{col}{r}"] = f"={col}{src[0]}" + "".join(
                    f"{op}{col}{x}" for x in src[1:])
            style_ligne_total(ws, r, 1, 5, cols_montant=(2, 3, 4))
            if "FONDS DE ROULEMENT" in label:
                marqueurs["FR"] = r
            elif "D'EXPLOITATION (2)" in label:
                marqueurs["BFE"] = r
            elif "HAO (3)" in label:
                marqueurs["BFH"] = r
            pile_vx = []
            continue
        if kind == "vsum":
            label, cles = ligne[1], ligne[2]
            ws.cell(r, 1, label)
            for col in ("B", "C") if avec_n1 else ("B",):
                ws[f"{col}{r}"] = "=" + "+".join(
                    f"{col}{marqueurs[k]}" for k in cles)
            style_ligne_total(ws, r, 1, 5, cols_montant=(2, 3, 4))
            marqueurs["BFG"] = r
            continue
        if kind == "vfr":
            ws.cell(r, 1, ligne[1])
            for col in ("B", "C") if avec_n1 else ("B",):
                ws[f"{col}{r}"] = f"={col}{marqueurs['FR']}-{col}{marqueurs['BFG']}"
            style_ligne_total(ws, r, 1, 5, cols_montant=(2, 3, 4))
            continue
    # variations
    if avec_n1:
        for rr in range(8, r + 1):
            if ws[f"B{rr}"].value and ws[f"C{rr}"].value \
                    and str(ws[f"B{rr}"].value).startswith("="):
                ws[f"D{rr}"] = f"=IF(OR(B{rr}=\"\",C{rr}=\"\"),\"\",B{rr}-C{rr})"
                ws[f"E{rr}"] = (f"=IF(OR(C{rr}=\"\",C{rr}=0),\"\","
                                f"(B{rr}-C{rr})/C{rr})")
                ws[f"E{rr}"].number_format = "0.0%"
    style_zone_donnees(ws, 8, r, 1, 5, cols_montant=(2, 3, 4))
    r += 2
    ws.cell(r, 1, "(a) CAFG = résultat net + dotations aux amortissements, "
                  "dépréciations et provisions - reprises + valeurs comptables "
                  "des cessions - produits des cessions. (b) Variations des "
                  "ratios en points. (c) Écarts de conversion à éliminer pour "
                  "ramener créances et dettes à leur valeur initiale.")
    largeurs(ws, {"A": 62, "B": 16, "C": 16, "D": 16, "E": 11})
    return {}


# --------------------------------------------------------------------------
# Point d'entrée
# --------------------------------------------------------------------------

def construire_notes(wb, avec_n1, ident, refs, notes=None):
    """Construit un jeu de notes (par défaut les 35 notes des associations).
    `refs` : {'ACTIF': {ref: row}, ...} pour les formules croisées.
    Renvoie les lignes de contrôle [(libellé, formule, attendu)] à reporter
    en feuille CONTROLES."""
    controles = []
    for spec in (notes if notes is not None else NOTES_ASSOCIATIONS):
        ws = wb.create_sheet(spec["feuille"])
        t = spec["type"]
        ids = {}
        if t == "table":
            ids = _render_table(ws, spec, avec_n1, ident)
        elif t == "texte":
            _render_texte(ws, spec, ident)
        elif t == "mouvements_brut":
            _render_mouvements(ws, spec, avec_n1, ident, "nd",
                               ["RUBRIQUES", "Montant brut à l'ouverture (A)",
                                "Augmentations (B)", "Diminutions (C)",
                                "Montant brut à la clôture (D = A + B - C)"])
        elif t in ("mouvements_amort", "mouvements_deprec"):
            _render_mouvements(ws, spec, avec_n1, ident, "nc",
                               ["RUBRIQUES", "Cumul à l'ouverture (A)",
                                "Augmentations : dotations (B)",
                                "Diminutions : reprises et sorties (C)",
                                "Cumul à la clôture (D = A + B - C)"])
        elif t == "plus_values":
            _render_plus_values(ws, spec, ident)
        elif t == "mouvements_prov":
            _render_mouvements_prov(ws, spec, avec_n1, ident)
        elif t == "special1":
            _render_special1(ws, spec, avec_n1, ident)
        elif t == "effectifs":
            _render_effectifs(ws, spec, ident)
        elif t == "budget":
            _render_budget(ws, spec, ident)
        elif t == "bailleur":
            _render_bailleur(ws, spec, ident)
        elif t == "synthese":
            _render_synthese(ws, spec, avec_n1, ident, refs)

        # ligne de contrôle note ↔ état
        ctl = spec.get("controle")
        if ctl and ids.get(ctl[2] if len(ctl) > 2 else "net"):
            etat, ref = ctl[0], ctl[1]
            row_note = ids[ctl[2] if len(ctl) > 2 else "net"]
            cols = COLS[spec.get("cols", "std5")]
            cell_note = f"'{spec['feuille']}'!{cols['n']}{row_note}"
            na = q(refs.get("NOM_ACTIF", "ACTIF"))
            npf = q(refs.get("NOM_PASSIF", "PASSIF"))
            ncr = q(refs.get("CR_NOM", "Compte de Resultat"))
            if etat == "ACTIF-net":
                cible = f"{na}!F{refs['ACTIF'][ref]}"
            elif etat == "ACTIF-D":
                cible = f"{na}!D{refs['ACTIF'][ref]}"
            elif etat == "PASSIF":
                cible = f"{npf}!D{refs['PASSIF'][ref]}"
            else:
                cible = f"{ncr}!D{refs['CR'][ref]}"
            if etat == "CR":
                # au CR les charges sont stockées signées (négatives)
                controles.append((f"{spec['feuille']} total vs poste {ref}",
                                  f"={cell_note}+{cible}", 0))
            else:
                controles.append((f"{spec['feuille']} total vs poste {ref}",
                                  f"={cell_note}-{cible}", 0))
    return controles


def _render_bailleur(ws, spec, ident):
    """NOTE 9 des projets de développement — fonds du bailleur."""
    r = entete_note(ws, spec["titre"], ident, 7)
    for i, h in enumerate(["Date des décaissements",
                           "Bailleur / sous-projet 1 : montant décaissé",
                           "Montant consommé", "Solde restant",
                           "Bailleur / sous-projet 2 : montant décaissé",
                           "Montant consommé", "Solde restant"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 7)
    for bloc, note in (("FONDS D'INVESTISSEMENT", ""),
                       ("FONDS D'ADMINISTRATION (2)", "")):
        r += 1
        bande_note(ws, r, 7, bloc)
        d0 = r + 1
        for _ in range(4):
            r += 1
            ws[f"D{r}"] = f"=B{r}-C{r}"
            ws[f"G{r}"] = f"=E{r}-F{r}"
        r += 1
        ws.cell(r, 1, "TOTAL " + bloc)
        for col in "BCDEFG":
            ws[f"{col}{r}"] = f"=SUM({col}{d0}:{col}{r-1})"
        style_zone_donnees(ws, d0, r - 1, 1, 7,
                           cols_montant=(2, 3, 4, 5, 6, 7))
        style_ligne_total(ws, r, 1, 7, cols_montant=(2, 3, 4, 5, 6, 7))
    r += 2
    ws.cell(r, 1, "Rappels balance (formules) — à ventiler par bailleur :")
    ws.cell(r, 1).font = F_GRAS
    rappels = [
        ("Fonds affectés aux investissements non consommés (soldes 162-164)",
         "162,163,164", "nc"),
        ("Fonds d'administration non consommés (soldes 462-464)",
         "462,463,464", "nc"),
        ("Quote-part de fonds d'administration consommée de l'exercice (702)",
         "702", "nc"),
    ]
    for lab, expr, mode in rappels:
        r += 1
        ws.cell(r, 1, lab)
        ws[f"B{r}"] = _f(expr, mode, "BALANCE")
        ws[f"B{r}"].number_format = FMT_MONTANT
    r += 2
    ws.cell(r, 1, "(1) une colonne par bailleur et/ou sous-projet. (2) le "
                  "montant consommé d'un exercice = solde du compte 702, à "
                  "subdiviser par nature de projet. Indiquer le niveau "
                  "d'utilisation des fonds en % et expliquer les retards "
                  "d'exécution éventuels.")
    largeurs(ws, {"A": 40, "B": 17, "C": 16, "D": 14, "E": 17, "F": 16,
                  "G": 14})
    return {}


def parties_depuis_specs(specs, decoupage):
    """Construit les parties de la fiche récapitulative depuis les specs de
    notes. `decoupage` : liste de (titre_partie, borne_min, borne_max) sur le
    numéro de tête de la note (5A -> 5)."""
    import re as _re
    parties = [(titre, []) for titre, _, _ in decoupage]
    for spec in specs:
        feuille = spec["feuille"]
        m = _re.match(r"NOTE (\d+)", feuille)
        if not m:
            continue
        num = int(m.group(1))
        for i, (_, lo, hi) in enumerate(decoupage):
            if lo <= num <= hi:
                parties[i][1].append((feuille.replace("NOTE", "Note"),
                                      spec["titre"]))
                break
    return parties
