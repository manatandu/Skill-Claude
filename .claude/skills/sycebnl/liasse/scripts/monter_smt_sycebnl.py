#!/usr/bin/env python3
"""
monter_smt_sycebnl.py — Monte les états financiers du SYSTÈME MINIMAL DE
TRÉSORERIE du SYCEBNL (Partie 4, ch. 4 du Journal officiel : bilan GA→HZ,
compte de résultat KA→KZC, 5 notes annexes) à partir d'une balance, dans un
classeur professionnel construit de toutes pièces.

    python monter_smt_sycebnl.py balance_N.xlsx [balance_N1.xlsx] \
        --sortie etats-smt-sycebnl.xlsx --entite "..." --identifiant "..." \
        --exercice "31/12/N" --duree 12

Entités visées : petites EBNL sous le seuil de 30 M FCFA de ressources
annuelles (Acte uniforme, art. 5-6) tenant une comptabilité de trésorerie.

Jeu d'états produit :
  - BILAN ACTIF (GA→GZ) et BILAN PASSIF (HA→HZ), une feuille chacun ;
  - COMPTE DE RÉSULTAT (KA→KZC : recettes/dépenses corrigées des variations
    de stocks, créances, dettes et des dotations) ;
  - NOTE 1 (acquisition et suivi des immobilisations), NOTE 2 (état des
    stocks), NOTE 3 (créances et dettes non échues), NOTE 4 (journal unique
    de trésorerie, ventilations officielles), NOTE 5 (dotations) ;
  - GARDE / BALANCE / BALANCE_N1 / CONTROLES / ANOMALIES.

Chaque montant est une FORMULE Excel (SUMIF sur BALANCE / BALANCE_N1) —
correspondance : references/correspondance-smt-sycebnl.tsv (construction du
moteur : le ch. 4 ne publie pas de table de correspondance).
"""

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from formules_sycebnl import (
    formule_tokens, set_lignes_max, q, nom_feuille,
    F_TITRE, F_SOUS_TITRE, F_ENTETE, F_NORMAL, F_GRAS,
    R_TITRE, R_ENTETE, R_BANDE, R_TOTAL, BORD_FIN, AL_CENTRE, AL_GAUCHE,
    FMT_MONTANT, style_entetes, style_zone_donnees, style_ligne_total,
    largeurs, style_titre, retirer_tirets, construire_identification,
    construire_fiche_notes, ordonner_feuilles,
    ecrire_cartouche, titre_etat, titre_note, entetes_bande, style_ligne,
    cadre, MOYEN, set_identite_etendue, construire_couverture,
    construire_garde_etafi, construire_fiche2, construire_controle_balance,
    construire_table_commentaires, construire_bilan_paysage,
    appliquer_police_arial, numeroter_pages, NOM_BALANCE, NOM_BALANCE_N1,
)
from monter_etats_sycebnl import lire_balance

NOM_ACTIF = "Bilan-Actif"
NOM_PASSIF = "Bilan-Passif"
CR_NOM = "Résultat"
PAGE_SYS = "SYCEBNL - SMT"


def entete_smt(ws, titre, page_ref, ident, ncols, note=False, taille=14):
    """Cartouche ETAFI + titre (vert pour un état, bleu nuit pour une
    note) : rend la ligne des en-têtes de colonnes."""
    ecrire_cartouche(ws, ident, page_ref, max(ncols, 5))
    if note:
        titre_note(ws, titre, max(ncols, 5), row=7)
    else:
        titre_etat(ws, titre, 1, max(ncols, 5), row=7, taille=taille)
    return 8


def _c(f):
    return f[1:] if f and f.startswith("=") else (f or "0")


def _tok(s):
    return [t.strip() for t in s.split(",") if t.strip()]


def f_actif(ref, bal):
    if ref == "GA":
        return formule_tokens(_tok("2"), "nd", bal)
    if ref == "GB":
        return formule_tokens(_tok("3"), "nd", bal)
    if ref == "GC":
        deb4 = _c(formule_tokens(_tok("4"), "d", bal))
        dep = _c(formule_tokens(_tok("49,590,591"), "c", bal))
        titres = _c(formule_tokens(_tok("50,51"), "nd", bal))
        return f"={deb4}-({dep})+({titres})"
    if ref == "GD":
        return formule_tokens(_tok("57"), "nd", bal)
    if ref == "GE":
        banques = _c(formule_tokens(_tok("52,53,55,56,58"), "nd", bal))
        dep = _c(formule_tokens(_tok("592,593,595"), "c", bal))
        return f"={banques}-({dep})"
    return None


def f_passif(ref, bal, kzc_row):
    if ref == "HA":
        return formule_tokens(_tok("10"), "nc", bal)
    if ref == "HB":
        f13 = _c(formule_tokens(_tok("13"), "nc", bal))
        col = "D" if bal == "BALANCE" else "E"
        return f"={q(CR_NOM)}!{col}{kzc_row}+({f13})"
    if ref == "HC":
        return formule_tokens(_tok("11,12,14,15,16,17,18,19"), "nc", bal)
    if ref == "HD":
        cred4 = _c(formule_tokens(_tok("4"), "c", bal, exclude=_tok("49")))
        p599 = _c(formule_tokens(_tok("599"), "nc", bal))
        return f"={cred4}+({p599})"
    return None


CR_LIGNES = [
    ("KA", "Revenus encaissés", "70", "nc", "4", False),
    ("KB", "Autres recettes sur activités",
     "71,72,75,77,78,79,82,84,86,88", "nc", "4", False),
    ("KX", "TOTAL DES REVENUS ENCAISSÉS (A)", None, None, "", True),
    ("JA", "Dépenses sur achats", "60!603", "nd", "4", False),
    ("JB", "Dépenses sur loyers", "622", "nd", "4", False),
    ("JC", "Dépenses sur salaires", "66", "nd", "4", False),
    ("JD", "Dépenses sur impôts et taxes", "64", "nd", "4", False),
    ("JE", "Charges d'intérêts", "67", "nd", "4", False),
    ("JF", "Autres dépenses sur activités", "61,62!622,63,65,81,83,87",
     "nd", "4", False),
    ("JX", "TOTAL DÉPENSES SUR CHARGES (B)", None, None, "", True),
    ("KZ", "SOLDE : excédent (+) ou insuffisance (-) de recettes (C = A - B)",
     None, None, "", True),
    ("VA", "+ Variations des stocks sur les achats [N - (N-1)]",
     None, None, "2", False),
    ("VB", "+ Variation des créances [N - (N-1)] — à saisir depuis la Note 3",
     None, None, "3", False),
    ("VC", "- Variation des dettes d'exploitation [N - (N-1)] — à saisir "
           "depuis la Note 3", None, None, "3", False),
    ("JG", "DOTATIONS AUX AMORTISSEMENTS", "68,69,85", "nd", "", False),
    ("KZC", "RÉSULTAT NET DE L'EXERCICE", None, None, "", True),
]


NIVEAUX_CR_SMT = {"KX": "section", "JX": "section", "KZ": "inter",
                  "KZC": "section"}


def construire_cr(wb, avec_n1, ident):
    ws = wb.create_sheet(CR_NOM)
    r = entete_smt(ws, "COMPTE DE RESULTAT", f"COMPTE DE RESULTAT\n{PAGE_SYS}",
                   ident, 5)
    for i, h in enumerate(["REF", "LIBELLES", "NOTE", "EXERCICE N",
                           "EXERCICE N-1"], start=1):
        ws.cell(r, i, h)
    entetes_bande(ws, r, r, 1, 5)
    ws.row_dimensions[r].height = 22
    debut = r + 1
    rows = {}
    for ref, lib, jetons, mode, note, total in CR_LIGNES:
        r += 1
        rows[ref] = r
        ws.cell(r, 1, ref)
        ws.cell(r, 2, lib)
        ws.cell(r, 3, note)
        if jetons:
            inc, exc = (jetons.split("!") + [""])[:2]
            for col, bal, actif in ((4, "BALANCE", True),
                                    (5, "BALANCE_N1", avec_n1)):
                if actif:
                    ws.cell(r, col).value = formule_tokens(
                        _tok(inc), mode, bal, exclude=_tok(exc))
        style_ligne(ws, r, 1, 5, NIVEAUX_CR_SMT.get(ref, "normal"),
                    cols_montant=(4, 5), col_ref=1)
        ws.row_dimensions[r].height = 22

    def pose(ref, fN, fN1):
        ws.cell(rows[ref], 4).value = fN
        if avec_n1:
            ws.cell(rows[ref], 5).value = fN1

    pose("KX", f"=D{rows['KA']}+D{rows['KB']}", f"=E{rows['KA']}+E{rows['KB']}")
    jd = "+".join(f"D{rows[x]}" for x in ("JA", "JB", "JC", "JD", "JE", "JF"))
    je = "+".join(f"E{rows[x]}" for x in ("JA", "JB", "JC", "JD", "JE", "JF"))
    pose("JX", f"={jd}", f"={je}")
    pose("KZ", f"=D{rows['KX']}-D{rows['JX']}", f"=E{rows['KX']}-E{rows['JX']}")
    if avec_n1:
        s_n = _c(formule_tokens(_tok("3"), "nd", "BALANCE"))
        s_n1 = _c(formule_tokens(_tok("3"), "nd", "BALANCE_N1"))
        ws.cell(rows["VA"], 4).value = f"=({s_n})-({s_n1})"
    else:
        v603 = _c(formule_tokens(_tok("603"), "nc", "BALANCE"))
        v73 = _c(formule_tokens(_tok("73"), "nc", "BALANCE"))
        ws.cell(rows["VA"], 4).value = f"=({v603})+({v73})"
    ws.cell(rows["VB"], 4).value = 0
    ws.cell(rows["VC"], 4).value = 0
    pose("KZC",
         f"=D{rows['KZ']}+D{rows['VA']}+D{rows['VB']}-D{rows['VC']}-D{rows['JG']}",
         f"=E{rows['KZ']}+E{rows['VA']}+E{rows['VB']}-E{rows['VC']}-E{rows['JG']}")
    cadre(ws, debut - 1, 1, r, 5, MOYEN)
    r += 2
    ws.cell(r, 1, "Les variations de créances (VB) et de dettes (VC) se "
                  "saisissent depuis l'état extra-comptable de la Note 3 ; si "
                  "la balance porte une classe 4 mouvementée (base "
                  "engagement), les laisser à zéro.")
    ws.cell(r, 1).font = F_NORMAL
    largeurs(ws, {"A": 6, "B": 60, "C": 6.5, "D": 15.7, "E": 15.7})
    return rows


def construire_bilan(wb, avec_n1, ident, kzc_row):
    lignes_a = [("GA", "Immobilisations (1)", "1"),
                ("GB", "Stocks", "2"),
                ("GC", "Adhérents, clients-usagers et autres débiteurs", "3"),
                ("GD", "Caisse", "4"),
                ("GE", "Banque (en + ou en -)", "4")]
    lignes_p = [("HA", "Dotations", "5"),
                ("HB", "Résultat net de l'exercice (en + ou en -)", ""),
                ("HC", "Autres fonds propres", ""),
                ("HD", "Fournisseurs et autres créditeurs", "3")]
    infos = {}
    for nom, lignes, total_ref, total_lab, f_poste, page in (
            (NOM_ACTIF, lignes_a, "GZ", "TOTAL ACTIF", f_actif, "PAGE 1/2"),
            (NOM_PASSIF, lignes_p, "HZ", "TOTAL PASSIF",
             lambda ref, bal: f_passif(ref, bal, kzc_row), "PAGE 2/2")):
        ws = wb.create_sheet(nom)
        cote = "ACTIF" if nom == NOM_ACTIF else "PASSIF"
        r = entete_smt(ws, "BILAN", f"BILAN {PAGE_SYS}\n{page}", ident, 5,
                       taille=16)
        ws.cell(r, 1, "REF")
        ws.cell(r, 2, cote)
        ws.cell(r, 3, "NOTE")
        ws.cell(r, 4, "EXERCICE N")
        ws.cell(r, 5, "EXERCICE N-1")
        entetes_bande(ws, r, r, 1, 5)
        ws.row_dimensions[r].height = 22
        premiere = r + 1
        for ref, lib, note in lignes:
            r += 1
            ws.cell(r, 1, ref)
            ws.cell(r, 2, lib)
            ws.cell(r, 3, note)
            ws.cell(r, 4).value = f_poste(ref, "BALANCE")
            if avec_n1:
                ws.cell(r, 5).value = f_poste(ref, "BALANCE_N1")
            style_ligne(ws, r, 1, 5, "normal", cols_montant=(4, 5), col_ref=1)
            ws.row_dimensions[r].height = 22
        r += 1
        ws.cell(r, 1, total_ref)
        ws.cell(r, 2, total_lab)
        ws.cell(r, 4).value = f"=SUM(D{premiere}:D{r-1})"
        if avec_n1:
            ws.cell(r, 5).value = f"=SUM(E{premiere}:E{r-1})"
        style_ligne(ws, r, 1, 5, "general", cols_montant=(4, 5), col_ref=1)
        ws.row_dimensions[r].height = 22
        cadre(ws, premiere - 1, 1, r, 5, MOYEN)
        infos[total_ref] = r
        infos[nom] = (premiere, r)
        r += 2
        ws.cell(r, 1, "(1) à faire figurer sur l'état de situation si "
                      "montants significatifs (Partie 4, ch. 4)."
                if cote == "ACTIF" else
                "Autres fonds propres : réserves, report à nouveau, "
                "subventions, fonds affectés/reportés, emprunts et "
                "provisions (le modèle SMT ne les distingue pas).")
        ws.cell(r, 1).font = F_NORMAL
        largeurs(ws, {"A": 6, "B": 52, "C": 6.5, "D": 15.7, "E": 15.7})
    return infos


def construire_note1(wb, bal, ident):
    ws = wb.create_sheet("NOTE 1 IMMOBILISATIONS")
    r = entete_smt(ws, "NOTE 1 : TABLEAU D'ACQUISITION ET DE SUIVI DU "
                       "MATERIEL, DU MOBILIER ET AUTRES IMMOBILISATIONS",
                   f"NOTE 1\n{PAGE_SYS}", ident, 7, note=True)
    for i, h in enumerate(["Date", "Désignation", "Montant",
                           "Date d'acquisition", "Durée d'utilité",
                           "Date de sortie", "Prix de cession"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 7)
    premiere = r + 1
    for l in bal:
        c = l["compte"]
        if c.startswith("2") and not c.startswith(("28", "29")) \
                and abs(l["sd"] - l["sc"]) > 0.005:
            r += 1
            ws.cell(r, 2, f"{c} — {l['libelle']}")
            ws.cell(r, 3).value = formule_tokens([c], "nd", "BALANCE")
    r += 1
    ws.cell(r, 2, "TOTAL IMMOBILISATIONS BRUTES")
    ws.cell(r, 3).value = (f"=SUM(C{premiere}:C{r-1})" if r > premiere else 0)
    r += 1
    ws.cell(r, 2, "Amortissements et dépréciations cumulés (28/29)")
    ws.cell(r, 3).value = formule_tokens(_tok("28,29"), "nc", "BALANCE")
    r += 1
    ws.cell(r, 2, "VALEUR NETTE (= poste GA du bilan)")
    ws.cell(r, 3).value = f"=C{r-2}-C{r-1}"
    style_zone_donnees(ws, premiere, r, 1, 7, cols_montant=(3, 7))
    style_ligne_total(ws, r, 1, 7, cols_montant=(3,))
    largeurs(ws, {"A": 11, "B": 48, "C": 15, "D": 15, "E": 13, "F": 13, "G": 15})


def construire_note2(wb, bal, avec_n1, ident):
    ws = wb.create_sheet("NOTE 2 STOCKS")
    r = entete_smt(ws, "NOTE 2 : ETAT DES STOCKS",
                   f"NOTE 2\n{PAGE_SYS}", ident, 5, note=True)
    for i, h in enumerate(["Référence", "Désignation", "Quantité",
                           "Prix unitaire", "Montant"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 5)
    premiere = r + 1
    for l in bal:
        if l["compte"].startswith("3") and not l["compte"].startswith("39") \
                and abs(l["sd"] - l["sc"]) > 0.005:
            r += 1
            ws.cell(r, 1, l["compte"])
            ws.cell(r, 2, l["libelle"])
            ws.cell(r, 5).value = formule_tokens([l["compte"]], "nd", "BALANCE")
    r += 3
    ws.cell(r, 2, "VALEUR DU STOCK FINAL")
    ws.cell(r, 5).value = formule_tokens(_tok("3"), "nd", "BALANCE")
    style_ligne_total(ws, r, 1, 5, cols_montant=(5,))
    r += 1
    ws.cell(r, 2, "VALEUR DU STOCK INITIAL")
    ws.cell(r, 5).value = (formule_tokens(_tok("3"), "nd", "BALANCE_N1")
                           if avec_n1 else 0)
    style_ligne_total(ws, r, 1, 5, cols_montant=(5,))
    style_zone_donnees(ws, premiere, r - 4, 1, 5, cols_montant=(5,))
    largeurs(ws, {"A": 12, "B": 46, "C": 12, "D": 14, "E": 16})


def construire_note3(wb, avec_n1, ident):
    ws = wb.create_sheet("NOTE 3 CREANCES-DETTES")
    r = entete_smt(ws, "NOTE 3 : ETAT DES CREANCES ET DES DETTES NON ECHUES",
                   f"NOTE 3\n{PAGE_SYS}", ident, 6, note=True)
    for bloc, mode in (("CRÉANCES — nom des clients-usagers et autres débiteurs", "d"),
                       ("DETTES — nom des fournisseurs et autres créditeurs", "c")):
        ws.cell(r, 1, bloc)
        ws.cell(r, 1).font = F_SOUS_TITRE
        r += 1
        for i, h in enumerate(["Date", "Nom", "Montant au 31/12/N",
                               "Montant au 01/01/N", "Variation en valeur",
                               "Variation en %"], start=1):
            ws.cell(r, i, h)
        style_entetes(ws, r, 1, 6)
        premiere = r + 1
        for _ in range(6):
            r += 1
            ws.cell(r, 5).value = f"=C{r}-D{r}"
            ws.cell(r, 6).value = f"=IF(D{r}=0,\"\",(C{r}-D{r})/D{r})"
        r += 1
        ws.cell(r, 2, "TOTAL DES " + ("CRÉANCES" if mode == "d" else "DETTES"))
        for col in "CDE":
            ws[f"{col}{r}"] = f"=SUM({col}{premiere}:{col}{r-1})"
        style_ligne_total(ws, r, 1, 6, cols_montant=(3, 4, 5))
        style_zone_donnees(ws, premiere, r - 1, 1, 6, cols_montant=(3, 4, 5))
        r += 1
        f_n = _c(formule_tokens(_tok("4"), mode, "BALANCE"))
        f_n1 = (_c(formule_tokens(_tok("4"), mode, "BALANCE_N1"))
                if avec_n1 else "0")
        ws.cell(r, 2, "Rappel balance (classe 4, soldes "
                      + ("débiteurs" if mode == "d" else "créditeurs")
                      + ") — contrôle")
        ws.cell(r, 3).value = f"={f_n}"
        ws.cell(r, 4).value = f"={f_n1}"
        ws.cell(r, 3).number_format = FMT_MONTANT
        ws.cell(r, 4).number_format = FMT_MONTANT
        r += 3
    ws.cell(r, 1, "Reporter la variation des créances en VB et celle des "
                  "dettes en VC du compte de résultat.")
    largeurs(ws, {"A": 12, "B": 40, "C": 18, "D": 18, "E": 16, "F": 12})


def construire_note4(wb, ident):
    ws = wb.create_sheet("NOTE 4 JOURNAL TRESORERIE")
    r = entete_smt(ws, "NOTE 4 : JOURNAL UNIQUE DE TRESORERIE",
                   f"NOTE 4\n{PAGE_SYS}", ident, 11, note=True)
    entetes = ["Dates", "Libellés", "Recettes", "Dépenses", "Solde",
               "Vent. recettes : Cotisations", "Subventions", "Autres",
               "Vent. dépenses : Achats de biens liés à l'activité",
               "Autres achats / Transport / Services extérieurs",
               "Salaires / Autres"]
    for i, h in enumerate(entetes, start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 11)
    r += 1
    ws.cell(r, 2, "Report à nouveau")
    for _ in range(14):
        r += 1
        ws.cell(r, 5).value = f"=E{r-1}+C{r}-D{r}"
    r += 1
    ws.cell(r, 2, "Solde à reporter")
    ws.cell(r, 5).value = f"=E{r-1}"
    style_zone_donnees(ws, 9, r, 1, 11,
                       cols_montant=(3, 4, 5, 6, 7, 8, 9, 10, 11))
    style_ligne_total(ws, r, 1, 11, cols_montant=(5,))
    largeurs(ws, {"A": 11, "B": 32, "C": 13, "D": 13, "E": 13, "F": 14,
                  "G": 13, "H": 11, "I": 18, "J": 20, "K": 14})


def construire_note5(wb, ident):
    ws = wb.create_sheet("NOTE 5 DOTATIONS")
    r = entete_smt(ws, "NOTE 5 : DOTATION",
                   f"NOTE 5\n{PAGE_SYS}", ident, 4, note=True)
    for i, h in enumerate(["Nom et prénoms des membres", "Nationalité",
                           "Montant", "Avec / sans droit d'entrée"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 4)
    blocs = [("Dotation non consomptible", "101,102"),
             ("Droit d'entrée", "103"),
             ("Dotation consomptible", "104")]
    data = []
    for lab, expr in blocs:
        r += 1
        ws.cell(r, 1, lab + " — rappel balance")
        ws.cell(r, 1).font = F_GRAS
        ws.cell(r, 3).value = formule_tokens(_tok(expr), "nc", "BALANCE")
        ws.cell(r, 3).number_format = FMT_MONTANT
        data.append(r)
        for _ in range(3):
            r += 1
    r += 1
    ws.cell(r, 1, "TOTAL")
    ws.cell(r, 3).value = "=" + "+".join(f"C{d}" for d in data)
    style_ligne_total(ws, r, 1, 4, cols_montant=(3,))
    style_zone_donnees(ws, 9, r - 1, 1, 4, cols_montant=(3,))
    largeurs(ws, {"A": 44, "B": 16, "C": 16, "D": 24})


def ecrire_balance_smt(wb, nom, bal):
    b = wb.create_sheet(nom_feuille(nom))
    entetes = ["Compte", "Intitulé", "Préfixe 2", "Préfixe 3", "Préfixe 4",
               "Solde final débit", "Solde final crédit"]
    b.append(entetes)
    style_entetes(b, 1, 1, len(entetes))
    for l in bal:
        c = l["compte"]
        b.append([c, l["libelle"], c[:2], c[:3], c[:4],
                  round(l["sd"], 2), round(l["sc"], 2)])
    style_zone_donnees(b, 2, b.max_row, 1, len(entetes), cols_montant=(6, 7))
    largeurs(b, {"A": 12, "B": 42, "C": 9, "D": 9, "E": 9, "F": 15, "G": 15})
    b.freeze_panes = "A2"


def detecter_anomalies_smt(bal, seuil=1.0):
    a = []
    sd = sum(l["sd"] for l in bal)
    sc = sum(l["sc"] for l in bal)
    if abs(sd - sc) > seuil:
        a.append({"gravite": "BLOQUANT", "compte": "", "libelle": "Balance entière",
                  "probleme": f"Balance déséquilibrée : débit {sd:,.2f} ≠ crédit {sc:,.2f}",
                  "solution": "Reprendre la saisie avant montage."})
    for l in bal:
        c, net = l["compte"], l["sd"] - l["sc"]
        if abs(net) <= seuil:
            continue
        cl = c[0] if c else ""
        if len(c) < 2 or cl not in "123456789":
            a.append({"gravite": "A_TRAITER", "compte": c, "libelle": l["libelle"],
                      "probleme": f"Compte non conforme au plan SYCEBNL (classe '{cl}').",
                      "solution": "Réaffecter au compte équivalent avant montage."})
        if cl == "9":
            a.append({"gravite": "INFO", "compte": c, "libelle": l["libelle"],
                      "probleme": "Compte de classe 9 (contributions volontaires en nature).",
                      "solution": "Hors bilan et hors compte de résultat par construction."})
    if any(l["compte"].startswith("4") and abs(l["sd"] - l["sc"]) > seuil for l in bal):
        a.append({"gravite": "INFO", "compte": "", "libelle": "Classe 4",
                  "probleme": "La balance porte des comptes de tiers : recettes/dépenses "
                              "incluent des montants non encaissés/décaissés (base engagement).",
                  "solution": "Laisser VB/VC à zéro dans ce cas ; ne les servir que pour "
                              "une balance de trésorerie pure (Note 3)."})
    a.append({"gravite": "INFO", "compte": "", "libelle": "SMT — seuil",
              "probleme": "Vérifier l'assujettissement : ressources annuelles ≤ 30 M FCFA "
                          "(Acte uniforme SYCEBNL, art. 5-6).",
              "solution": "Au-delà, monter le jeu du Système normal "
                          "(monter_etats_sycebnl.py) ou des projets (monter_projets.py)."})
    return a


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("balance_N")
    ap.add_argument("balance_N1", nargs="?")
    ap.add_argument("--sortie", default="etats-smt-sycebnl.xlsx")
    ap.add_argument("--entite", default="")
    ap.add_argument("--identifiant", default="")
    ap.add_argument("--exercice", default="")
    ap.add_argument("--duree", default="12")
    ap.add_argument("--adresse", default="")
    ap.add_argument("--sigle", default="")
    ap.add_argument("--ntd", default="")
    args = ap.parse_args()
    set_identite_etendue(args.adresse, args.sigle, args.ntd)

    bal, idx = lire_balance(args.balance_N)
    print(f"Balance N : {len(bal)} comptes. Colonnes repérées : {idx}")
    bal1 = None
    if args.balance_N1:
        bal1, _ = lire_balance(args.balance_N1)
        print(f"Balance N-1 : {len(bal1)} comptes.")
    avec_n1 = bal1 is not None
    set_lignes_max(max(len(bal), len(bal1 or [])) + 20)

    ident = (args.entite, args.identifiant, args.exercice, args.duree)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cr_rows = construire_cr(wb, avec_n1, ident)
    infos_bilan = construire_bilan(wb, avec_n1, ident, cr_rows["KZC"])

    construire_note1(wb, bal, ident)
    construire_note2(wb, bal, avec_n1, ident)
    construire_note3(wb, avec_n1, ident)
    construire_note4(wb, ident)
    construire_note5(wb, ident)

    ecrire_balance_smt(wb, "BALANCE", bal)
    if avec_n1:
        ecrire_balance_smt(wb, "BALANCE_N1", bal1)

    ctl = wb.create_sheet("CONTROLES")
    n = max(len(bal), 1)
    ctl.append(["Contrôle", "Valeur", "Attendu"])
    style_entetes(ctl, 1, 1, 3)
    B = q(NOM_BALANCE)
    for lab, f, att in [
            ("Total solde débit balance", f"=SUM({B}!F2:F{n+1})", ""),
            ("Total solde crédit balance", f"=SUM({B}!G2:G{n+1})", ""),
            ("Écart balance (doit être 0)", "=B2-B3", 0),
            ("Total actif (GZ)", f"={q(NOM_ACTIF)}!D{infos_bilan['GZ']}", ""),
            ("Total passif (HZ)", f"={q(NOM_PASSIF)}!D{infos_bilan['HZ']}", ""),
            ("Écart actif - passif (doit être 0)", "=B5-B6", 0),
            ("Résultat net (compte de résultat, KZC)",
             f"={q(CR_NOM)}!D{cr_rows['KZC']}", "")]:
        ctl.append([lab, f, att])
    style_zone_donnees(ctl, 2, ctl.max_row, 1, 3, cols_montant=(2,))
    largeurs(ctl, {"A": 56, "B": 20, "C": 10})

    anomalies = detecter_anomalies_smt(bal)
    an = wb.create_sheet("ANOMALIES")
    an.append(["Gravité", "Compte", "Intitulé", "Problème", "Solution proposée"])
    style_entetes(an, 1, 1, 5)
    ordre_g = {"BLOQUANT": 0, "A_TRAITER": 1, "A_VERIFIER": 2, "MINEUR": 3, "INFO": 4}
    for x in sorted(anomalies, key=lambda z: ordre_g.get(z["gravite"], 9)):
        an.append([x["gravite"], x["compte"], x["libelle"], x["probleme"], x["solution"]])
    style_zone_donnees(an, 2, max(an.max_row, 2), 1, 5)
    largeurs(an, {"A": 12, "B": 12, "C": 26, "D": 62, "E": 62})

    construire_controle_balance(wb, avec_n1, len(bal), len(bal1 or []))
    construire_couverture(wb, ident, "LIASSE SMT")
    construire_garde_etafi(
        wb, ident,
        bandeau="ETATS FINANCIERS NORMALISES\nDU SYSTEME COMPTABLE DES "
                "ENTITES A BUT NON LUCRATIF (SYCEBNL)",
        sous_bandeau="Associations, Ordres Professionnels, Fondations "
                     "et Assimilées",
        systeme="SYSTEME MINIMAL DE TRESORERIE",
        documents=["Fiche d'identification et renseignements divers",
                   "Bilan (actif et passif)",
                   "Compte de résultat",
                   "Notes annexes 1 à 5"])
    construire_fiche2(wb, ident, "EQUIPE DE L'ENTITE A BUT NON LUCRATIF")
    pa_a, pa_p = infos_bilan[NOM_ACTIF], infos_bilan[NOM_PASSIF]
    construire_bilan_paysage(
        wb, ident,
        {"feuille": NOM_ACTIF, "lig_debut": pa_a[0], "lig_fin": pa_a[1],
         "col_note": "C", "libelle": "ACTIF",
         "cols": [("EXERCICE N", "D"), ("EXERCICE N-1", "E")]},
        {"feuille": NOM_PASSIF, "lig_debut": pa_p[0], "lig_fin": pa_p[1],
         "col_note": "C", "libelle": "PASSIF",
         "cols": [("EXERCICE N", "D"), ("EXERCICE N-1", "E")]},
        titre="BILAN", page_ref=f"BILAN {PAGE_SYS}\nPAGE 1/1")
    construire_identification(wb, ident, "SYCEBNL",
                              "Système minimal de trésorerie")
    parties = [("Partie 1 : Notes sur le bilan",
                [("Note 1", "Tableau d'acquisition et de suivi du matériel, "
                            "du mobilier et des cautions"),
                 ("Note 2", "État des stocks"),
                 ("Note 3", "État des créances et des dettes non échues"),
                 ("Note 5", "Dotations")]),
               ("Partie 2 : Notes sur le compte de résultat",
                [("Note 4", "Journal unique de trésorerie")])]
    construire_fiche_notes(wb, parties, ident)
    construire_table_commentaires(wb, parties, ident)
    ordonner_feuilles(wb, [NOM_BALANCE, NOM_BALANCE_N1, "CONTROLE BALANCE",
                           "Couverture", "Garde", "Fiche 1", "Fiche 2",
                           "Bilan paysage", NOM_ACTIF, NOM_PASSIF, CR_NOM,
                           "NOTES ANNEXES",
                           "NOTE 1 IMMOBILISATIONS", "NOTE 2 STOCKS",
                           "NOTE 3 CREANCES-DETTES",
                           "NOTE 4 JOURNAL TRESORERIE", "NOTE 5 DOTATIONS",
                           "TABLE COMMENTAIRE", "CONTROLES", "ANOMALIES"])
    retirer_tirets(wb)
    appliquer_police_arial(wb)
    numeroter_pages(wb)
    wb.save(args.sortie)
    print(f"États SMT SYCEBNL écrits : {args.sortie}")
    bloquants = [a for a in anomalies if a["gravite"] in ("BLOQUANT", "A_TRAITER")]
    print(f"Anomalies : {len(anomalies)} dont {len(bloquants)} à traiter avant remise.")


if __name__ == "__main__":
    main()
