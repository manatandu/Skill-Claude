"""
theme_etafi.py — Charte graphique « ETAFI » des liasses (SYSCOHADA / SYCEBNL).

Reproduit la présentation d'une liasse fiscale professionnelle réelle
(logiciel ETAFI, dépôt DGI) : cartouche d'identification de 6 lignes en tête
de chaque page, titres d'états en Arial Black vert, bandeaux d'en-têtes bleu
pâle, quatre niveaux de lignes remarquables, format de montant comptable
(zéro affiché « - »), bordures fines/filet, et pages spéciales (Couverture,
Garde, Fiche 1, Fiche 2, CONTROLE BALANCE, NOTES ANNEXES, TABLE COMMENTAIRE,
Bilan paysage).

Palette exacte du modèle :
  CCFFFF  bandeaux d'en-têtes de colonnes        FFFFCC  lignes de rubriques
  C0C0C0  totaux intermédiaires / bandes grises  008000  totaux de section
  000080  TOTAL GENERAL / bandeau garde          003366  titres de notes, TFT
  CCFFCC  en-têtes CONTROLE BALANCE              FFCC99  titre CONTROLE BALANCE
  CCCCFF  bandes de parties TABLE COMMENTAIRE    660066  bandeau Couverture

Niveaux de lignes (style_ligne) :
  normal    ligne ordinaire (Arial 9, filets)
  rubrique  intitulé de rubrique en gras sur fond jaune pâle FFFFCC
  inter     total intermédiaire en gras sur fond gris C0C0C0
  section   total de section en gras blanc sur fond vert 008000
  general   TOTAL GENERAL en gras blanc sur fond bleu nuit 000080
  cle       ligne clef (TFT trésorerie ZA/ZF/ZG) en blanc sur 003366
  bande     bande de section sans référence (gras sur gris C0C0C0)
"""

import copy

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Palette, polices, bordures, formats
# --------------------------------------------------------------------------

C_ENTETE = "CCFFFF"       # bandeaux d'en-têtes de colonnes
C_RUBRIQUE = "FFFFCC"     # lignes de rubriques / cases codes Fiche 1
C_GRIS = "C0C0C0"         # totaux intermédiaires, bandes grises
C_SECTION = "008000"      # totaux de section (fond) et titres d'états (texte)
C_NAVY = "000080"         # TOTAL GENERAL, bandeau de garde
C_NOTE = "003366"         # titres de notes, lignes clefs TFT
C_CTRL_ENT = "CCFFCC"     # en-têtes CONTROLE BALANCE
C_CTRL_TITRE = "FFCC99"   # titre EQUILIBRE DE LA BALANCE
C_PARTIE_TC = "CCCCFF"    # bandes de parties TABLE COMMENTAIRE
C_COUV = "660066"         # bandeau de couverture
C_GARDE_TXT = "CC99FF"    # texte du bandeau navy de la garde

F_DONNEE = Font(name="Arial", size=9)
F_DONNEE_G = Font(name="Arial", size=9, bold=True)
F_BLANC_G = Font(name="Arial", size=9, bold=True, color="FFFFFF")
F_CARTOUCHE = Font(name="Arial", size=9)
F_PAGE_REF = Font(name="Arial", size=8)
F_ENTETE_COL = Font(name="Arial", size=9, bold=True)
F_TITRE_ETAT = Font(name="Arial Black", size=16, bold=True, color=C_SECTION)
F_TITRE_ETAT_M = Font(name="Arial Black", size=14, bold=True, color=C_SECTION)
F_TITRE_NOTE = Font(name="Arial Black", size=11, bold=True, color=C_NOTE)
F_VERDICT = Font(name="Arial Black", size=10, bold=True, color=C_NAVY)

FIN = Side(style="thin", color="000000")
FILET = Side(style="hair", color="000000")
MOYEN = Side(style="medium", color="000000")
POINTILLE = Side(style="dashed", color="000000")

B_DONNEE = Border(top=FILET, bottom=FILET, left=FIN, right=FIN)
B_FIN = Border(top=FIN, bottom=FIN, left=FIN, right=FIN)
B_SOULIGNE = Border(bottom=POINTILLE)

AL_CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_GAUCHE = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_DROITE = Alignment(horizontal="right", vertical="center")

# Format comptable du modèle : milliers en espaces, zéro affiché « - ».
FMT_MONTANT = '_-* #,##0\\ _€_-;\\-* #,##0\\ _€_-;_-* "-"\\ _€_-;_-@_-'
FMT_PCT = "0.0%"

# Noms normalisés des feuilles de balance dans les classeurs produits.
NOM_BALANCE = "BALANCE N"
NOM_BALANCE_N1 = "BALANCE N-1"
_ALIAS_FEUILLES = {"BALANCE": NOM_BALANCE, "BALANCE_N1": NOM_BALANCE_N1}


def nom_feuille(nom):
    """Nom réel d'une feuille (traduit les alias historiques BALANCE...)."""
    return _ALIAS_FEUILLES.get(nom, nom)


def q(nom):
    """Référence de feuille prête pour une formule : 'Bilan-Actif' devient
    'Bilan-Actif' entre apostrophes (espaces, tirets, accents obligent)."""
    import re
    nom = nom_feuille(nom)
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", nom):
        return nom
    return "'" + nom.replace("'", "''") + "'"


# --------------------------------------------------------------------------
# Identité de l'entité (cartouche)
# --------------------------------------------------------------------------

_IDENT_DEFAUT = {"entite": "", "ncc": "", "exercice": "", "duree": "12",
                 "adresse": "", "sigle": "", "ntd": ""}
_IDENT_EXTRA = {"adresse": "", "sigle": "", "ntd": ""}


def set_identite_etendue(adresse="", sigle="", ntd=""):
    """Champs d'identité additionnels du cartouche (options CLI)."""
    _IDENT_EXTRA.update({"adresse": adresse or "", "sigle": sigle or "",
                         "ntd": ntd or ""})


def normaliser_ident(ident):
    """Accepte le tuple historique (entite, identifiant, exercice, duree)
    ou un dict, et rend le dict complet du cartouche."""
    d = dict(_IDENT_DEFAUT)
    d.update(_IDENT_EXTRA)
    if isinstance(ident, dict):
        d.update({k: (v or "") for k, v in ident.items()})
    else:
        entite, identifiant, exercice, duree = ident
        d.update({"entite": entite or "", "ncc": identifiant or "",
                  "exercice": exercice or "", "duree": duree or "12"})
    return d


def _texte_exercice(exercice):
    ex = str(exercice or "").strip()
    if ex.isdigit() and len(ex) == 4:
        return f"Exercice clos le 31-12-{ex}"
    return f"Exercice clos le {ex}" if ex else "Exercice clos le"


# --------------------------------------------------------------------------
# Aides bas niveau
# --------------------------------------------------------------------------

def fusion(ws, r1, c1, r2, c2):
    if (r1, c1) != (r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2, end_column=c2)


def cadre(ws, r1, c1, r2, c2, cote=MOYEN):
    """Trace un cadre (bordure extérieure) autour d'une plage, en préservant
    les bordures intérieures existantes."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            b = copy.copy(cell.border) if cell.border else Border()
            if r == r1:
                b.top = cote
            if r == r2:
                b.bottom = cote
            if c == c1:
                b.left = cote
            if c == c2:
                b.right = cote
            cell.border = b


def hauteurs(ws, spec):
    for r, h in spec.items():
        ws.row_dimensions[r].height = h


def largeurs(ws, spec):
    for lettre, l in spec.items():
        ws.column_dimensions[lettre].width = l


# --------------------------------------------------------------------------
# Cartouche de page (lignes 1 à 6) et titres
# --------------------------------------------------------------------------

def ecrire_cartouche(ws, ident, page_ref, col_max):
    """Cartouche d'identification du modèle en tête de page :
      L1  - 0 -                                   (numéro de page, cf.
      L2                     [réf. page]           numeroter_pages)
      L3  Dénomination sociale : X
      L4  Adresse : ....          Sigle usuel : X
      L5  NCC : X     Exercice clos le X   Durée (en mois) : X
      L6  NTD : X
    Rend la ligne du titre (7)."""
    d = normaliser_ident(ident)
    ws.sheet_view.showGridLines = False
    c = ws.cell(1, 1, "- 0 -")
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "@"
    # réf. de page en haut à droite (deux dernières colonnes)
    c_ref = max(2, col_max - 1)
    fusion(ws, 2, c_ref, 2, col_max)
    c = ws.cell(2, c_ref, page_ref)
    c.font = F_PAGE_REF
    c.alignment = AL_CENTRE
    c.number_format = "@"
    for cc in range(c_ref, col_max + 1):
        ws.cell(2, cc).border = Border(top=FIN, bottom=FIN, left=FIN)
    # lignes d'identité
    c = ws.cell(3, 1, f"Dénomination sociale : {d['entite']}".rstrip())
    c.font = F_CARTOUCHE
    ws.cell(4, 1, "Adresse :").font = F_CARTOUCHE
    fin_adr = max(2, col_max - 3)
    fusion(ws, 4, 2, 4, fin_adr)
    c = ws.cell(4, 2, d["adresse"])
    c.font = F_CARTOUCHE
    for cc in range(2, fin_adr + 1):
        ws.cell(4, cc).border = B_SOULIGNE
    c = ws.cell(4, max(3, col_max - 1), "Sigle usuel :")
    c.font = F_CARTOUCHE
    c.alignment = AL_DROITE
    c = ws.cell(4, col_max, d["sigle"])
    c.font = F_CARTOUCHE
    c.border = B_SOULIGNE
    c = ws.cell(5, 1, f"N° de compte contribuable (NCC) : {d['ncc']}".rstrip())
    c.font = F_CARTOUCHE
    c = ws.cell(5, max(3, col_max - 3), _texte_exercice(d["exercice"]))
    c.font = F_CARTOUCHE
    c = ws.cell(5, max(3, col_max - 1), f"Durée (en mois) : {d['duree']}")
    c.font = F_CARTOUCHE
    c = ws.cell(6, 1, f"N° de télédéclarant (NTD) : {d['ntd']}".rstrip())
    c.font = F_CARTOUCHE
    hauteurs(ws, {1: 12, 2: 26, 3: 15, 4: 15, 5: 15, 6: 15, 7: 28})
    return 7


def titre_etat(ws, texte, col_min, col_max, row=7, taille=16):
    """Titre d'état : Arial Black vert, centré, filet inférieur moyen."""
    fusion(ws, row, col_min, row, col_max)
    c = ws.cell(row, col_min, texte)
    c.font = F_TITRE_ETAT if taille >= 16 else F_TITRE_ETAT_M
    c.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(col_min, col_max + 1):
        ws.cell(row, cc).border = Border(bottom=MOYEN)
    return row + 1


def titre_note(ws, texte, col_max, row=7):
    """Titre de note annexe : Arial Black 11 bleu nuit, centré."""
    fusion(ws, row, 1, row, col_max)
    c = ws.cell(row, 1, texte)
    c.font = F_TITRE_NOTE
    c.alignment = AL_CENTRE
    return row + 1


# --------------------------------------------------------------------------
# Bandeaux d'en-têtes et niveaux de lignes
# --------------------------------------------------------------------------

def entetes_bande(ws, row_min, row_max, col_min, col_max):
    """Bandeau d'en-têtes de colonnes : fond CCFFFF, Arial 9 gras, cadre
    extérieur moyen, séparations intérieures fines."""
    for r in range(row_min, row_max + 1):
        for c in range(col_min, col_max + 1):
            cell = ws.cell(r, c)
            cell.font = F_ENTETE_COL
            cell.fill = PatternFill("solid", fgColor=C_ENTETE)
            cell.alignment = AL_CENTRE
            cell.border = B_FIN
    cadre(ws, row_min, col_min, row_max, col_max, MOYEN)


_NIVEAUX = {
    "rubrique": (C_RUBRIQUE, F_DONNEE_G),
    "inter": (C_GRIS, F_DONNEE_G),
    "section": (C_SECTION, F_BLANC_G),
    "general": (C_NAVY, F_BLANC_G),
    "cle": (C_NOTE, F_BLANC_G),
    "bande": (C_GRIS, F_DONNEE_G),
}


def style_ligne(ws, row, col_min, col_max, niveau="normal", cols_montant=(),
                col_ref=None):
    """Habille une ligne de tableau selon son niveau (cf. en-tête de module).
    `col_ref` : colonne du code REF (centrée, gras si ligne remarquable)."""
    fond, police = _NIVEAUX.get(niveau, (None, F_DONNEE))
    for c in range(col_min, col_max + 1):
        cell = ws.cell(row, c)
        cell.border = B_DONNEE
        if fond and c != col_ref:
            cell.fill = PatternFill("solid", fgColor=fond)
        cell.font = police if c != col_ref else (
            F_DONNEE_G if niveau != "normal" else F_DONNEE)
        if c in cols_montant:
            cell.number_format = FMT_MONTANT
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = Alignment(vertical="center")
        elif c == col_ref:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif cell.alignment is None or cell.alignment.horizontal is None:
            cell.alignment = AL_GAUCHE
    if niveau == "general":
        for c in range(col_min, col_max + 1):
            b = copy.copy(ws.cell(row, c).border)
            b.bottom = MOYEN
            ws.cell(row, c).border = b


# --------------------------------------------------------------------------
# Compatibilité : aides historiques réécrites dans la charte ETAFI
# --------------------------------------------------------------------------

def style_titre(ws, cell_range, texte):
    debut, fin = cell_range.split(":")
    r = int("".join(ch for ch in debut if ch.isdigit()))
    c1 = ws[debut].column
    c2 = ws[fin].column
    titre_etat(ws, texte, c1, c2, row=r, taille=14)


def style_entetes(ws, row, col_min, col_max):
    entetes_bande(ws, row, row, col_min, col_max)


def style_zone_donnees(ws, row_min, row_max, col_min, col_max,
                       cols_montant=(), bandes=False):
    for r in range(row_min, row_max + 1):
        gras = any(ws.cell(r, c).font is not None and ws.cell(r, c).font.bold
                   for c in range(col_min, col_max + 1))
        for c in range(col_min, col_max + 1):
            cell = ws.cell(r, c)
            cell.border = B_DONNEE
            cell.font = F_DONNEE_G if gras else F_DONNEE
            if c in cols_montant:
                cell.number_format = FMT_MONTANT


def style_ligne_total(ws, row, col_min, col_max, cols_montant=()):
    style_ligne(ws, row, col_min, col_max, "inter", cols_montant)


def format_montants(ws, cellules):
    for ref in cellules:
        ws[ref].number_format = FMT_MONTANT


# --------------------------------------------------------------------------
# Nettoyage typographique, polices, numérotation
# --------------------------------------------------------------------------

def retirer_tirets(wb):
    """Remplace tirets cadratins (—) et demi-cadratins (–) par un tiret
    simple dans toutes les cellules texte : aucun « — » dans les livrables."""
    from openpyxl.cell.cell import MergedCell
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c, MergedCell):
                    continue
                v = c.value
                if isinstance(v, str) and ("—" in v or "–" in v):
                    v = v.replace(" — ", " - ").replace("—", "-")
                    v = v.replace(" – ", " - ").replace("–", "-")
                    c.value = v


def appliquer_police_arial(wb):
    """Police de corps Arial 9 partout : les Calibri par défaut restants
    passent en Arial (Arial Black / Times conservés, tailles >= 12 aussi)."""
    from openpyxl.cell.cell import MergedCell
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c, MergedCell) or c.value is None:
                    continue
                f = c.font
                if f is None or f.name in (None, "Calibri"):
                    taille = 9 if (f is None or f.size is None
                                   or f.size <= 11) else f.size
                    c.font = Font(name="Arial", size=taille,
                                  bold=bool(f and f.bold),
                                  italic=bool(f and f.italic),
                                  color=f.color if f else None)


def numeroter_pages(wb):
    """Numérote « - n - » (cellule A1) les pages porteuses du cartouche,
    dans l'ordre final des feuilles."""
    n = 0
    for ws in wb.worksheets:
        try:
            v = ws["A1"].value
        except Exception:
            continue
        if isinstance(v, str) and v.strip().startswith("-"):
            n += 1
            ws["A1"] = f"- {n} -"


def ordonner_feuilles(wb, ordre):
    """Réordonne les feuilles : celles citées dans `ordre` d'abord (dans cet
    ordre), les autres ensuite dans leur ordre actuel."""
    ordre = [nom_feuille(n) for n in ordre]
    pos = {nom: i for i, nom in enumerate(ordre)}
    base = len(ordre)
    actuels = {ws.title: i for i, ws in enumerate(wb._sheets)}
    wb._sheets.sort(key=lambda w: pos.get(w.title, base + actuels[w.title]))


# --------------------------------------------------------------------------
# Couverture
# --------------------------------------------------------------------------

def construire_couverture(wb, ident, titre_liasse, pays=""):
    """Page de couverture : cadre, bande verticale bleu nuit, nom de
    l'entité, bandeau violet « LIASSE ... »."""
    d = normaliser_ident(ident)
    ws = wb.create_sheet("Couverture")
    ws.sheet_view.showGridLines = False
    largeurs(ws, {"A": 1.3, "B": 0.8, "C": 1.3, "D": 8.7, "E": 10.5,
                  "F": 11.5, "G": 11.5, "H": 13.2, "I": 11.5, "J": 8.7,
                  "K": 11.5, "L": 1.3, "M": 0.8})
    hauteurs(ws, {1: 4, 2: 4, 31: 33, 53: 5})
    for r in range(3, 53):
        if r != 31:
            ws.row_dimensions[r].height = 13
    cadre(ws, 2, 2, 52, 13, MOYEN)
    for r in range(3, 19):
        for c in (3, 4):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=C_NOTE)
    if pays:
        fusion(ws, 4, 6, 4, 9)
        c = ws.cell(4, 6, pays)
        c.font = Font(name="Times New Roman", size=12, bold=True)
        c.alignment = AL_CENTRE
    fusion(ws, 26, 5, 29, 10)
    c = ws.cell(26, 5, d["entite"])
    c.font = Font(name="Arial", size=22, bold=True)
    c.alignment = AL_CENTRE
    fusion(ws, 31, 5, 31, 11)
    c = ws.cell(31, 5, titre_liasse)
    c.font = Font(name="Arial Black", size=20, bold=True, color="FFFFFF")
    c.alignment = AL_CENTRE
    for cc in range(5, 12):
        ws.cell(31, cc).fill = PatternFill("solid", fgColor=C_COUV)
    fusion(ws, 34, 5, 34, 11)
    c = ws.cell(34, 5, _texte_exercice(d["exercice"]))
    c.font = Font(name="Arial", size=12, bold=True)
    c.alignment = AL_CENTRE
    return ws


# --------------------------------------------------------------------------
# Garde
# --------------------------------------------------------------------------

def construire_garde_etafi(wb, ident, bandeau, sous_bandeau, systeme,
                           documents, lignes_admin=(), centre_depot=""):
    """Page de garde du modèle : en-tête administratif, bande grise du
    référentiel, bandeau bleu nuit, désignation de l'entité, système,
    liste des documents déposés, zone réservée à l'administration."""
    d = normaliser_ident(ident)
    ws = wb.create_sheet("Garde")
    ws.sheet_view.showGridLines = False
    largeurs(ws, {"A": 1.0, "B": 19.8, "C": 11.3, "D": 11.0, "E": 3.2,
                  "F": 5.2, "G": 5.2, "H": 3.2, "I": 11.0, "J": 11.3,
                  "K": 19.8, "L": 1.0})
    hauteurs(ws, {2: 20, 3: 20, 4: 20, 8: 12, 9: 14, 12: 44, 13: 22,
                  16: 18, 20: 16, 22: 18, 26: 18, 28: 18, 30: 14, 31: 17,
                  32: 30, 33: 14})
    r = 2
    for i, ligne in enumerate(lignes_admin):
        fusion(ws, r, 3, r, 10)
        c = ws.cell(r, 3, ligne)
        c.font = Font(name="Arial", size=14 if i == 0 else 10, bold=True)
        c.alignment = AL_CENTRE
        r += 1
    if centre_depot:
        fusion(ws, 8, 3, 8, 10)
        c = ws.cell(8, 3, "CENTRE DE DEPOT DE")
        c.font = Font(name="Arial", size=10)
        c.alignment = AL_CENTRE
        fusion(ws, 9, 3, 9, 10)
        c = ws.cell(9, 3, centre_depot)
        c.font = Font(name="Arial", size=11, bold=True)
        c.alignment = AL_CENTRE
    fusion(ws, 12, 2, 12, 11)
    c = ws.cell(12, 2, bandeau)
    c.font = Font(name="Arial", size=14, bold=True)
    c.fill = PatternFill("solid", fgColor=C_GRIS)
    c.alignment = AL_CENTRE
    for cc in range(2, 12):
        ws.cell(12, cc).fill = PatternFill("solid", fgColor=C_GRIS)
        ws.cell(12, cc).border = Border(top=MOYEN, bottom=FIN)
    fusion(ws, 13, 2, 13, 11)
    c = ws.cell(13, 2, sous_bandeau)
    c.font = Font(name="Arial", size=17, bold=True, color=C_GARDE_TXT)
    c.alignment = AL_CENTRE
    for cc in range(2, 12):
        ws.cell(13, cc).fill = PatternFill("solid", fgColor=C_NAVY)
        ws.cell(13, cc).border = Border(bottom=MOYEN)
    fusion(ws, 16, 3, 16, 10)
    c = ws.cell(16, 3, _texte_exercice(d["exercice"]) +
                f"   -   Durée (en mois) : {d['duree']}")
    c.font = Font(name="Arial", size=14, bold=True)
    c.alignment = AL_CENTRE
    fusion(ws, 20, 2, 20, 11)
    c = ws.cell(20, 2, "DESIGNATION DE L'ENTITE")
    c.font = Font(name="Arial", size=12, bold=True)
    c.alignment = AL_CENTRE
    for cc in range(2, 12):
        ws.cell(20, cc).border = Border(top=FIN, bottom=FIN)
    champs = [(22, "DENOMINATION SOCIALE :", d["entite"]),
              (24, "SIGLE USUEL :", d["sigle"]),
              (26, "ADRESSE COMPLETE :", d["adresse"]),
              (28, "N° de compte contribuable (NCC) :", d["ncc"]),
              (30, "N° de télédéclarant (NTD) :", d["ntd"])]
    for r, lab, val in champs:
        c = ws.cell(r, 2, lab)
        c.font = Font(name="Arial", size=9, bold=(r >= 28))
        fusion(ws, r, 5, r, 10)
        c = ws.cell(r, 5, val)
        c.font = Font(name="Arial", size=12, bold=True)
        c.alignment = AL_CENTRE
        for cc in range(5, 11):
            ws.cell(r, cc).border = B_SOULIGNE
    fusion(ws, 32, 2, 32, 11)
    c = ws.cell(32, 2, systeme)
    c.font = Font(name="Arial", size=16, bold=True)
    c.alignment = AL_CENTRE
    for cc in range(2, 12):
        ws.cell(32, cc).border = Border(top=POINTILLE)
    ws.cell(34, 2, "Documents déposés").font = Font(name="Arial", size=10,
                                                    bold=True)
    ws.cell(34, 8, "Réservé à l'administration").font = Font(
        name="Arial", size=10, bold=True)
    r = 35
    for doc in documents:
        fusion(ws, r, 2, r, 5)
        c = ws.cell(r, 2, doc)
        c.font = F_DONNEE
        for cc in range(2, 6):
            ws.cell(r, cc).border = B_FIN
        c = ws.cell(r, 6, "X")
        c.font = Font(name="Arial Black", size=8, bold=True)
        c.alignment = AL_CENTRE
        c.border = B_FIN
        r += 1
    zone = [("Date de dépôt", 1),
            ("", 3),
            ("Nom de l'agent ayant réceptionné le dépôt", 1),
            ("", 3),
            ("Signature de l'agent et cachet du service", 1),
            ("", 3)]
    rz = 35
    for texte, h in zone:
        fusion(ws, rz, 8, rz + h - 1, 11)
        if texte:
            c = ws.cell(rz, 8, texte)
            c.font = Font(name="Arial", size=8)
            c.alignment = Alignment(horizontal="center", vertical="top")
        for rr in range(rz, rz + h):
            for cc in range(8, 12):
                ws.cell(rr, cc).border = B_FIN
        rz += h
    r = max(r, rz) + 1
    for lab in ("Nombre de pages déposées par exemplaire :",
                "Nombre d'exemplaires déposés :"):
        fusion(ws, r, 2, r, 5)
        c = ws.cell(r, 2, lab)
        c.font = F_DONNEE
        for cc in range(2, 7):
            ws.cell(r, cc).border = B_FIN
        ws.cell(r, 6).border = B_FIN
        r += 1
    return ws


# --------------------------------------------------------------------------
# Fiche 1 : identification et renseignements divers
# --------------------------------------------------------------------------

def construire_fiche1(wb, ident, referentiel, systeme, page_ref="FICHE 1"):
    """Fiche d'identification à cases codes (ZA, ZB, ...) sur fond FFFFCC,
    adaptée du modèle : champs connus pré-remplis, le reste à compléter."""
    d = normaliser_ident(ident)
    ws = wb.create_sheet("Fiche 1")
    NB = 10
    ecrire_cartouche(ws, ident, page_ref, NB)
    fusion(ws, 7, 1, 7, NB)
    c = ws.cell(7, 1, "FICHE D'IDENTIFICATION ET RENSEIGNEMENTS DIVERS")
    c.font = Font(name="Arial", size=11, bold=True, color=C_SECTION)
    c.alignment = AL_CENTRE
    fusion(ws, 8, 1, 8, NB)
    c = ws.cell(8, 1, f"{referentiel} - {systeme}")
    c.font = Font(name="Arial", size=9, bold=True)
    c.alignment = AL_CENTRE
    ex = str(d["exercice"]).strip()
    debut = f"01-01-{ex}" if ex.isdigit() and len(ex) == 4 else ""
    fin = f"31-12-{ex}" if ex.isdigit() and len(ex) == 4 else ex
    champs = [
        ("ZA", "EXERCICE COMPTABLE", f"DU : {debut}    AU : {fin}"),
        ("ZB", "DATE D'ARRETE EFFECTIF DES COMPTES", ""),
        ("ZC", "EXERCICE PRECEDENT CLOS LE", ""),
        ("ZD", "DUREE DE L'EXERCICE PRECEDENT (EN MOIS)", ""),
        ("ZE", "N° REGISTRE (RCCM, F92, CONVENTION...) ET GREFFE", ""),
        ("ZF", "N° REPERTOIRE DES ENTITES", ""),
        ("ZG", "N° DE CAISSE SOCIALE", ""),
        ("ZH", "N° CODE IMPORTATEUR", ""),
        ("ZI", "CODE ACTIVITE PRINCIPALE", ""),
        ("ZJ", "DESIGNATION DE L'ENTITE ET SIGLE",
         (d["entite"] + (f"  ({d['sigle']})" if d["sigle"] else "")).strip()),
        ("ZK", "N° DE TELEPHONE, ADRESSE E-MAIL, BOITE POSTALE, VILLE", ""),
        ("ZL", "ADRESSE GEOGRAPHIQUE COMPLETE (IMMEUBLE, RUE, QUARTIER, "
               "VILLE, PAYS)", d["adresse"]),
        ("ZM", "DESIGNATION PRECISE DE L'ACTIVITE PRINCIPALE EXERCEE", ""),
        ("ZN", "% DE CAPACITE DE PRODUCTION UTILE", ""),
        ("ZO", "NOM, ADRESSE, TELEPHONE, E-MAIL ET QUALITE DE LA PERSONNE "
               "A CONTACTER EN CAS DE DEMANDE", ""),
        ("ZP", "NOM, ADRESSE, TELEPHONE ET E-MAIL DU SALARIE OU DU "
               "PROFESSIONNEL COMPTABLE AYANT ETABLI LES ETATS FINANCIERS",
         ""),
        ("ZQ", "NOM, ADRESSE, TELEPHONE, E-MAIL ET N° D'INSCRIPTION A "
               "L'ORDRE DE L'EXPERT-COMPTABLE AYANT DELIVRE L'ATTESTATION "
               "DE VISA", ""),
        ("ZR", "NOM, ADRESSE, TELEPHONE, E-MAIL ET N° D'INSCRIPTION DU "
               "COMMISSAIRE AUX COMPTES, LE CAS ECHEANT", ""),
        ("ZS", "ETATS FINANCIERS APPROUVES PAR L'ORGANE COMPETENT "
               "(OUI / NON)", ""),
        ("ZT", "NOM DU SIGNATAIRE DES ETATS FINANCIERS", ""),
        ("ZU", "QUALITE DU SIGNATAIRE DES ETATS FINANCIERS", ""),
        ("ZV", "DATE DE SIGNATURE", ""),
        ("ZW", "DOMICILIATIONS BANCAIRES (BANQUE ET NUMERO DE COMPTE)", ""),
    ]
    r = 9
    for code, lab, val in champs:
        r += 1
        c = ws.cell(r, 1, code)
        c.font = Font(name="Arial", size=8, bold=True)
        c.fill = PatternFill("solid", fgColor=C_RUBRIQUE)
        c.alignment = AL_CENTRE
        c.border = B_FIN
        fusion(ws, r, 2, r, 6)
        c = ws.cell(r, 2, lab)
        c.font = Font(name="Arial", size=8)
        c.alignment = AL_GAUCHE
        fusion(ws, r, 7, r, NB)
        c = ws.cell(r, 7, val)
        c.font = Font(name="Arial", size=8, bold=True)
        c.alignment = AL_GAUCHE
        for cc in range(2, NB + 1):
            ws.cell(r, cc).border = Border(bottom=FILET)
        ws.row_dimensions[r].height = 24
    r += 2
    fusion(ws, r, 2, r, 5)
    ws.cell(r, 2, "Signature").font = Font(name="Arial", size=8)
    for cc in range(2, 6):
        ws.cell(r, cc).border = Border(top=FIN)
    largeurs(ws, {"A": 5, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12,
                  "G": 11, "H": 11, "I": 11, "J": 11})
    return ws


def construire_fiche2(wb, ident, sous_titre, page_ref="FICHE 2", lignes=20):
    """Fiche 2 : tableau de l'équipe / des dirigeants de l'entité."""
    ws = wb.create_sheet("Fiche 2")
    NB = 10
    ecrire_cartouche(ws, ident, page_ref, NB)
    fusion(ws, 7, 1, 7, NB)
    c = ws.cell(7, 1, "FICHE D'IDENTIFICATION ET RENSEIGNEMENTS DIVERS")
    c.font = Font(name="Arial", size=12, bold=True, color=C_SECTION)
    c.alignment = AL_CENTRE
    fusion(ws, 8, 1, 8, NB)
    c = ws.cell(8, 1, sous_titre)
    c.font = Font(name="Arial", size=11, bold=True)
    c.alignment = AL_CENTRE
    entetes = [("Nom et Prénoms", 1, 2), ("Nationalité", 3, 3),
               ("Autres nationalités (à préciser) (2)", 4, 4),
               ("Qualité", 5, 5), ("N° d'identification fiscale", 6, 7),
               ("Adresse (BP, ville, pays, adresse géographique et "
                "adresse e-mail)", 8, 10)]
    r = 10
    for lab, c1, c2 in entetes:
        fusion(ws, r, c1, r, c2)
        ws.cell(r, c1, lab)
    entetes_bande(ws, r, r, 1, NB)
    ws.row_dimensions[r].height = 34
    for i in range(lignes):
        r += 1
        ws.row_dimensions[r].height = 26
        for lab, c1, c2 in entetes:
            fusion(ws, r, c1, r, c2)
        for cc in range(1, NB + 1):
            ws.cell(r, cc).border = B_FIN
    r += 2
    c = ws.cell(r, 1, "(2) Mentionner les autres nationalités le cas "
                      "échéant.")
    c.font = Font(name="Arial", size=8)
    largeurs(ws, {"A": 12, "B": 16, "C": 13, "D": 13, "E": 13, "F": 8,
                  "G": 8, "H": 11, "I": 11, "J": 11})
    return ws


# --------------------------------------------------------------------------
# CONTROLE BALANCE
# --------------------------------------------------------------------------

def construire_controle_balance(wb, avec_n1, n_lignes, n_lignes_n1,
                                cols=("F", "G", "H", "I")):
    """Feuille CONTROLE BALANCE (équilibre de la balance) : sommes des
    colonnes de la balance et verdicts Equilibre / Déséquilibre.
    `cols` : lettres (solde final débit, solde final crédit, mouvements
    débit, mouvements crédit) dans les feuilles de balance."""
    ws = wb.create_sheet("CONTROLE BALANCE")
    ws.sheet_view.showGridLines = False
    largeurs(ws, {"A": 22.5, "B": 20.7, "C": 20.7, "D": 20.7, "E": 20.7})
    sfd, sfc, mvd, mvc = cols
    fusion(ws, 1, 1, 1, 5)
    c = ws.cell(1, 1, "EQUILIBRE DE LA BALANCE")
    c.font = Font(name="Arial", size=20, bold=True)
    c.alignment = AL_CENTRE
    for cc in range(1, 6):
        ws.cell(1, cc).fill = PatternFill("solid", fgColor=C_CTRL_TITRE)
    ws.row_dimensions[1].height = 26

    def bloc(r0, nom, feuille, n):
        fusion(ws, r0, 1, r0 + 2, 1)
        c = ws.cell(r0, 1, nom)
        c.font = Font(name="Arial", size=14, bold=True)
        c.fill = PatternFill("solid", fgColor=C_GRIS)
        c.alignment = AL_CENTRE
        heads = [("Mouvements Débit", mvd, False),
                 ("Mouvements Crédit", mvc, False),
                 ("Solde Final Débit", sfd, True),
                 ("Solde Final Crédit", sfc, True)]
        for i, (lab, lettre, vert) in enumerate(heads):
            cc = 2 + i
            c = ws.cell(r0, cc, lab)
            c.font = Font(name="Arial", size=10, bold=True)
            c.alignment = AL_CENTRE
            if vert:
                c.fill = PatternFill("solid", fgColor=C_CTRL_ENT)
            c.border = B_FIN
            c = ws.cell(r0 + 1, cc,
                        f"=SUM({q(feuille)}!{lettre}2:{lettre}{n})")
            c.font = Font(name="Arial", size=11, bold=True)
            c.fill = PatternFill("solid", fgColor=C_GRIS)
            c.number_format = "#,##0"
            c.border = B_DONNEE
        for c1, c2 in ((2, 3), (4, 5)):
            fusion(ws, r0 + 2, c1, r0 + 2, c2)
            la = get_column_letter(c1)
            lb = get_column_letter(c2)
            rr = r0 + 1
            c = ws.cell(r0 + 2, c1,
                        f'=IF(ROUND({la}{rr}-{lb}{rr},0)=0,"Equilibre",'
                        f'"Déséquilibre : écart de "'
                        f'&TEXT({la}{rr}-{lb}{rr},"#,##0"))')
            c.font = F_VERDICT
            c.alignment = AL_CENTRE
            for cc in range(c1, c2 + 1):
                ws.cell(r0 + 2, cc).border = B_DONNEE
        ws.row_dimensions[r0].height = 50
        ws.row_dimensions[r0 + 1].height = 21
        ws.row_dimensions[r0 + 2].height = 54
        cadre(ws, r0, 1, r0 + 2, 5, MOYEN)

    bloc(2, "BALANCE N", NOM_BALANCE, max(2, n_lignes + 1))
    if avec_n1:
        bloc(5, "BALANCE N-1", NOM_BALANCE_N1, max(2, n_lignes_n1 + 1))
    return ws


# --------------------------------------------------------------------------
# NOTES ANNEXES (fiche récapitulative) et TABLE COMMENTAIRE
# --------------------------------------------------------------------------

def construire_fiche_notes_etafi(wb, parties, ident, page_ref="NOTES ANNEXES",
                                 note_pied=None):
    """Feuille NOTES ANNEXES : fiche récapitulative des notes présentées.
    `parties` : liste de (titre_partie, [(numero, intitule), ...])."""
    ws = wb.create_sheet("NOTES ANNEXES")
    NB = 10
    ecrire_cartouche(ws, ident, page_ref, NB)
    fusion(ws, 7, 1, 7, NB)
    c = ws.cell(7, 1, "FICHE RECAPITULATIVE NOTES ANNEXES PRESENTEES (1)")
    c.font = F_TITRE_NOTE
    c.alignment = AL_CENTRE
    ws.row_dimensions[7].height = 22
    r = 8
    ws.cell(r, 1, "NOTES")
    fusion(ws, r, 2, r, 8)
    ws.cell(r, 2, "INTITULES")
    ws.cell(r, 9, "A (2)")
    ws.cell(r, 10, "N/A (2)")
    entetes_bande(ws, r, r, 1, NB)
    ws.row_dimensions[r].height = 19
    for titre_partie, lignes in parties:
        r += 1
        fusion(ws, r, 1, r, NB)
        c = ws.cell(r, 1, titre_partie)
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = AL_CENTRE
        for cc in range(1, NB + 1):
            ws.cell(r, cc).fill = PatternFill("solid", fgColor=C_GRIS)
            ws.cell(r, cc).border = B_FIN
        ws.row_dimensions[r].height = 16
        for numero, intitule in lignes:
            r += 1
            c = ws.cell(r, 1, numero)
            c.font = Font(name="Arial", size=10)
            c.alignment = AL_CENTRE
            fusion(ws, r, 2, r, 8)
            c = ws.cell(r, 2, intitule)
            c.font = Font(name="Arial", size=10)
            c.alignment = AL_GAUCHE
            for cc in range(1, NB + 1):
                ws.cell(r, cc).border = B_FIN
    r += 2
    fusion(ws, r, 1, r, NB)
    c = ws.cell(r, 1, note_pied or
                "(1) Les notes non documentées ne doivent pas être jointes "
                "aux états financiers ; dans une note, les lignes non "
                "chiffrées doivent être supprimées.")
    c.font = Font(name="Arial", size=8)
    c.alignment = AL_GAUCHE
    ws.row_dimensions[r].height = 30
    r += 1
    fusion(ws, r, 1, r, NB)
    c = ws.cell(r, 1, "(2) A : applicable ; N/A : non applicable. Cocher "
                      "la colonne correspondante pour chaque note.")
    c.font = Font(name="Arial", size=8)
    c.alignment = AL_GAUCHE
    largeurs(ws, {"A": 11.5, "B": 19, "C": 10, "D": 8, "E": 17, "F": 9.5,
                  "G": 8, "H": 21.5, "I": 6.3, "J": 6.3})
    return ws


def construire_table_commentaires(wb, parties, ident, par_page=4):
    """Feuille TABLE COMMENTAIRE : pour chaque note, sa référence (case
    grise) et une zone haute de commentaire libre ; bandes de parties
    lilas, blocs répétés par page comme dans le modèle."""
    d = normaliser_ident(ident)
    ws = wb.create_sheet("TABLE COMMENTAIRE")
    ws.sheet_view.showGridLines = False
    NB = 8
    largeurs(ws, {"A": 11.2, "B": 14.8, "C": 11.5, "D": 19.8, "E": 14.5,
                  "F": 10.3, "G": 13.5, "H": 6.3})
    plates = []
    for titre_partie, lignes in parties:
        premier = True
        for numero, intitule in lignes:
            plates.append((titre_partie if premier else None,
                           numero, intitule))
            premier = False
    nb_pages = max(1, (len(plates) + par_page - 1) // par_page)

    r = 1
    page = 0
    for i, (partie, numero, intitule) in enumerate(plates):
        if i % par_page == 0:
            page += 1
            fusion(ws, r, 1, r, 6)
            c = ws.cell(r, 1, "TABLE DES COMMENTAIRES")
            c.font = Font(name="Arial", size=11, bold=True, color=C_SECTION)
            c.alignment = AL_CENTRE
            fusion(ws, r, 7, r, 8)
            c = ws.cell(r, 7, f"COMMENTAIRES\nPAGE {page}/{nb_pages}")
            c.font = F_PAGE_REF
            c.alignment = AL_CENTRE
            for cc in range(7, 9):
                ws.cell(r, cc).border = Border(top=FIN, bottom=FIN, left=FIN)
            ws.row_dimensions[r].height = 26
            c = ws.cell(r + 1, 1, f"Dénomination sociale : {d['entite']}")
            c.font = Font(name="Arial", size=8)
            c = ws.cell(r + 2, 1, "Adresse : " + d["adresse"])
            c.font = Font(name="Arial", size=8)
            c = ws.cell(r + 2, 6, "Sigle usuel : " + d["sigle"])
            c.font = Font(name="Arial", size=8)
            c = ws.cell(r + 3, 1,
                        f"N° de compte contribuable (NCC) : {d['ncc']}")
            c.font = Font(name="Arial", size=8)
            c = ws.cell(r + 3, 5, _texte_exercice(d["exercice"]))
            c.font = Font(name="Arial", size=8)
            c = ws.cell(r + 3, 7, f"Durée (en mois) : {d['duree']}")
            c.font = Font(name="Arial", size=8)
            c = ws.cell(r + 4, 1, f"N° de télédéclarant (NTD) : {d['ntd']}")
            c.font = Font(name="Arial", size=8)
            r += 5
        if partie:
            fusion(ws, r, 1, r, NB)
            c = ws.cell(r, 1, partie)
            c.font = Font(name="Arial Black", size=10, bold=True)
            c.alignment = AL_CENTRE
            for cc in range(1, NB + 1):
                ws.cell(r, cc).fill = PatternFill("solid",
                                                  fgColor=C_PARTIE_TC)
                ws.cell(r, cc).border = Border(bottom=FIN)
            ws.row_dimensions[r].height = 17
            r += 1
        c = ws.cell(r, 1, numero)
        c.font = Font(name="Arial", size=8)
        c.fill = PatternFill("solid", fgColor=C_GRIS)
        c.alignment = AL_CENTRE
        c.border = B_FIN
        fusion(ws, r, 2, r, NB)
        c = ws.cell(r, 2, intitule)
        c.font = Font(name="Arial", size=8, bold=True)
        c.alignment = AL_GAUCHE
        for cc in range(2, NB + 1):
            ws.cell(r, cc).border = B_FIN
        ws.row_dimensions[r].height = 16
        r += 1
        fusion(ws, r, 1, r, NB)
        for cc in range(1, NB + 1):
            ws.cell(r, cc).border = B_FIN
        ws.cell(r, 1).alignment = Alignment(horizontal="left",
                                            vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 110
        r += 1
    return ws


# --------------------------------------------------------------------------
# Bilan paysage (actif et passif côte à côte)
# --------------------------------------------------------------------------

_FONDS_NIVEAU = {C_RUBRIQUE: "rubrique", C_GRIS: "inter",
                 C_SECTION: "section", C_NAVY: "general", C_NOTE: "cle"}


def _niveau_depuis_fond(cell):
    f = cell.fill
    if f is None or f.fill_type != "solid":
        return "normal"
    rgb = f.fgColor.rgb
    if not isinstance(rgb, str):
        return "normal"
    return _FONDS_NIVEAU.get(rgb[-6:].upper(), "normal")


def construire_bilan_paysage(wb, ident, cote_actif, cote_passif,
                             titre="BILAN", page_ref=None,
                             nom="Bilan paysage"):
    """Bilan sur une page : actif à gauche, passif à droite, par liens vers
    les feuilles Bilan-Actif / Bilan-Passif (aucune re-saisie).

    Chaque côté est un dict : {"feuille": nom, "lig_debut": r1, "lig_fin":
    r2, "cols": [(entete, lettre_source), ...], "col_note": lettre|None,
    "libelle": "ACTIF"/"PASSIF"}. Les niveaux (rubrique, totaux) sont relus
    depuis le fond des cellules sources."""
    ws = wb.create_sheet(nom)
    nca = len(cote_actif["cols"])
    ncp = len(cote_passif["cols"])
    # colonnes : REF | LIBELLE | NOTE | montants actif || REF | LIBELLE |
    # NOTE | montants passif
    ca0 = 1
    cp0 = 3 + nca + 1
    col_max = cp0 + 2 + ncp
    if page_ref is None:
        page_ref = f"{titre}\nPAGE 1/1"
    ecrire_cartouche(ws, ident, page_ref, col_max)
    titre_etat(ws, titre, 3, col_max - 2, row=7, taille=16)

    def entetes_cote(c0, cote):
        ws.cell(8, c0, "REF")
        ws.cell(8, c0 + 1, cote["libelle"])
        ws.cell(8, c0 + 2, "NOTE")
        fusion(ws, 8, c0, 9, c0)
        fusion(ws, 8, c0 + 1, 9, c0 + 1)
        fusion(ws, 8, c0 + 2, 9, c0 + 2)
        for i, (entete, _lettre) in enumerate(cote["cols"]):
            ws.cell(8, c0 + 3 + i, entete)
            fusion(ws, 8, c0 + 3 + i, 9, c0 + 3 + i)
        entetes_bande(ws, 8, 9, c0, c0 + 2 + len(cote["cols"]))

    entetes_cote(ca0, cote_actif)
    entetes_cote(cp0, cote_passif)
    ws.row_dimensions[8].height = 24
    ws.row_dimensions[9].height = 24

    def ecrire_cote(c0, cote):
        src = wb[cote["feuille"]]
        qsrc = q(cote["feuille"])
        r_out = 10
        for r in range(cote["lig_debut"], cote["lig_fin"] + 1):
            ref = src.cell(r, 1).value
            lib = src.cell(r, 2).value
            if lib is None and ref is None:
                continue
            niveau = _niveau_depuis_fond(src.cell(r, 2))
            ws.cell(r_out, c0, ref)
            c = ws.cell(r_out, c0 + 1, lib)
            c.alignment = AL_GAUCHE
            if cote.get("col_note"):
                ws.cell(r_out, c0 + 2,
                        src[f"{cote['col_note']}{r}"].value)
            monts = []
            for i, (_entete, lettre) in enumerate(cote["cols"]):
                cc = c0 + 3 + i
                ws.cell(r_out, cc, f"={qsrc}!{lettre}{r}")
                monts.append(cc)
            style_ligne(ws, r_out, c0, c0 + 2 + len(cote["cols"]),
                        niveau, cols_montant=monts, col_ref=c0)
            ws.row_dimensions[r_out].height = 18
            r_out += 1
        return r_out

    fin_a = ecrire_cote(ca0, cote_actif)
    fin_p = ecrire_cote(cp0, cote_passif)
    cadre(ws, 8, ca0, max(fin_a, fin_p) - 1, ca0 + 2 + nca, MOYEN)
    cadre(ws, 8, cp0, max(fin_a, fin_p) - 1, cp0 + 2 + ncp, MOYEN)
    spec = {get_column_letter(ca0): 5, get_column_letter(ca0 + 1): 34,
            get_column_letter(ca0 + 2): 6,
            get_column_letter(cp0): 5, get_column_letter(cp0 + 1): 34,
            get_column_letter(cp0 + 2): 6}
    for i in range(nca):
        spec[get_column_letter(ca0 + 3 + i)] = 13.5
    for i in range(ncp):
        spec[get_column_letter(cp0 + 3 + i)] = 13.5
    largeurs(ws, spec)
    return ws


# --------------------------------------------------------------------------
# Alias historiques (moteurs v2) : pointent désormais sur la charte ETAFI
# --------------------------------------------------------------------------

BLEU_FONCE = C_NOTE
BLEU_CLAIR = C_ENTETE
GRIS_CLAIR = C_GRIS
OR_TOTAL = C_RUBRIQUE
F_TITRE = F_TITRE_ETAT_M
F_SOUS_TITRE = Font(name="Arial", size=10, bold=True)
F_ENTETE = F_ENTETE_COL
F_NORMAL = F_DONNEE
F_GRAS = F_DONNEE_G
R_TITRE = PatternFill("solid", fgColor=C_ENTETE)
R_ENTETE = PatternFill("solid", fgColor=C_ENTETE)
R_BANDE = PatternFill("solid", fgColor="FFFFFF")
R_TOTAL = PatternFill("solid", fgColor=C_GRIS)
BORD_FIN = B_FIN
