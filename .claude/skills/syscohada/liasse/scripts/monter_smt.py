#!/usr/bin/env python3
"""
monter_smt.py — Monte les états financiers du SYSTÈME MINIMAL DE TRÉSORERIE
(SYSCOHADA / AUDCIF, Titre X) à partir d'une balance, dans un classeur
professionnel construit de toutes pièces (il n'existe pas de gabarit Excel
officiel du SMT).

    python monter_smt.py balance_N.xlsx [balance_N1.xlsx] \
        --sortie etats-smt.xlsx --entite "..." --identifiant "..." \
        --exercice "31/12/N" --duree 12

Jeu d'états produit (AUDCIF Titre X, ch. 2 et 3) :
  - BILAN ACTIF et BILAN PASSIF (une feuille chacun) ;
  - COMPTE DE RESULTAT (recettes/dépenses corrigées des variations
    d'inventaire et des amortissements, G = C - D + E - F) ;
  - NOTE 1 (tableau de suivi du matériel, du mobilier et des cautions,
    pré-alimenté depuis la classe 2 de la balance), NOTE 2 (état des stocks),
    NOTE 3 (créances et dettes non échues), NOTE 4 (journal de trésorerie)
    et les deux journaux de suivi (créances impayées, dettes à payer) ;
  - BALANCE / BALANCE_N1 / CONTROLES / ANOMALIES + page de GARDE.

Chaque montant est une FORMULE Excel (SUMIF sur BALANCE / BALANCE_N1) :
l'origine de tout chiffre se retrace jusqu'au compte. La correspondance
compte → poste est documentée dans references/correspondance-smt.tsv
(construction du moteur : le Titre X ne publie pas de table officielle).

Les lignes « variation des créances / des dettes » (VB/VC) restent à saisir
depuis l'inventaire extra-comptable (NOTE 3) : une balance ne les porte pas.
"""

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from formules import (
    formule_tokens, set_lignes_max, q, nom_feuille,
    F_TITRE, F_SOUS_TITRE, F_ENTETE, F_NORMAL, F_GRAS,
    R_TITRE, R_ENTETE, R_BANDE, R_TOTAL, BORD_FIN, AL_CENTRE, AL_GAUCHE,
    FMT_MONTANT, style_entetes, style_zone_donnees, style_ligne_total,
    largeurs, style_titre, retirer_tirets, construire_identification,
    construire_fiche_notes, ordonner_feuilles,
    ecrire_cartouche, titre_etat, titre_note, entetes_bande, style_ligne,
    cadre, MOYEN, set_identite_etendue, construire_couverture,
    construire_garde_etafi, construire_fiche2, construire_controle_balance,
    construire_table_commentaires, construire_bilan_paysage,
    appliquer_police_arial, numeroter_pages, NOM_BALANCE, NOM_BALANCE_N1,
)
from monter_liasse import lire_balance

NOM_ACTIF = "Bilan-Actif"
NOM_PASSIF = "Bilan-Passif"
CR_NOM = "Résultat"
PAGE_SYS = "SYSCOHADA - SMT"


def entete_smt(ws, titre, page_ref, ident, ncols, note=False, taille=14):
    """Cartouche ETAFI + titre (vert pour un état, bleu nuit pour une
    note) : rend la ligne des en-têtes de colonnes."""
    ecrire_cartouche(ws, ident, page_ref, max(ncols, 5))
    if note:
        titre_note(ws, titre, max(ncols, 5), row=7)
    else:
        titre_etat(ws, titre, 1, max(ncols, 5), row=7, taille=taille)
    return 8


def _c(f):
    return f[1:] if f and f.startswith("=") else (f or "0")


def _tok(s):
    return [t.strip() for t in s.split(",") if t.strip()]


# --------------------------------------------------------------------------
# Formules des postes SMT (voir references/correspondance-smt.tsv)
# --------------------------------------------------------------------------

def f_actif(ref, bal):
    if ref == "SA1":
        return formule_tokens(_tok("2"), "nd", bal)
    if ref == "SA2":
        return formule_tokens(_tok("3"), "nd", bal)
    if ref == "SA3":
        deb4 = _c(formule_tokens(_tok("4"), "d", bal))
        dep = _c(formule_tokens(_tok("49,590,591"), "c", bal))
        titres = _c(formule_tokens(_tok("50,51"), "nd", bal))
        return f"={deb4}-({dep})+({titres})"
    if ref == "SA4":
        return formule_tokens(_tok("57"), "nd", bal)
    if ref == "SA5":
        banques = _c(formule_tokens(_tok("52,53,54,55,56,58"), "nd", bal))
        dep = _c(formule_tokens(_tok("592,593,594"), "c", bal))
        return f"={banques}-({dep})"
    return None


def f_passif(ref, bal, kzc_row):
    if ref == "SP1":
        return formule_tokens(_tok("10,11,12,14,15"), "nc", bal)
    if ref == "SP2":
        f13 = _c(formule_tokens(_tok("13"), "nc", bal))
        col = "D" if bal == "BALANCE" else "E"
        return f"={q(CR_NOM)}!{col}{kzc_row}+({f13})"
    if ref == "SP3":
        return formule_tokens(_tok("16,17,18,19"), "nc", bal)
    if ref == "SP4":
        cred4 = _c(formule_tokens(_tok("4"), "c", bal, exclude=_tok("49")))
        p599 = _c(formule_tokens(_tok("599"), "nc", bal))
        return f"={cred4}+({p599})"
    return None


CR_LIGNES = [
    # (ref, libellé, jetons, mode, total?)
    ("KA", "Recettes sur ventes ou prestations de services", "70", "nc", False),
    ("KB", "Autres recettes sur activités", "71,72,75,77,78,79,82,84,86,88", "nc", False),
    ("KX", "TOTAL DES RECETTES SUR PRODUITS (A)", None, None, True),
    ("JA", "Dépenses sur achats", "60!603", "nd", False),
    ("JB", "Dépenses sur loyers", "622", "nd", False),
    ("JC", "Dépenses sur salaires", "66", "nd", False),
    ("JD", "Dépenses sur impôts et taxes", "64,89", "nd", False),
    ("JE", "Charges d'intérêts", "67", "nd", False),
    ("JF", "Autres dépenses sur activités", "61,62!622,63,65,81,83,87", "nd", False),
    ("JX", "TOTAL DÉPENSES SUR CHARGES (B)", None, None, True),
    ("KZ", "SOLDE : Excédent (+) ou insuffisance (-) de recettes (C = A - B)", None, None, True),
    ("VA", "Variation des stocks [N - (N-1)] (D)", None, None, False),
    ("VB", "Variation des créances [N - (N-1)] (D) — à saisir depuis la NOTE 3", None, None, False),
    ("VC", "Variation des dettes d'exploitation [N - (N-1)] (E) — à saisir depuis la NOTE 3", None, None, False),
    ("JG", "DOTATIONS AUX AMORTISSEMENTS (F)", "68,69,85", "nd", False),
    ("KZC", "RÉSULTAT DE L'EXERCICE (G = C + D stocks + D créances - E dettes - F)", None, None, True),
]


# --------------------------------------------------------------------------
# Construction des feuilles
# --------------------------------------------------------------------------

def construire_bilan(wb, avec_n1, ident, cr_kzc_row):
    lignes_a = [
        ("SA1", "Immobilisations (1)", "1"),
        ("SA2", "Stocks", "2"),
        ("SA3", "Clients et débiteurs divers", "3"),
        ("SA4", "Caisse", "4"),
        ("SA5", "Banque (en + ou en -)", "4"),
    ]
    lignes_p = [
        ("SP1", "Compte exploitant", "1"),
        ("SP2", "Résultat de l'exercice (en + ou en -)", ""),
        ("SP3", "Emprunt", ""),
        ("SP4", "Fournisseurs et créditeurs divers", "3"),
    ]
    infos = {}
    for nom, lignes, total_ref, total_lab, f_poste, page in (
            (NOM_ACTIF, lignes_a, "SAZ", "TOTAL ACTIF", f_actif, "PAGE 1/2"),
            (NOM_PASSIF, lignes_p, "SPZ", "TOTAL PASSIF",
             lambda ref, bal: f_passif(ref, bal, cr_kzc_row), "PAGE 2/2")):
        ws = wb.create_sheet(nom)
        cote = "ACTIF" if nom == NOM_ACTIF else "PASSIF"
        r = entete_smt(ws, "BILAN", f"BILAN {PAGE_SYS}\n{page}", ident, 5,
                       taille=16)
        ws.cell(r, 1, "REF")
        ws.cell(r, 2, cote)
        ws.cell(r, 3, "NOTE")
        ws.cell(r, 4, "EXERCICE N")
        ws.cell(r, 5, "EXERCICE N-1")
        entetes_bande(ws, r, r, 1, 5)
        ws.row_dimensions[r].height = 22
        premiere = r + 1
        for ref, lib, note in lignes:
            r += 1
            ws.cell(r, 1, ref)
            ws.cell(r, 2, lib)
            ws.cell(r, 3, note)
            ws.cell(r, 4).value = f_poste(ref, "BALANCE")
            if avec_n1:
                ws.cell(r, 5).value = f_poste(ref, "BALANCE_N1")
            style_ligne(ws, r, 1, 5, "normal", cols_montant=(4, 5), col_ref=1)
            ws.row_dimensions[r].height = 22
        r += 1
        ws.cell(r, 1, total_ref)
        ws.cell(r, 2, total_lab)
        ws.cell(r, 4).value = f"=SUM(D{premiere}:D{r-1})"
        if avec_n1:
            ws.cell(r, 5).value = f"=SUM(E{premiere}:E{r-1})"
        style_ligne(ws, r, 1, 5, "general", cols_montant=(4, 5), col_ref=1)
        ws.row_dimensions[r].height = 22
        cadre(ws, premiere - 1, 1, r, 5, MOYEN)
        infos[total_ref] = r
        infos[nom] = (premiere, r)
        r += 2
        ws.cell(r, 1, "(1) À faire figurer à l'actif si montants significatifs "
                      "(AUDCIF Titre X). Registre des immobilisations : NOTE 1."
                if cote == "ACTIF" else
                "Le compte exploitant regroupe apports, prélèvements et "
                "réserves de l'exploitant (comptes 10, 11, 12, 14, 15).")
        ws.cell(r, 1).font = F_NORMAL
        largeurs(ws, {"A": 6, "B": 52, "C": 6.5, "D": 15.7, "E": 15.7})
    return infos


NIVEAUX_CR_SMT = {"KX": "section", "JX": "section", "KZ": "inter",
                  "KZC": "section"}


def construire_cr(wb, avec_n1, ident):
    ws = wb.create_sheet(CR_NOM)
    r = entete_smt(ws, "COMPTE DE RESULTAT",
                   f"COMPTE DE RESULTAT\n{PAGE_SYS}", ident, 5)
    ws.cell(r, 1, "REF")
    ws.cell(r, 2, "RUBRIQUES")
    ws.cell(r, 3, "NOTE")
    ws.cell(r, 4, "EXERCICE N")
    ws.cell(r, 5, "EXERCICE N-1")
    entetes_bande(ws, r, r, 1, 5)
    ws.row_dimensions[r].height = 22
    debut = r + 1
    rows = {}
    for ref, lib, jetons, mode, total in CR_LIGNES:
        r += 1
        rows[ref] = r
        ws.cell(r, 1, ref)
        ws.cell(r, 2, lib)
        note = "4" if ref in ("KA", "KB", "JA", "JB", "JC", "JD", "JF") else \
               ("2" if ref == "VA" else ("3" if ref in ("VB", "VC") else ""))
        ws.cell(r, 3, note)
        if jetons:
            inc, exc = (jetons.split("!") + [""])[:2]
            for col, bal, actif in ((4, "BALANCE", True), (5, "BALANCE_N1", avec_n1)):
                if actif:
                    ws.cell(r, col).value = formule_tokens(
                        _tok(inc), mode, bal, exclude=_tok(exc))
        style_ligne(ws, r, 1, 5, NIVEAUX_CR_SMT.get(ref, "normal"),
                    cols_montant=(4, 5), col_ref=1)
        ws.row_dimensions[r].height = 22
    # formules d'agrégation et lignes spéciales
    def pose(ref, fN, fN1=None):
        ws.cell(rows[ref], 4).value = fN
        if avec_n1 and fN1:
            ws.cell(rows[ref], 5).value = fN1

    pose("KX", f"=D{rows['KA']}+D{rows['KB']}", f"=E{rows['KA']}+E{rows['KB']}")
    ja_jf_d = "+".join(f"D{rows[x]}" for x in ("JA", "JB", "JC", "JD", "JE", "JF"))
    ja_jf_e = "+".join(f"E{rows[x]}" for x in ("JA", "JB", "JC", "JD", "JE", "JF"))
    pose("JX", f"={ja_jf_d}", f"={ja_jf_e}")
    pose("KZ", f"=D{rows['KX']}-D{rows['JX']}", f"=E{rows['KX']}-E{rows['JX']}")
    # VA : variation des stocks
    if avec_n1:
        s_n = _c(formule_tokens(_tok("3"), "nd", "BALANCE"))
        s_n1 = _c(formule_tokens(_tok("3"), "nd", "BALANCE_N1"))
        pose("VA", f"=({s_n})-({s_n1})")
    else:
        v603 = _c(formule_tokens(_tok("603"), "nc", "BALANCE"))
        v73 = _c(formule_tokens(_tok("73"), "nc", "BALANCE"))
        pose("VA", f"=({v603})+({v73})")
    # VB / VC : saisie manuelle (inventaire extra-comptable)
    ws.cell(rows["VB"], 4).value = 0
    ws.cell(rows["VC"], 4).value = 0
    pose("KZC", f"=D{rows['KZ']}+D{rows['VA']}+D{rows['VB']}-D{rows['VC']}-D{rows['JG']}",
         f"=E{rows['KZ']}+E{rows['VA']}+E{rows['VB']}-E{rows['VC']}-E{rows['JG']}")

    cadre(ws, debut - 1, 1, r, 5, MOYEN)
    r += 2
    ws.cell(r, 1, "Lecture (AUDCIF Titre X) : G = C - D + E - F. Les variations "
                  "sont ici affichées en N - (N-1) : une hausse des stocks ou des "
                  "créances augmente le résultat, une hausse des dettes le diminue. "
                  "VB et VC se saisissent depuis la NOTE 3 (inventaire extra-comptable).")
    ws.cell(r, 1).font = F_NORMAL
    largeurs(ws, {"A": 6, "B": 60, "C": 6.5, "D": 15.7, "E": 15.7})
    return rows


def construire_note1(wb, bal, ident):
    ws = wb.create_sheet("NOTE 1 IMMOBILISATIONS")
    r = entete_smt(ws, "NOTE 1 : TABLEAU SMT DE SUIVI DU MATERIEL, DU "
                       "MOBILIER ET DES CAUTIONS",
                   f"NOTE 1\n{PAGE_SYS}", ident, 7, note=True)
    for i, h in enumerate(["Compte", "Désignation", "Montant",
                           "Date d'acquisition", "Durée d'utilité",
                           "Date de sortie", "Prix de cession"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 7)
    premiere = r + 1
    for l in bal:
        c = l["compte"]
        if c.startswith("2") and not c.startswith(("28", "29")) \
                and abs(l["sd"] - l["sc"]) > 0.005:
            r += 1
            ws.cell(r, 1, c)
            ws.cell(r, 2, l["libelle"])
            ws.cell(r, 3).value = formule_tokens([c], "nd", "BALANCE")
    r += 1
    ws.cell(r, 2, "TOTAL IMMOBILISATIONS BRUTES")
    ws.cell(r, 3).value = (f"=SUM(C{premiere}:C{r-1})" if r > premiere else 0)
    r += 1
    ws.cell(r, 2, "Amortissements et dépréciations cumulés (28/29)")
    ws.cell(r, 3).value = formule_tokens(_tok("28,29"), "nc", "BALANCE")
    r += 1
    ws.cell(r, 2, "VALEUR NETTE (= poste Immobilisations du bilan)")
    ws.cell(r, 3).value = f"=C{r-2}-C{r-1}"
    style_zone_donnees(ws, premiere, r, 1, 7, cols_montant=(3, 7))
    style_ligne_total(ws, r, 1, 7, cols_montant=(3,))
    r += 2
    ws.cell(r, 1, "Compléter dates, durées et sorties depuis le registre : "
                  "la balance ne porte que les montants.")
    largeurs(ws, {"A": 12, "B": 46, "C": 16, "D": 16, "E": 14, "F": 14, "G": 16})


def construire_note2(wb, bal, avec_n1, ident):
    ws = wb.create_sheet("NOTE 2 STOCKS")
    r = entete_smt(ws, "NOTE 2 : ETAT DES STOCKS",
                   f"NOTE 2\n{PAGE_SYS}", ident, 5, note=True)
    for i, h in enumerate(["Référence", "Désignation", "Quantité",
                           "Prix unitaire", "Montant"], start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 5)
    premiere = r + 1
    for l in bal:
        if l["compte"].startswith("3") and not l["compte"].startswith("39") \
                and abs(l["sd"] - l["sc"]) > 0.005:
            r += 1
            ws.cell(r, 1, l["compte"])
            ws.cell(r, 2, l["libelle"])
            ws.cell(r, 5).value = formule_tokens([l["compte"]], "nd", "BALANCE")
    for _ in range(3):     # lignes vierges pour le détail article par article
        r += 1
    r += 1
    ws.cell(r, 2, "VALEUR DU STOCK FINAL")
    ws.cell(r, 5).value = formule_tokens(_tok("3"), "nd", "BALANCE")
    style_ligne_total(ws, r, 1, 5, cols_montant=(5,))
    r += 1
    ws.cell(r, 2, "VALEUR DU STOCK INITIAL")
    ws.cell(r, 5).value = (formule_tokens(_tok("3"), "nd", "BALANCE_N1")
                           if avec_n1 else 0)
    style_ligne_total(ws, r, 1, 5, cols_montant=(5,))
    r += 1
    ws.cell(r, 2, "VARIATION DES STOCKS [final - initial]")
    ws.cell(r, 5).value = f"=E{r-2}-E{r-1}"
    style_ligne_total(ws, r, 1, 5, cols_montant=(5,))
    style_zone_donnees(ws, premiere, r - 3, 1, 5, cols_montant=(5,))
    largeurs(ws, {"A": 12, "B": 46, "C": 12, "D": 14, "E": 16})


def construire_note3(wb, avec_n1, ident):
    ws = wb.create_sheet("NOTE 3 CREANCES-DETTES")
    r = entete_smt(ws, "NOTE 3 : ETAT DES CREANCES ET DES DETTES NON ECHUES",
                   f"NOTE 3\n{PAGE_SYS}", ident, 6, note=True)
    for bloc, rappel_mode in (("CRÉANCES — clients-usagers et autres débiteurs", "creances"),
                              ("DETTES — fournisseurs et autres créditeurs", "dettes")):
        ws.cell(r, 1, bloc)
        ws.cell(r, 1).font = F_SOUS_TITRE
        r += 1
        for i, h in enumerate(["Date", "Nom", "Montant au 31/12/N",
                               "Montant au 01/01/N", "Variation en valeur",
                               "Variation en %"], start=1):
            ws.cell(r, i, h)
        style_entetes(ws, r, 1, 6)
        premiere = r + 1
        for _ in range(6):
            r += 1
            ws.cell(r, 5).value = f"=C{r}-D{r}"
            ws.cell(r, 6).value = f"=IF(D{r}=0,\"\",(C{r}-D{r})/D{r})"
        r += 1
        ws.cell(r, 2, "TOTAL DES " + ("CRÉANCES" if rappel_mode == "creances" else "DETTES"))
        for col in "CDE":
            ws[f"{col}{r}"] = f"=SUM({col}{premiere}:{col}{r-1})"
        style_ligne_total(ws, r, 1, 6, cols_montant=(3, 4, 5))
        style_zone_donnees(ws, premiere, r - 1, 1, 6, cols_montant=(3, 4, 5))
        r += 1
        if rappel_mode == "creances":
            f_n = _c(formule_tokens(_tok("4"), "d", "BALANCE"))
            f_n1 = _c(formule_tokens(_tok("4"), "d", "BALANCE_N1")) if avec_n1 else "0"
        else:
            f_n = _c(formule_tokens(_tok("4"), "c", "BALANCE"))
            f_n1 = _c(formule_tokens(_tok("4"), "c", "BALANCE_N1")) if avec_n1 else "0"
        ws.cell(r, 2, "Rappel balance (classe 4, "
                      + ("soldes débiteurs" if rappel_mode == "creances" else "soldes créditeurs")
                      + ") — contrôle")
        ws.cell(r, 3).value = f"={f_n}"
        ws.cell(r, 4).value = f"={f_n1}"
        ws.cell(r, 3).number_format = FMT_MONTANT
        ws.cell(r, 4).number_format = FMT_MONTANT
        r += 3
    ws.cell(r, 1, "Reporter la variation des créances en VB et celle des dettes "
                  "en VC du compte de résultat.")
    largeurs(ws, {"A": 12, "B": 40, "C": 18, "D": 18, "E": 16, "F": 12})


def construire_note4(wb, ident):
    ws = wb.create_sheet("NOTE 4 JOURNAL TRESORERIE")
    r = entete_smt(ws, "NOTE 4 : JOURNAL DE TRESORERIE SMT",
                   f"NOTE 4\n{PAGE_SYS}", ident, 11, note=True)
    entetes = ["Date", "Libellés", "Recettes", "Dépenses", "Solde",
               "Vent. recettes : Ventes", "Autres", "Matériel et mobilier",
               "Vent. dépenses : Achats", "Loyers / Salaires / Impôts", "Autres"]
    for i, h in enumerate(entetes, start=1):
        ws.cell(r, i, h)
    style_entetes(ws, r, 1, 11)
    r += 1
    ws.cell(r, 2, "Report à nouveau")
    for _ in range(14):
        r += 1
        ws.cell(r, 5).value = f"=E{r-1}+C{r}-D{r}"
    r += 1
    ws.cell(r, 2, "Solde à reporter")
    ws.cell(r, 5).value = f"=E{r-1}"
    style_zone_donnees(ws, 9, r, 1, 11, cols_montant=(3, 4, 5, 6, 7, 8, 9, 10, 11))
    style_ligne_total(ws, r, 1, 11, cols_montant=(5,))
    largeurs(ws, {"A": 11, "B": 34, "C": 13, "D": 13, "E": 13, "F": 13,
                  "G": 11, "H": 15, "I": 13, "J": 22, "K": 11})


def construire_journaux_suivi(wb, ident):
    ws = wb.create_sheet("JOURNAUX DE SUIVI")
    r = entete_smt(ws, "JOURNAUX DE SUIVI SMT : CREANCES IMPAYEES ET "
                       "DETTES A PAYER",
                   f"JOURNAUX DE SUIVI\n{PAGE_SYS}", ident, 5, note=True)
    for titre, tiers in (("JOURNAL DE SUIVI DES CRÉANCES IMPAYÉES", "Nom du client"),
                         ("JOURNAL DE SUIVI DES DETTES À PAYER", "Nom du fournisseur")):
        ws.cell(r, 1, titre)
        ws.cell(r, 1).font = F_SOUS_TITRE
        r += 1
        for i, h in enumerate(["Date", "N° facture", tiers, "Montant",
                               "Date de paiement"], start=1):
            ws.cell(r, i, h)
        style_entetes(ws, r, 1, 5)
        premiere = r + 1
        r += 8
        ws.cell(r, 3, "TOTAL")
        ws.cell(r, 4).value = f"=SUM(D{premiere}:D{r-1})"
        style_zone_donnees(ws, premiere, r - 1, 1, 5, cols_montant=(4,))
        style_ligne_total(ws, r, 1, 5, cols_montant=(4,))
        r += 3
    largeurs(ws, {"A": 12, "B": 14, "C": 40, "D": 16, "E": 16})


def construire_garde(wb, ident, avec_n1):
    construire_garde_etafi(
        wb, ident,
        bandeau="ETATS FINANCIERS NORMALISES\nDU SYSTEME COMPTABLE OHADA "
                "(SYSCOHADA REVISE - AUDCIF)",
        sous_bandeau="Entités relevant du Système minimal de trésorerie",
        systeme="SYSTEME MINIMAL DE TRESORERIE",
        documents=["Fiche d'identification et renseignements divers",
                   "Bilan (actif et passif)",
                   "Compte de résultat",
                   "Notes annexes 1 à 4",
                   "Journaux de suivi (créances impayées, dettes à payer)"])


def ecrire_balance(wb, nom, bal):
    b = wb.create_sheet(nom_feuille(nom))
    entetes = ["Compte", "Intitulé", "Préfixe 2", "Préfixe 3", "Préfixe 4",
               "Solde final débit", "Solde final crédit",
               "Mouvement débit", "Mouvement crédit"]
    b.append(entetes)
    style_entetes(b, 1, 1, len(entetes))
    for l in bal:
        c = l["compte"]
        b.append([c, l["libelle"], c[:2], c[:3], c[:4],
                  round(l["sd"], 2), round(l["sc"], 2),
                  round(l.get("md", 0.0), 2), round(l.get("mc", 0.0), 2)])
    style_zone_donnees(b, 2, b.max_row, 1, len(entetes), cols_montant=(6, 7, 8, 9))
    largeurs(b, {"A": 12, "B": 42, "C": 9, "D": 9, "E": 9, "F": 15, "G": 15,
                 "H": 15, "I": 15})
    b.freeze_panes = "A2"


def detecter_anomalies_smt(bal, seuil=1.0):
    a = []
    sd = sum(l["sd"] for l in bal)
    sc = sum(l["sc"] for l in bal)
    if abs(sd - sc) > seuil:
        a.append({"gravite": "BLOQUANT", "compte": "", "libelle": "Balance entière",
                  "probleme": f"Balance déséquilibrée : débit {sd:,.2f} ≠ crédit {sc:,.2f}",
                  "solution": "Reprendre la saisie avant montage."})
    for l in bal:
        c, net = l["compte"], l["sd"] - l["sc"]
        if abs(net) <= seuil:
            continue
        cl = c[0] if c else ""
        if len(c) < 2 or cl not in "12345678":
            a.append({"gravite": "A_TRAITER", "compte": c, "libelle": l["libelle"],
                      "probleme": f"Compte non conforme au plan OHADA (classe '{cl}').",
                      "solution": "Réaffecter au compte équivalent avant montage."})
        if cl == "6" and (l["sc"] - l["sd"]) > seuil:
            a.append({"gravite": "A_VERIFIER", "compte": c, "libelle": l["libelle"],
                      "probleme": "Compte de dépense (classe 6) au solde créditeur.",
                      "solution": "Contrôler l'imputation."})
        if cl == "7" and net > seuil:
            a.append({"gravite": "A_VERIFIER", "compte": c, "libelle": l["libelle"],
                      "probleme": "Compte de recette (classe 7) au solde débiteur.",
                      "solution": "Contrôler l'imputation."})
    if any(l["compte"].startswith("4") and abs(l["sd"] - l["sc"]) > seuil for l in bal):
        a.append({"gravite": "INFO", "compte": "", "libelle": "Classe 4",
                  "probleme": "La balance porte des comptes de tiers : recettes et "
                              "dépenses incluent alors des montants non encaissés/décaissés "
                              "(base engagement).",
                  "solution": "Dans ce cas, laisser VB et VC à zéro : le résultat est déjà "
                              "en base engagement. Ne les servir que pour une balance de "
                              "trésorerie pure, depuis l'inventaire extra-comptable (NOTE 3)."})
    a.append({"gravite": "INFO", "compte": "", "libelle": "SMT — seuils",
              "probleme": "Vérifier l'assujettissement au SMT : CA HT ≤ 60 M FCFA "
                          "(négoce), 40 M (artisanat), 30 M (services) — AUDCIF art. 13.",
              "solution": "Au-delà des seuils, monter la liasse du Système normal "
                          "(monter_liasse.py)."})
    return a


def construire_controles(wb, bal, avec_n1, kzc_row, infos_bilan):
    ctl = wb.create_sheet("CONTROLES")
    n = max(len(bal), 1)
    ctl.append(["Contrôle", "Valeur", "Attendu"])
    style_entetes(ctl, 1, 1, 3)
    B = q(NOM_BALANCE)
    lignes = [
        ("Total solde débit balance", f"=SUM({B}!F2:F{n+1})", ""),
        ("Total solde crédit balance", f"=SUM({B}!G2:G{n+1})", ""),
        ("Écart balance (doit être 0)", "=B2-B3", 0),
        ("Total actif", f"={q(NOM_ACTIF)}!D{infos_bilan['SAZ']}", ""),
        ("Total passif", f"={q(NOM_PASSIF)}!D{infos_bilan['SPZ']}", ""),
        ("Écart actif - passif (doit être 0)", "=B5-B6", 0),
        ("Résultat (compte de résultat, G)", f"={q(CR_NOM)}!D{kzc_row}", ""),
    ]
    for lab, f, att in lignes:
        ctl.append([lab, f, att])
    style_zone_donnees(ctl, 2, ctl.max_row, 1, 3, cols_montant=(2,))
    largeurs(ctl, {"A": 56, "B": 20, "C": 10})
    return ctl


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("balance_N")
    ap.add_argument("balance_N1", nargs="?")
    ap.add_argument("--sortie", default="etats-smt.xlsx")
    ap.add_argument("--entite", default="")
    ap.add_argument("--identifiant", default="")
    ap.add_argument("--exercice", default="")
    ap.add_argument("--duree", default="12")
    ap.add_argument("--adresse", default="")
    ap.add_argument("--sigle", default="")
    ap.add_argument("--ntd", default="")
    args = ap.parse_args()
    set_identite_etendue(args.adresse, args.sigle, args.ntd)

    bal, idx = lire_balance(args.balance_N)
    print(f"Balance N : {len(bal)} comptes. Colonnes repérées : {idx}")
    bal1 = None
    if args.balance_N1:
        bal1, _ = lire_balance(args.balance_N1)
        print(f"Balance N-1 : {len(bal1)} comptes.")
    avec_n1 = bal1 is not None
    set_lignes_max(max(len(bal), len(bal1 or [])) + 20)

    ident = (args.entite, args.identifiant, args.exercice, args.duree)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ordre : CR d'abord pour connaître la ligne KZC, bilan ensuite
    cr_rows = construire_cr(wb, avec_n1, ident)
    kzc_row = cr_rows["KZC"]
    infos_bilan = construire_bilan(wb, avec_n1, ident, kzc_row)

    construire_note1(wb, bal, ident)
    construire_note2(wb, bal, avec_n1, ident)
    construire_note3(wb, avec_n1, ident)
    construire_note4(wb, ident)
    construire_journaux_suivi(wb, ident)

    ecrire_balance(wb, "BALANCE", bal)
    if avec_n1:
        ecrire_balance(wb, "BALANCE_N1", bal1)

    anomalies = detecter_anomalies_smt(bal)
    construire_controles(wb, bal, avec_n1, kzc_row, infos_bilan)
    construire_controle_balance(wb, avec_n1, len(bal), len(bal1 or []),
                                cols=("F", "G", "H", "I"))

    an = wb.create_sheet("ANOMALIES")
    an.append(["Gravité", "Compte", "Intitulé", "Problème", "Solution proposée"])
    style_entetes(an, 1, 1, 5)
    ordre = {"BLOQUANT": 0, "A_TRAITER": 1, "A_VERIFIER": 2, "MINEUR": 3, "INFO": 4}
    for x in sorted(anomalies, key=lambda z: ordre.get(z["gravite"], 9)):
        an.append([x["gravite"], x["compte"], x["libelle"], x["probleme"], x["solution"]])
    style_zone_donnees(an, 2, max(an.max_row, 2), 1, 5)
    largeurs(an, {"A": 12, "B": 12, "C": 26, "D": 62, "E": 62})

    construire_couverture(wb, ident, "LIASSE SMT")
    construire_garde(wb, ident, avec_n1)
    construire_identification(wb, ident, "SYSCOHADA révisé",
                              "Système minimal de trésorerie")
    construire_fiche2(wb, ident, "DIRIGEANTS ET RESPONSABLES DE L'ENTITE")
    pa_a, pa_p = infos_bilan[NOM_ACTIF], infos_bilan[NOM_PASSIF]
    construire_bilan_paysage(
        wb, ident,
        {"feuille": NOM_ACTIF, "lig_debut": pa_a[0], "lig_fin": pa_a[1],
         "col_note": "C", "libelle": "ACTIF",
         "cols": [("EXERCICE N", "D"), ("EXERCICE N-1", "E")]},
        {"feuille": NOM_PASSIF, "lig_debut": pa_p[0], "lig_fin": pa_p[1],
         "col_note": "C", "libelle": "PASSIF",
         "cols": [("EXERCICE N", "D"), ("EXERCICE N-1", "E")]},
        titre="BILAN", page_ref=f"BILAN {PAGE_SYS}\nPAGE 1/1")
    parties = [("Partie 1 : Notes sur le bilan",
                [("NOTE 1", "Tableau SMT de suivi du matériel, du mobilier "
                            "et des cautions"),
                 ("NOTE 2", "État des stocks"),
                 ("NOTE 3", "État des créances et des dettes non échues")]),
               ("Partie 2 : Notes sur le compte de résultat et pièces de "
                "tenue",
                [("NOTE 4", "Journal de trésorerie SMT"),
                 ("", "Journal de suivi des créances impayées"),
                 ("", "Journal de suivi des dettes à payer")])]
    construire_fiche_notes(wb, parties, ident,
                           note_pied="(1) A : applicable ; N/A : non "
                           "applicable. Ne pas joindre les notes non "
                           "documentées ; supprimer les lignes non chiffrées "
                           "avant remise.")
    construire_table_commentaires(wb, parties, ident)
    ordonner_feuilles(wb, [NOM_BALANCE, NOM_BALANCE_N1, "CONTROLE BALANCE",
                           "Couverture", "Garde", "Fiche 1", "Fiche 2",
                           "Bilan paysage", NOM_ACTIF, NOM_PASSIF, CR_NOM,
                           "NOTES ANNEXES",
                           "NOTE 1 IMMOBILISATIONS", "NOTE 2 STOCKS",
                           "NOTE 3 CREANCES-DETTES",
                           "NOTE 4 JOURNAL TRESORERIE",
                           "JOURNAUX DE SUIVI",
                           "TABLE COMMENTAIRE", "CONTROLES", "ANOMALIES"])
    retirer_tirets(wb)
    appliquer_police_arial(wb)
    numeroter_pages(wb)

    wb.save(args.sortie)
    print(f"États SMT écrits : {args.sortie}")
    bloquants = [a for a in anomalies if a["gravite"] in ("BLOQUANT", "A_TRAITER")]
    print(f"Anomalies : {len(anomalies)} dont {len(bloquants)} à traiter avant remise.")


if __name__ == "__main__":
    main()
