#!/usr/bin/env python3
"""
notes_sn.py — Alimentation des notes annexes du Système normal SYSCOHADA
dans le gabarit officiel (assets/gabarit-liasse.xlsx).

Chaque ligne de note reçoit une FORMULE Excel (SUMIF sur la feuille BALANCE,
et BALANCE_N1 pour l'exercice précédent) plutôt qu'une valeur figée : tout
chiffre de l'annexe est retraçable jusqu'aux comptes de la balance.

Sources des affectations ligne par ligne :
  - AUDCIF, Titre IX, ch. 6 (liste officielle des notes) et ch. 7
    (correspondance postes/comptes), recoupés avec le plan de comptes
    (comptes/references/plan-comptes.tsv) pour chaque numéro cité.
  - Les lignes purement déclaratives (identité, effectifs, réévaluations,
    engagements) restent vierges : elles ne se déduisent pas d'une balance.

Syntaxe des jetons : "271,277!2771" = comptes 271 et 277, sauf 2771.
Modes : nd (net débiteur), nc (net créditeur), d (débits seuls),
        c (crédits seuls) — voir formules.py.

Trois familles de notes :
  - SOLDES : ventilation de soldes de clôture. Entièrement en formules.
  - MOUVEMENTS (3A, 3C, 3D, 28) : ouverture depuis BALANCE_N1, clôture par
    la formule du gabarit ; les colonnes de flux utilisent les colonnes de
    mouvement de la balance si présentes, sinon la variation nette N/N-1
    (=MAX(0, N-N-1)) est posée en augmentation ou en diminution — toujours
    en formule, jamais en valeur muette.
  - DÉCLARATIVES (1, 2, 3B, 3E, 8A/3F, 13, 16B, 16C, 27B, 31, 32, 33, 35,
    36) : seul l'en-tête (entité, exercice) est rempli.
"""

from formules import formule_tokens, FMT_MONTANT


# --------------------------------------------------------------------------
# Notes de type SOLDES : {feuille: {"n": colonne N, "n1": colonne N-1,
#                                   "lignes": {row: (jetons, mode)}}}
# --------------------------------------------------------------------------

NOTES_SOLDES = {
    # NOTE 4 — Immobilisations financières (comptes 26/27, dépréc. 296/297)
    "NOTE 4": {"n": "B", "n1": "C", "lignes": {
        13: ("26", "nd"),                # Titres de participation
        14: ("271,277,278", "nd"),       # Prêts et créances (+ rattachées, diverses)
        15: ("272", "nd"),               # Prêts au personnel
        16: ("273", "nd"),               # Créances sur l'État
        17: ("274", "nd"),               # Titres immobilisés
        18: ("275", "nd"),               # Dépôts et cautionnements versés
        19: ("276", "nd"),               # Intérêts courus
        21: ("296", "nc"),               # Dépréciations titres de participation
        22: ("297", "nc"),               # Dépréciations autres immobilisations
    }},
    # NOTE 5 — Actif circulant HAO / Dettes circulantes HAO
    "NOTE 5": {"n": "C", "n1": "D", "lignes": {
        14: ("485", "nd"),               # Créances sur cessions d'immobilisations
        15: ("488", "nd"),               # Autres créances HAO
        17: ("498", "nc"),               # Dépréciations des créances HAO
        27: ("481", "nc"),               # Fournisseurs d'investissements
        28: ("482", "nc"),               # F. d'investissements, effets à payer
        30: ("484,4998", "nc"),          # Autres dettes HAO
    }},
    # NOTE 6 — Stocks et en-cours (31 à 38, dépréciations 39)
    "NOTE 6": {"n": "C", "n1": "D", "lignes": {
        12: ("31", "nd"), 13: ("32", "nd"), 14: ("33", "nd"), 15: ("34", "nd"),
        16: ("35", "nd"), 17: ("36", "nd"), 18: ("37", "nd"), 19: ("38", "nd"),
        22: ("39", "nc"),
    }},
    # NOTE 7 — Clients (41x), dépréciations 491, clients créditeurs 419x
    "NOTE 7": {"n": "B", "n1": "C", "lignes": {
        12: ("411", "nd"),               # Clients
        13: ("412,413", "nd"),           # Effets à recevoir, valeurs impayées
        16: ("414", "nd"),               # Créances sur cessions courantes d'immo
        17: ("415", "nd"),               # Effets escomptés non échus
        18: ("416", "nd"),               # Créances litigieuses ou douteuses
        19: ("418", "nd"),               # Produits à recevoir
        22: ("491", "nc"),               # Dépréciations
        26: ("4191", "nc"),              # Avances reçues (hors groupe)
        27: ("4192", "nc"),              # Avances reçues groupe
        28: ("419!4191,4192", "nc"),     # Autres clients créditeurs
    }},
    # NOTE 8 — Autres créances (soldes débiteurs des tiers)
    "NOTE 8": {"n": "B", "n1": "C", "lignes": {
        14: ("42", "d"),                 # Personnel
        15: ("43", "d"),                 # Organismes sociaux
        16: ("44", "d"),                 # État et collectivités publiques
        17: ("451,452,458", "d"),        # Organismes internationaux
        18: ("46", "d"),                 # Apporteurs, associés et groupe
        19: ("475", "d"),                # Compte transitoire révision SYSCOHADA
        20: ("471,472,473,474,476", "d"),  # Autres débiteurs divers + CCA
        21: ("185", "d"),                # Comptes permanents non bloqués
        22: ("186,187", "d"),            # Comptes de liaison charges / produits
        23: ("188", "d"),                # Comptes de liaison sociétés en participation
        26: ("492,493,494,495,496,497", "nc"),  # Dépréciations
    }},
    # NOTE 9 — Titres de placement (50x, dépréciations 590)
    "NOTE 9": {"n": "B", "n1": "C", "lignes": {
        11: ("501", "nd"), 12: ("502", "nd"), 13: ("503", "nd"),
        14: ("504", "nd"), 15: ("505", "nd"), 16: ("506", "nd"),
        17: ("508", "nd"), 19: ("590", "nc"),
    }},
    # NOTE 10 — Valeurs à encaisser (51x, dépréciations 591)
    "NOTE 10": {"n": "B", "n1": "C", "lignes": {
        10: ("511", "nd"), 11: ("512", "nd"), 12: ("513", "nd"),
        13: ("514", "nd"), 14: ("515", "nd"), 15: ("518", "nd"),
        17: ("591", "nc"),
    }},
    # NOTE 11 — Disponibilités (52/53/54/55/57/58, dépréciations 592-594)
    "NOTE 11": {"n": "C", "n1": "D", "lignes": {
        12: ("521", "d"), 13: ("522", "d"), 14: ("525", "d"),
        15: ("523,524", "d"), 16: ("526", "d"), 17: ("531,532", "d"),
        18: ("533,538", "d"), 19: ("536", "d"), 20: ("54", "d"),
        21: ("57", "nd"), 22: ("55", "nd"), 23: ("581,582", "nd"),
        25: ("592,593,594", "nc"),
    }},
    # NOTE 12 — (section Transferts de charges)
    "NOTE 12": {"n": "B", "n1": "C", "lignes": {
        25: ("781", "nc"),               # Transferts de charges d'exploitation
        29: ("787", "nc"),               # Transferts de charges financières
    }},
    # NOTE 14 — Primes et réserves
    "NOTE 14": {"n": "C", "n1": "D", "lignes": {
        10: ("1052", "nc"), 11: ("1051", "nc"), 12: ("1053", "nc"),
        13: ("1054", "nc"), 14: ("1058", "nc"),
        17: ("111", "nc"), 18: ("112", "nc"), 19: ("1131", "nc"),
        20: ("1132", "nc"), 21: ("1133,1134,1138", "nc"),
        24: ("118", "nc"),
    }},
    # NOTE 15A — Subventions d'investissement (141x/148) et prov. réglementées
    "NOTE 15A": {"n": "C", "n1": "D", "lignes": {
        10: ("1411", "nc"), 11: ("1412", "nc"), 12: ("1413", "nc"),
        13: ("1414", "nc"), 14: ("1415", "nc"), 15: ("1416", "nc"),
        16: ("1417", "nc"), 17: ("1418,148", "nc"),
        20: ("151", "nc"), 21: ("152", "nc"), 22: ("154", "nc"),
        23: ("155", "nc"), 24: ("156", "nc"), 25: ("157", "nc"),
        26: ("153,158", "nc"),
    }},
    # NOTES 15B — Autres fonds propres (avances conditionnées 167)
    "NOTES 15B": {"n": "C", "n1": "D", "lignes": {
        11: ("167", "nc"),
    }},
    # NOTE 16A — Dettes financières et ressources assimilées
    "NOTE 16A": {"n": "B", "n1": "C", "lignes": {
        7: ("161", "nc"), 8: ("162", "nc"), 9: ("163", "nc"),
        10: ("164", "nc"), 11: ("165", "nc"), 12: ("166", "nc"),
        13: ("167", "nc"), 14: ("168", "nc"), 15: ("181,182,183", "nc"),
        16: ("184", "nc"),
        19: ("172", "nc"), 20: ("173", "nc"), 21: ("174", "nc"),
        22: ("176", "nc"), 23: ("178", "nc"),
        26: ("191", "nc"), 27: ("192", "nc"), 28: ("193", "nc"),
        29: ("194", "nc"), 30: ("195", "nc"), 31: ("196", "nc"),
        33: ("197", "nc"), 38: ("198", "nc"),
    }},
    # NOTE 17 — Fournisseurs d'exploitation (401/402/408, débiteurs 409x)
    # 404 (acquisitions courantes d'immobilisations) rattaché à la ligne
    # "dettes en compte" pour que le total de la note recoupe le poste DJ.
    "NOTE 17": {"n": "B", "n1": "C", "lignes": {
        10: ("401,404", "nc"),           # Dettes en compte
        11: ("402", "nc"),               # Effets à payer
        13: ("408", "nc"),               # Factures non parvenues
        17: ("4091,4093", "d"),          # Avances et acomptes versés
        18: ("4092", "d"),               # Avances groupe
        19: ("4094,4098", "d"),          # Autres fournisseurs débiteurs
    }},
    # NOTE 18 — Dettes fiscales et sociales (42/43/44, soldes créditeurs)
    "NOTE 18": {"n": "B", "n1": "C", "lignes": {
        11: ("421", "c"),
        12: ("422", "c"),
        13: ("423,424,425,426,427,428", "c"),
        14: ("431", "c"),
        15: ("432", "c"),
        16: ("433,438", "c"),
        19: ("441", "c"),
        20: ("442,446", "c"),
        21: ("443,444,445", "c"),
        22: ("447", "c"),
        23: ("448,449", "c"),
    }},
    # NOTE 19 — Autres dettes et provisions pour risques à court terme
    "NOTE 19": {"n": "B", "n1": "C", "lignes": {
        10: ("451,452,458", "c"),        # Organismes internationaux
        12: ("461", "c"),                # Apporteurs, opérations sur le capital
        13: ("462", "c"),                # Associés, comptes courants
        14: ("465", "c"),                # Dividendes à payer
        15: ("466", "c"),                # Groupe, comptes courants
        16: ("463,467", "c"),            # Autres dettes associés
        19: ("471", "c"),                # Créditeurs divers
        23: ("472", "c"),                # Versements restant sur titres
        24: ("475", "c"),                # Compte transitoire révision
        25: ("473,474,476,477", "c"),    # Autres créditeurs divers + PCA
        28: ("185", "c"),                # Comptes permanents non bloqués
        29: ("186,187", "c"),            # Comptes de liaison
        30: ("188", "c"),                # Liaison sociétés en participation
        35: ("499!4998,599", "nc"),      # Provisions pour risques à court terme
    }},
    # NOTE 20 — Banques, crédit d'escompte et de trésorerie (trésorerie passif)
    "NOTE 20": {"n": "B", "n1": "C", "lignes": {
        12: ("564", "c"), 13: ("565", "c"),
        17: ("521", "c"), 18: ("522", "c"), 19: ("523,524,525", "c"),
        20: ("526,566", "c"), 21: ("561", "c"),
    }},
    # NOTE 21 — Chiffre d'affaires et autres produits (701/702-706/707, 72, 75)
    "NOTE 21": {"n": "B", "n1": "C", "lignes": {
        13: ("7011", "nc"), 14: ("7012", "nc"), 15: ("7013,7014", "nc"),
        16: ("7015,7019", "nc"),
        19: ("702,703,704", "nc"),       # Ventes de produits fabriqués (total)
        25: ("7051,7061", "nc"), 26: ("7052,7062", "nc"),
        27: ("7053,7054,7063,7064", "nc"), 28: ("7055,7065,7059,7069", "nc"),
        31: ("707", "nc"),
        36: ("72", "nc"), 38: ("75", "nc"),
    }},
    # NOTE 22 — Achats (601/602/604/605/608)
    "NOTE 22": {"n": "B", "n1": "C", "lignes": {
        11: ("6011", "nd"), 12: ("6012", "nd"),
        13: ("6013,6014,6015,6019", "nd"),
        16: ("6021", "nd"), 17: ("6022", "nd"),
        18: ("6023,6024,6025,6029", "nd"),
        21: ("6041", "nd"), 22: ("6042", "nd"), 23: ("6043", "nd"),
        24: ("6044,6046", "nd"), 25: ("6051", "nd"), 26: ("6052", "nd"),
        27: ("6053", "nd"), 28: ("6054", "nd"), 29: ("6047,6055", "nd"),
        30: ("6056", "nd"), 31: ("6057,6058", "nd"), 32: ("608", "nd"),
        33: ("6045", "nd"), 34: ("6049,6059", "nd"),
    }},
    # NOTE 23 — Transports (61x)
    "NOTE 23": {"n": "B", "n1": "C", "lignes": {
        10: ("612", "nd"), 11: ("613", "nd"), 12: ("614", "nd"),
        13: ("616", "nd"), 14: ("618", "nd"),
    }},
    # NOTE 24 — Services extérieurs (62x/63x)
    "NOTE 24": {"n": "B", "n1": "C", "lignes": {
        10: ("621", "nd"), 11: ("622", "nd"), 12: ("623", "nd"),
        13: ("624", "nd"), 14: ("625", "nd"), 15: ("626", "nd"),
        16: ("627", "nd"), 17: ("628", "nd"), 18: ("631", "nd"),
        19: ("632", "nd"), 20: ("633", "nd"), 21: ("634", "nd"),
        22: ("635", "nd"), 23: ("637,638", "nd"),
    }},
    # NOTE 25 — Impôts et taxes (64x)
    "NOTE 25": {"n": "B", "n1": "C", "lignes": {
        11: ("641", "nd"), 12: ("645", "nd"), 13: ("646", "nd"),
        14: ("647", "nd"), 15: ("648", "nd"),
    }},
    # NOTE 26 — Autres charges (65x)
    "NOTE 26": {"n": "B", "n1": "C", "lignes": {
        10: ("651", "nd"),
        12: ("652", "nd"),
        13: ("654", "nd"),
        14: ("6581", "nd"),
        15: ("6582,6583", "nd"),
        16: ("656,657,6588", "nd"),
        17: ("659", "nd"),
    }},
    # NOTE 27A — Charges de personnel (66x)
    "NOTE 27A": {"n": "B", "n1": "C", "lignes": {
        11: ("661,662", "nd"), 12: ("663", "nd"), 13: ("664", "nd"),
        14: ("666", "nd"), 15: ("667", "nd"), 16: ("668", "nd"),
    }},
    # NOTE 29 — Charges et revenus financiers (67x / 77x)
    "NOTE 29": {"n": "B", "n1": "C", "lignes": {
        10: ("671", "nd"), 11: ("672", "nd"), 12: ("673", "nd"),
        13: ("674", "nd"), 14: ("675", "nd"), 15: ("676", "nd"),
        16: ("677", "nd"), 18: ("678", "nd"), 19: ("679", "nd"),
        22: ("771,775", "nc"), 23: ("772", "nc"), 24: ("773", "nc"),
        25: ("774", "nc"), 26: ("776", "nc"), 27: ("777", "nc"),
        28: ("778", "nc"), 29: ("779", "nc"),
    }},
    # NOTE 30 — Autres charges et produits HAO (83x-88x)
    "NOTE 30": {"n": "B", "n1": "C", "lignes": {
        10: ("831,833,837", "nd"),
        13: ("834", "nd"), 14: ("835", "nd"), 15: ("836", "nd"),
        16: ("839", "nd"), 17: ("85", "nd"), 18: ("87", "nd"),
        22: ("841,843,844,847", "nc"),
        25: ("845", "nc"), 26: ("846", "nc"), 27: ("848", "nc"),
        28: ("849", "nc"), 29: ("86", "nc"), 30: ("88", "nc"),
    }},
}

# --------------------------------------------------------------------------
# Corrections/compléments de formules de totalisation dans le gabarit.
# Certaines cellules du modèle Excel officiel portent des zéros ou des
# formules tronquées ; elles sont remplacées par la vraie totalisation.
# Chaque entrée : (feuille, cellule, formule) — {n}/{n1} substitués par la
# colonne correspondante de la note quand pertinent.
# --------------------------------------------------------------------------

TOTAUX_FIXES = [
    ("NOTE 4",  "B23", "=B20-B21-B22"), ("NOTE 4",  "C23", "=C20-C21-C22"),
    ("NOTE 5",  "C16", "=SUM(C14:C15)"), ("NOTE 5",  "D16", "=SUM(D14:D15)"),
    ("NOTE 5",  "C19", "=C16-C17"),      ("NOTE 5",  "D19", "=D16-D17"),
    ("NOTE 5",  "C31", "=SUM(C27:C30)"), ("NOTE 5",  "D31", "=SUM(D27:D30)"),
    ("NOTE 6",  "C24", "=C20-C22"),      ("NOTE 6",  "D24", "=D20-D22"),
    ("NOTE 7",  "B24", "=B20-B22"),      ("NOTE 7",  "C24", "=C20-C22"),
    ("NOTE 7",  "B29", "=SUM(B26:B28)"), ("NOTE 7",  "C29", "=SUM(C26:C28)"),
    ("NOTE 8",  "B28", "=B24-B26"),      ("NOTE 8",  "C28", "=C24-C26"),
    ("NOTE 9",  "B18", "=SUM(B11:B17)"), ("NOTE 9",  "C18", "=SUM(C11:C17)"),
    ("NOTE 9",  "B21", "=B18-B19"),      ("NOTE 9",  "C21", "=C18-C19"),
    ("NOTE 10", "B16", "=SUM(B10:B15)"), ("NOTE 10", "C16", "=SUM(C10:C15)"),
    ("NOTE 10", "B19", "=B16-B17"),      ("NOTE 10", "C19", "=C16-C17"),
    ("NOTE 11", "C27", "=C24-C25"),      ("NOTE 11", "D27", "=D24-D25"),
    ("NOTE 14", "C15", "=SUM(C10:C14)"), ("NOTE 14", "D15", "=SUM(D10:D14)"),
    ("NOTE 14", "C22", "=SUM(C17:C21)"), ("NOTE 14", "D22", "=SUM(D17:D21)"),
    ("NOTE 15A", "C18", "=SUM(C10:C17)"), ("NOTE 15A", "D18", "=SUM(D10:D17)"),
    ("NOTE 15A", "C27", "=SUM(C20:C26)"), ("NOTE 15A", "D27", "=SUM(D20:D26)"),
    ("NOTE 15A", "C29", "=C18+C27"),     ("NOTE 15A", "D29", "=D18+D27"),
    ("NOTES 15B", "C15", "=SUM(C10:C14)"), ("NOTES 15B", "D15", "=SUM(D10:D14)"),
    ("NOTE 16A", "B24", "=SUM(B19:B23)"), ("NOTE 16A", "C24", "=SUM(C19:C23)"),
    ("NOTE 17", "B15", "=SUM(B10:B14)"), ("NOTE 17", "C15", "=SUM(C10:C14)"),
    ("NOTE 17", "B20", "=SUM(B17:B19)"), ("NOTE 17", "C20", "=SUM(C17:C19)"),
    ("NOTE 19", "B17", "=SUM(B12:B16)"), ("NOTE 19", "C17", "=SUM(C12:C16)"),
    ("NOTE 19", "B26", "=SUM(B19:B25)"), ("NOTE 19", "C26", "=SUM(C19:C25)"),
    ("NOTE 19", "B31", "=SUM(B28:B30)"), ("NOTE 19", "C31", "=SUM(C28:C30)"),
    ("NOTE 19", "B33", "=B10+B17+B26+B31"), ("NOTE 19", "C33", "=C10+C17+C26+C31"),
    ("NOTE 20", "B14", "=B12+B13"),      ("NOTE 20", "C14", "=C12+C13"),
    ("NOTE 20", "B22", "=SUM(B17:B21)"), ("NOTE 20", "C22", "=SUM(C17:C21)"),
    ("NOTE 20", "B24", "=B14+B22"),      ("NOTE 20", "C24", "=C14+C22"),
    ("NOTE 21", "B23", "=SUM(B19:B22)"), ("NOTE 21", "C23", "=SUM(C19:C22)"),
    ("NOTE 29", "B30", "=SUM(B22:B29)"), ("NOTE 29", "C30", "=SUM(C22:C29)"),
    ("NOTE 30", "B20", "=SUM(B10:B19)"), ("NOTE 30", "C20", "=SUM(C10:C19)"),
    ("NOTE 30", "B31", "=SUM(B22:B30)"), ("NOTE 30", "C31", "=SUM(C22:C30)"),
    ("NOTE 30", "B33", "=B31-B20"),      ("NOTE 30", "C33", "=C31-C20"),
    # NOTE 3A : le sous-total corporel du modèle saute la ligne 21 (Terrains
    # hors immeuble de placement) — corrigé pour couvrir B21:B27.
    ("NOTE 3A", "B20", "=SUM(B21:B27)"), ("NOTE 3A", "C20", "=SUM(C21:C27)"),
    ("NOTE 3A", "D20", "=SUM(D21:D27)"), ("NOTE 3A", "E20", "=SUM(E21:E27)"),
    ("NOTE 3A", "F20", "=SUM(F21:F27)"), ("NOTE 3A", "G20", "=SUM(G21:G27)"),
    ("NOTE 3A", "H20", "=SUM(H21:H27)"),
    ("NOTE 3A", "B28", "=SUM(B29:B30)"), ("NOTE 3A", "C28", "=SUM(C29:C30)"),
    ("NOTE 3C", "F26", "=B26+C26-E26"),
]

# --------------------------------------------------------------------------
# Notes de type MOUVEMENTS
# --------------------------------------------------------------------------
# NOTE 3A (immobilisations brutes) : B ouverture / C acquisitions /
# F cessions / H clôture (formule du gabarit). Jetons = comptes bruts.
NOTE_3A_LIGNES = {
    16: "211,2181,2191",
    17: "212,213,214,2193",
    18: "215,216",
    19: "217,218!2181,2198",
    21: "22!2281",
    22: "2281",
    23: "231,232,233,237,2391!2315,2325",
    24: "2315,2325",
    25: "234,235,238,2392,2393",
    26: "24!245,2495",
    27: "245,2495",
    29: "251",
    30: "252",
    32: "26",
    33: "27",
}

# NOTE 3C (amortissements) : B ouverture / C dotations / E sorties /
# F clôture (formule du gabarit). Jetons = comptes d'amortissements (28x).
NOTE_3C_LIGNES = {
    15: "2811",
    16: "2812,2813,2814",
    17: "2815,2816",
    18: "2817,2818",
    20: "282",
    22: "2831,2832,2833,2837",
    24: "2834,2835,2838",
    25: "284!2845",
    26: "2845",
}

# NOTE 28 (provisions et dépréciations inscrites au bilan) :
# B ouverture (N-1) / dotations C-D-E / reprises F-G-H / I clôture (formule).
# Chaque ligne : (jetons bilan, colonne de dotation, colonne de reprise).
NOTE_28_LIGNES = {
    12: ("15", "E", "H"),                 # Provisions réglementées (HAO)
    13: ("19", "D", "G"),                 # Provisions financières R&C
    14: ("29", "C", "F"),                 # Dépréciations des immobilisations
    17: ("39", "C", "F"),                 # Dépréciations des stocks
    18: ("498", "E", "H"),                # Dépréciations actif circulant HAO
    19: ("490", "C", "F"),                # Dépréciations fournisseurs
    20: ("491", "C", "F"),                # Dépréciations clients
    22: ("492,493,494,495,496,497", "C", "F"),  # Dépréciations autres créances
    23: ("590", "D", "G"),                # Dépréciations titres de placement
    24: ("591", "D", "G"),                # Dépréciations valeurs à encaisser
    25: ("592,593,594", "D", "G"),        # Dépréciations disponibilités
    26: ("499!4998", "C", "F"),           # Provisions risques CT exploitation
    27: ("599", "D", "G"),                # Provisions risques CT financières
}

# NOTE 3D (plus/moins-values de cession) : injection aux sous-totaux par
# famille — VNC depuis 81x (débits), prix de cession depuis 82x (crédits).
NOTE_3D_FAMILLES = [
    ("SOUS TOTAL : IMMOBILISATIONS INCORPORELLES", "811", "821"),
    ("SOUS TOTAL : IMMOBILISATIONS CORPORELLES", "812", "822"),
    ("SOUS TOTAL : IMMOBILISATIONS FINANCIERES", "816", "826"),
    ("TOTAL GENERAL", "81", "82"),
]


# --------------------------------------------------------------------------
# Moteur d'injection
# --------------------------------------------------------------------------

def _parse_jetons(s):
    """'24!245,2495' -> (["24"], ["245","2495"]) ; '26,27' -> inclusion seule."""
    if "!" in s:
        inc, exc = s.split("!", 1)
        return ([t.strip() for t in inc.split(",") if t.strip()],
                [t.strip() for t in exc.split(",") if t.strip()])
    return ([t.strip() for t in s.split(",") if t.strip()], [])


def _cible(ws, coord):
    """Cellule d'écriture réelle : si coord tombe dans une plage fusionnée,
    on vise l'ancre (coin haut-gauche) ; si l'ancre porte déjà une valeur
    (libellé, formule posée plus tôt), on renonce plutôt que d'écraser."""
    from openpyxl.cell.cell import MergedCell
    cell = ws[coord]
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            anchor = ws.cell(rng.min_row, rng.min_col)
            if anchor.coordinate == coord:
                return anchor          # coord est l'ancre : écriture voulue
            return anchor if anchor.value in (None, "", 0) else None
    return None


def _pose(ws, coord, formule):
    if formule is None:
        return
    cell = _cible(ws, coord)
    if cell is None:
        return
    cell.value = formule
    cell.number_format = FMT_MONTANT


def _diff_pos(jetons, exc, mode, a="BALANCE", b="BALANCE_N1"):
    """=MAX(0, somme(N) - somme(N-1)) — variation nette posée en formule."""
    fa = formule_tokens(jetons, mode, a, exclude=exc)[1:]
    fb = formule_tokens(jetons, mode, b, exclude=exc)[1:]
    return f"=MAX(0,({fa})-({fb}))"


def injecter_notes_soldes(wb, avec_n1):
    for feuille, spec in NOTES_SOLDES.items():
        if feuille not in wb.sheetnames:
            continue
        ws = wb[feuille]
        for row, (jetons, mode) in spec["lignes"].items():
            inc, exc = _parse_jetons(jetons)
            f = formule_tokens(inc, mode, "BALANCE", exclude=exc)
            if f:
                _pose(ws, f"{spec['n']}{row}", f)
            if avec_n1:
                f1 = formule_tokens(inc, mode, "BALANCE_N1", exclude=exc)
                if f1:
                    _pose(ws, f"{spec['n1']}{row}", f1)


def injecter_note_3a(wb, avec_n1, avec_mvt):
    if "NOTE 3A" not in wb.sheetnames:
        return
    ws = wb["NOTE 3A"]
    for row, jetons in NOTE_3A_LIGNES.items():
        inc, exc = _parse_jetons(jetons)
        if avec_n1:
            _pose(ws, f"B{row}", formule_tokens(inc, "nd", "BALANCE_N1", exclude=exc))
            if avec_mvt:
                _pose(ws, f"C{row}", formule_tokens(inc, "md", "BALANCE", exclude=exc))
                _pose(ws, f"F{row}", formule_tokens(inc, "mc", "BALANCE", exclude=exc))
            else:
                _pose(ws, f"C{row}", _diff_pos(inc, exc, "nd"))
                _pose(ws, f"F{row}", _diff_pos(inc, exc, "nd", "BALANCE_N1", "BALANCE"))
        else:
            # sans N-1 : tout le brut de clôture est posé en colonne A
            # (ouverture inconnue) — signalé dans ANOMALIES par le moteur
            _pose(ws, f"B{row}", formule_tokens(inc, "nd", "BALANCE", exclude=exc))


def injecter_note_3c(wb, avec_n1, avec_mvt):
    if "NOTE 3C" not in wb.sheetnames:
        return
    ws = wb["NOTE 3C"]
    for row, jetons in NOTE_3C_LIGNES.items():
        inc, exc = _parse_jetons(jetons)
        if avec_n1:
            _pose(ws, f"B{row}", formule_tokens(inc, "nc", "BALANCE_N1", exclude=exc))
            if avec_mvt:
                _pose(ws, f"C{row}", formule_tokens(inc, "mc", "BALANCE", exclude=exc))
                _pose(ws, f"E{row}", formule_tokens(inc, "md", "BALANCE", exclude=exc))
            else:
                _pose(ws, f"C{row}", _diff_pos(inc, exc, "nc"))
                _pose(ws, f"E{row}", _diff_pos(inc, exc, "nc", "BALANCE_N1", "BALANCE"))
        else:
            _pose(ws, f"B{row}", formule_tokens(inc, "nc", "BALANCE", exclude=exc))


def injecter_note_3d(wb):
    if "NOTE 3D" not in wb.sheetnames:
        return
    ws = wb["NOTE 3D"]
    # colonnes du modèle : C brut A / D amort B / E VNC / F prix / G +/- value
    for label, cpt_vnc, cpt_prix in NOTE_3D_FAMILLES:
        for row in ws.iter_rows(min_col=2, max_col=2):
            cell = row[0]
            if cell.value and label.lower() in str(cell.value).lower():
                r = cell.row
                _pose(ws, f"E{r}", formule_tokens([cpt_vnc], "nd", "BALANCE"))
                _pose(ws, f"F{r}", formule_tokens([cpt_prix], "nc", "BALANCE"))
                _pose(ws, f"G{r}", f"=F{r}-E{r}")
                break


def injecter_note_28(wb, avec_n1):
    if "NOTE 28" not in wb.sheetnames:
        return
    ws = wb["NOTE 28"]
    for row, (jetons, col_dot, col_rep) in NOTE_28_LIGNES.items():
        inc, exc = _parse_jetons(jetons)
        if avec_n1:
            _pose(ws, f"B{row}", formule_tokens(inc, "nc", "BALANCE_N1", exclude=exc))
            # variation nette posée en dotation (si hausse) / reprise (si baisse) ;
            # à défaut de l'échéancier, la note reste juste en clôture et
            # documentée « présentation en net » dans le README.
            _pose(ws, f"{col_dot}{row}", _diff_pos(inc, exc, "nc"))
            _pose(ws, f"{col_rep}{row}", _diff_pos(inc, exc, "nc", "BALANCE_N1", "BALANCE"))
        else:
            _pose(ws, f"B{row}", formule_tokens(inc, "nc", "BALANCE", exclude=exc))
    # sous-totaux complets (le modèle n'en portait qu'une partie)
    for col in "BCDEFGHI":
        _pose(ws, f"{col}15", f"=SUM({col}12:{col}14)")
        _pose(ws, f"{col}28", f"=SUM({col}17:{col}27)")
        _pose(ws, f"{col}30", f"={col}15+{col}28")


def corriger_totaux(wb):
    for feuille, coord, formule in TOTAUX_FIXES:
        if feuille in wb.sheetnames:
            ws = wb[feuille]
            cell = _cible(ws, coord)
            if cell is None:
                continue
            cell.value = formule
            cell.number_format = FMT_MONTANT


# --------------------------------------------------------------------------
# NOTE 34 — fiche de synthèse : formules croisées vers les états
# --------------------------------------------------------------------------

def injecter_note_34(wb, avec_n1):
    if "NOTE 34" not in wb.sheetnames:
        return
    ws = wb["NOTE 34"]
    CR = "'Compte de Résultat'"

    def construire(col_cr, bal):
        """Formules de la fiche pour une colonne du CR (E=N, F=N-1) et une
        feuille balance (BALANCE / BALANCE_N1)."""
        def s(tokens, mode):
            return formule_tokens(tokens.split(","), mode, bal)[1:]
        return {
            "10": f"={CR}!{col_cr}16", "11": f"={CR}!{col_cr}12",
            "12": f"={CR}!{col_cr}30", "13": f"={CR}!{col_cr}32",
            "14": f"={CR}!{col_cr}35", "15": f"={CR}!{col_cr}41",
            "16": f"={CR}!{col_cr}42", "18": f"={CR}!{col_cr}50",
            "21": "=" + s("654", "nd"),
            "22": "=" + s("754", "nc"),
            # CAFG : revenus financiers (hors gains de change) + reprises fin.
            "24": f"={CR}!{col_cr}36+{CR}!{col_cr}37-" + s("776", "nc"),
            "25": "=" + s("776", "nc"),
            "26": f"={CR}!{col_cr}38",
            "27": f"={CR}!{col_cr}44-" + s("86", "nc") + "-" + s("848", "nc"),
            "28": "=" + s("848", "nc"),
            "29": f"={CR}!{col_cr}39+{CR}!{col_cr}40-" + s("676", "nd"),
            "30": "=" + s("676", "nd"),
            "31": f"={CR}!{col_cr}48", "32": f"={CR}!{col_cr}49",
        }

    for suffixe, f in construire("E", "BALANCE").items():
        ws["C" + suffixe] = f
        ws["C" + suffixe].number_format = FMT_MONTANT
    if avec_n1:
        for suffixe, f in construire("F", "BALANCE_N1").items():
            ws["D" + suffixe] = f
            ws["D" + suffixe].number_format = FMT_MONTANT


# --------------------------------------------------------------------------
# En-têtes d'identification sur toutes les feuilles
# --------------------------------------------------------------------------

_LABELS = [
    ("désignation entité", "entite"),
    ("numéro d’identification", "identifiant"),
    ("numéro d'identification", "identifiant"),
    ("exercice clos le", "exercice"),
    ("durée (en mois)", "duree"),
]


def remplir_entetes(wb, entite="", identifiant="", exercice="", duree=""):
    vals = {"entite": entite, "identifiant": identifiant,
            "exercice": exercice, "duree": duree}
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                bas = v.strip().lower()
                for label, cle in _LABELS:
                    if bas.startswith(label) and vals[cle]:
                        base = v.rstrip()
                        if not base.endswith(":"):
                            base = base.rstrip() + " :"
                        try:
                            cell.value = f"{base} {vals[cle]}"
                        except AttributeError:
                            pass  # cellule fusionnée non maîtresse
                        break


# --------------------------------------------------------------------------
# Point d'entrée
# --------------------------------------------------------------------------

def injecter_notes(wb, avec_n1, avec_mvt, entite="", identifiant="",
                   exercice="", duree=""):
    """Alimente l'ensemble des notes mécanisables du gabarit officiel."""
    injecter_notes_soldes(wb, avec_n1)
    injecter_note_3a(wb, avec_n1, avec_mvt)
    injecter_note_3c(wb, avec_n1, avec_mvt)
    injecter_note_3d(wb)
    injecter_note_28(wb, avec_n1)
    injecter_note_34(wb, avec_n1)
    corriger_totaux(wb)
    remplir_entetes(wb, entite, identifiant, exercice, duree)
