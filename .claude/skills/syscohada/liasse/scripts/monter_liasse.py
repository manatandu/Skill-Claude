#!/usr/bin/env python3
"""
monter_liasse.py — Monte la liasse SYSCOHADA (Système normal) à partir d'une
balance générale, en s'appuyant sur le gabarit officiel OHADA.

    python monter_liasse.py balance_N.xlsx [balance_N1.xlsx] \
        --gabarit ../assets/gabarit-liasse.xlsx \
        --correspondance ../references/correspondance.tsv \
        --sortie liasse.xlsx \
        --entite "ACME SARL" --identifiant "CD/KIN/RCCM/..." \
        --exercice "31/12/2025" --duree 12

Ce que le moteur produit (v3) :
  - Bilan actif (brut / amort-dépréc / net), bilan passif, compte de résultat
    et TFT — chaque poste est écrit sous forme de FORMULE Excel (SUMIF) qui
    pointe vers la feuille BALANCE (et BALANCE_N1 si fournie) : tout chiffre
    de la liasse est retraçable jusqu'aux comptes qui l'alimentent.
  - Les NOTES ANNEXES mécanisables (voir notes_sn.py) : notes de soldes
    entièrement en formules, notes de mouvements (3A, 3C, 3D, 28) alimentées
    depuis N-1 et les colonnes de mouvement quand elles existent, NOTE 34
    (fiche de synthèse) en formules croisées. Les notes déclaratives restent
    des gabarits à compléter (en-tête d'identification pré-rempli partout).
  - Feuilles GARDE (page de garde), BALANCE, BALANCE_N1, CONTROLES
    (équilibres + recoupements notes/bilan, tous en formules), ANOMALIES.

Conditions du TFT (formules détaillées dans references/tft-formules-praticien.md) :
  - ZA, FB à FE exigent la balance N-1 ; FF à FQ exigent en plus les colonnes
    de mouvement de l'exercice. À défaut, les postes restent vides et une
    anomalie INFO le signale — jamais un chiffre approximé en silence.
"""

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib_mapping import (
    charger_maquette, normaliser_compte, compte_dans_expr, token_match, Expr,
)
from formules import (
    formule_tokens, formule_expr, retirer_tirets,
    F_TITRE, F_SOUS_TITRE, F_ENTETE, F_NORMAL, F_GRAS,
    R_TITRE, R_ENTETE, R_BANDE, R_TOTAL, BORD_FIN, AL_CENTRE, AL_GAUCHE,
    FMT_MONTANT, style_entetes, style_zone_donnees, largeurs,
)
import notes_sn


# --------------------------------------------------------------------------
# Lecture souple d'une balance
# --------------------------------------------------------------------------

ENTETES = {
    "compte":   ["cpte", "compte", "n° compte", "numero", "num compte", "code"],
    "libelle":  ["intitul", "libell", "designation", "désignation"],
    "sd":       ["s.f. debit", "sf debit", "solde final debit", "solde debit",
                 "sf_d", "final debit", "debit final"],
    "sc":       ["s.f. credit", "sf credit", "solde final credit", "solde credit",
                 "sf_c", "final credit", "credit final"],
    "md":       ["mouvement debit", "mvt debit", "m.debit", "mouvements debit",
                 "debit mouvement", "debit periode", "mouv. debit"],
    "mc":       ["mouvement credit", "mvt credit", "m.credit", "mouvements credit",
                 "credit mouvement", "credit periode", "mouv. credit"],
}


def _norm(x):
    return str(x).strip().lower().replace("é", "e").replace("è", "e")


def _trouver_colonnes(entetes):
    idx = {}
    for i, cell in enumerate(entetes):
        h = _norm(cell)
        for cle, motifs in ENTETES.items():
            if cle in idx:
                continue
            if any(m in h for m in motifs):
                idx[cle] = i
    return idx


def lire_balance(chemin):
    """Retourne (liste de {compte, libelle, sd, sc, md, mc}, diagnostic)."""
    ext = os.path.splitext(chemin)[1].lower()
    if ext in (".xls",):
        import xlrd
        wb = xlrd.open_workbook(chemin)
        sh = wb.sheet_by_index(0)
        for s in wb.sheets():
            if "balance" in s.name.lower():
                sh = s
                break
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)]
    elif ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(chemin, data_only=True)
        sh = wb.active
        for name in wb.sheetnames:
            if "balance" in name.lower():
                sh = wb[name]
                break
        rows = [[c for c in row] for row in sh.iter_rows(values_only=True)]
    elif ext in (".csv", ".tsv"):
        import csv
        delim = "\t" if ext == ".tsv" else None
        with open(chemin, encoding="utf-8-sig") as f:
            if delim is None:
                sample = f.read(2048)
                f.seek(0)
                delim = "\t" if sample.count("\t") > sample.count(",") else ","
            rows = [r for r in csv.reader(f, delimiter=delim)]
    else:
        raise ValueError(f"Extension non gérée : {ext}")

    idx, debut = {}, 0
    for r, row in enumerate(rows[:30]):
        cand = _trouver_colonnes(row)
        if "compte" in cand and ("sd" in cand or "sc" in cand):
            idx, debut = cand, r + 1
            break
    if "compte" not in idx:
        raise ValueError("Impossible de repérer la colonne des comptes.")

    lignes = []
    for row in rows[debut:]:
        if idx["compte"] >= len(row):
            continue
        compte = normaliser_compte(row[idx["compte"]])
        if not compte:
            continue

        def num(i):
            if i is None or i >= len(row):
                return 0.0
            try:
                return float(row[i] or 0)
            except (TypeError, ValueError):
                return 0.0

        lignes.append({
            "compte": compte,
            "libelle": str(row[idx["libelle"]]).strip() if "libelle" in idx
                       and idx["libelle"] < len(row) else "",
            "sd": num(idx.get("sd")), "sc": num(idx.get("sc")),
            "md": num(idx.get("md")), "mc": num(idx.get("mc")),
        })
    return lignes, idx


# --------------------------------------------------------------------------
# Détection d'anomalies (inchangé sur le fond — voir references/anomalies.md)
# --------------------------------------------------------------------------

def detecter_anomalies(bal, rubs, seuil=1.0):
    a = []
    sd = sum(l["sd"] for l in bal)
    sc = sum(l["sc"] for l in bal)
    if abs(sd - sc) > seuil:
        a.append({
            "gravite": "BLOQUANT", "compte": "", "libelle": "Balance entière",
            "probleme": f"Balance déséquilibrée : débit {sd:,.2f} ≠ crédit {sc:,.2f} (écart {sd-sc:,.2f})",
            "solution": "Reprendre la saisie : la liasse ne peut pas boucler tant que la balance ne boucle pas.",
        })
    elif abs(sd - sc) > 0.005:
        a.append({
            "gravite": "MINEUR", "compte": "", "libelle": "Balance entière",
            "probleme": f"Écart d'arrondi de {sd-sc:,.5f} entre débit et crédit.",
            "solution": "Simple arrondi, sans effet sur la présentation. Peut être ignoré.",
        })

    all_exprs = []
    for r in rubs.values():
        all_exprs.append(r.brut)
        all_exprs.append(r.amort)

    for l in bal:
        c = l["compte"]
        net = l["sd"] - l["sc"]
        if abs(net) <= seuil:
            continue
        cl = c[0] if c else ""

        if len(c) < 2 or cl not in "12345678":
            a.append({
                "gravite": "A_TRAITER", "compte": c, "libelle": l["libelle"],
                "probleme": f"Compte non conforme au plan OHADA (classe '{cl}').",
                "solution": "Rapprocher du libellé et réaffecter au compte OHADA équivalent avant montage.",
            })
            continue

        if c[:3] in ("585", "588"):
            a.append({
                "gravite": "A_TRAITER", "compte": c, "libelle": l["libelle"],
                "probleme": "Virement interne (585/588) non soldé. L'AUDCIF ne lui donne aucun poste.",
                "solution": "Solder le compte à la clôture (Titre VII, classe 5) : un résidu est une erreur d'inventaire.",
            })
            continue

        affecte = any(compte_dans_expr(c, e) for e in all_exprs)
        if not affecte:
            a.append({
                "gravite": "A_TRAITER", "compte": c, "libelle": l["libelle"],
                "probleme": "Solde non capté par aucune rubrique. Il fera fuir l'équilibre.",
                "solution": "Vérifier le numéro (préfixe OHADA) ou compléter la maquette pour ce cas.",
            })

        if cl == "6" and (l["sc"] - l["sd"]) > seuil:
            a.append({
                "gravite": "A_VERIFIER", "compte": c, "libelle": l["libelle"],
                "probleme": "Compte de charge (classe 6) au solde créditeur.",
                "solution": "Contrôler : transfert de charge, RRR obtenus ou erreur d'imputation.",
            })
        if cl == "7" and net > seuil:
            a.append({
                "gravite": "A_VERIFIER", "compte": c, "libelle": l["libelle"],
                "probleme": "Compte de produit (classe 7) au solde débiteur.",
                "solution": "Contrôler : RRR accordés, annulation de produit ou erreur d'imputation.",
            })
        if cl == "2" and c[:2] not in ("28", "29") and (l["sc"] - l["sd"]) > seuil:
            a.append({
                "gravite": "A_VERIFIER", "compte": c, "libelle": l["libelle"],
                "probleme": "Immobilisation (classe 2 hors amort/déprec) au solde créditeur.",
                "solution": "Vérifier : cession non soldée, avoir sur immo ou mauvaise ventilation.",
            })

    cl6_7_8 = any(l["compte"][0] in "678" and abs(l["sd"] - l["sc"]) > seuil for l in bal)
    cl13 = any(l["compte"].startswith("13") and abs(l["sd"] - l["sc"]) > seuil for l in bal)
    if cl6_7_8 and cl13:
        a.append({
            "gravite": "A_VERIFIER", "compte": "", "libelle": "Résultat",
            "probleme": "Classes 6/7/8 ouvertes ET classe 13 mouvementée : risque de double résultat.",
            "solution": "Fournir soit une balance avant clôture (6/7/8 ouverts), soit après (13 seul).",
        })
    if not cl6_7_8:
        a.append({
            "gravite": "INFO", "compte": "", "libelle": "Résultat",
            "probleme": "Classes 6/7/8 non mouvementées : le compte de résultat ressortira à zéro.",
            "solution": "Pour un compte de résultat renseigné, fournir la balance avant affectation.",
        })
    return a


# --------------------------------------------------------------------------
# Calcul des montants par rubrique (contrôles internes en Python ;
# les cellules, elles, reçoivent des formules)
# --------------------------------------------------------------------------

def somme(bal, expr, mode):
    t = 0.0
    for l in bal:
        c = l["compte"]
        if not compte_dans_expr(c, expr):
            continue
        sd, sc = l["sd"], l["sc"]
        if expr.sens == "debiteur":
            t += sd
        elif expr.sens == "crediteur":
            t += sc
        elif mode in ("brut_actif", "charge"):
            t += sd - sc
        else:
            t += sc - sd
    return t


def calculer(bal, rubs):
    val = {}
    for ref, r in rubs.items():
        if r.formule:
            continue
        if r.etat == "BILAN-ACTIF":
            brut = somme(bal, r.brut, "brut_actif")
            amort = somme(bal, r.amort, "amort")
            val[ref] = {"brut": brut, "amort": amort, "net": brut - amort}
        elif r.etat == "BILAN-PASSIF":
            val[ref] = somme(bal, r.brut, "passif")
        elif r.etat == "COMPTE-DE-RESULTAT":
            s = r.brut.signe
            if s == "+":
                val[ref] = somme(bal, r.brut, "produit")
            elif s == "-":
                val[ref] = somme(bal, r.brut, "charge")
            else:
                cls = r.brut.include[0][0] if r.brut.include else "6"
                val[ref] = somme(bal, r.brut, "charge" if cls == "6" else "var")
    return val


def _mode_cr(r):
    s = r.brut.signe
    if s == "+":
        return "nc"
    if s == "-":
        return "nd"
    cls = r.brut.include[0][0] if r.brut.include else "6"
    return "nd" if cls == "6" else "nc"


# --------------------------------------------------------------------------
# Générateurs de formules par rubrique
# --------------------------------------------------------------------------

def _corps(f):
    """Enlève le '=' de tête d'une formule (pour composer)."""
    return f[1:] if f and f.startswith("=") else (f or "0")


def f_rub_brut(rubs, ref, feuille):
    return formule_expr(rubs[ref].brut, "nd", feuille)


def f_rub_amort(rubs, ref, feuille):
    return formule_expr(rubs[ref].amort, "nc", feuille)


def f_rub_net(rubs, ref, feuille):
    fb = f_rub_brut(rubs, ref, feuille)
    fa = f_rub_amort(rubs, ref, feuille)
    if fb and fa:
        return f"=({_corps(fb)})-({_corps(fa)})"
    return fb or (f"=-({_corps(fa)})" if fa else None)


def f_rub_passif(rubs, ref, feuille):
    return formule_expr(rubs[ref].brut, "nc", feuille)


def f_rub_cr(rubs, ref, feuille):
    return formule_expr(rubs[ref].brut, _mode_cr(rubs[ref]), feuille)


def f_somme_nets(rubs, refs, feuille):
    corps = "+".join(f"({_corps(f_rub_net(rubs, r, feuille))})" for r in refs
                     if f_rub_net(rubs, r, feuille))
    return "=" + corps if corps else None


# --------------------------------------------------------------------------
# Coordonnées des cellules dans le gabarit officiel
# --------------------------------------------------------------------------

ACTIF_ROW = {  # col D=brut, E=amort, F=net(N), G=net(N-1)
    "AE": 11, "AF": 12, "AG": 13, "AH": 14, "AJ": 16, "AK": 18, "AL": 20,
    "AM": 21, "AN": 22, "AP": 23, "AR": 25, "AS": 26, "BA": 28, "BB": 29,
    "BH": 31, "BI": 32, "BJ": 33, "BQ": 35, "BR": 36, "BS": 37, "BU": 39,
}
PASSIF_ROW = {  # col D=net(N), E=net(N-1)
    "CA": 10, "CB": 11, "CD": 12, "CE": 13, "CF": 14, "CG": 15, "CH": 16,
    "CJ": 17, "CL": 18, "CM": 19, "DA": 21, "DB": 22, "DC": 23, "DH": 26,
    "DI": 27, "DJ": 28, "DK": 29, "DM": 30, "DN": 31, "DQ": 33, "DR": 34,
    "DV": 36,
}
RES_ROW = {  # col E=net(N), F=net(N-1)
    "TA": 9, "RA": 10, "RB": 11, "TB": 13, "TC": 14, "TD": 15, "TE": 17,
    "TF": 18, "TG": 19, "TH": 20, "TI": 21, "RC": 22, "RD": 23, "RE": 24,
    "RF": 25, "RG": 26, "RH": 27, "RI": 28, "RJ": 29, "RK": 31, "TJ": 33,
    "RL": 34, "TK": 36, "TL": 37, "TM": 38, "RM": 39, "RN": 40, "TN": 43,
    "TO": 44, "RO": 45, "RP": 46, "RQ": 48, "RS": 49,
}


def injecter(wb, rubs, avec_n1):
    """Écrit les FORMULES des postes dans ACTIF, PASSIF, Compte de Résultat."""
    A, P, R = wb["ACTIF"], wb["PASSIF"], wb["Compte de Résultat"]

    for ref, row in ACTIF_ROW.items():
        fb = f_rub_brut(rubs, ref, "BALANCE")
        fa = f_rub_amort(rubs, ref, "BALANCE")
        A.cell(row, 4).value = fb or 0
        A.cell(row, 5).value = fa or 0
        A.cell(row, 6).value = f"=D{row}-E{row}"
        for col in (4, 5, 6, 7):
            A.cell(row, col).number_format = FMT_MONTANT
        if avec_n1:
            fn = f_rub_net(rubs, ref, "BALANCE_N1")
            A.cell(row, 7).value = fn or 0

    for mref, row in PASSIF_ROW.items():
        ref = "CI" if mref == "CJ" else mref
        if mref == "CJ":
            # résultat net = résultat du CR + solde éventuel de la classe 13 :
            # la liasse boucle que la balance soit avant ou après affectation
            f13 = formule_tokens(["13"], "nc", "BALANCE")
            P.cell(row, 4).value = f"='Compte de Résultat'!E50+({_corps(f13)})"
            if avec_n1:
                f13b = formule_tokens(["13"], "nc", "BALANCE_N1")
                P.cell(row, 5).value = f"='Compte de Résultat'!F50+({_corps(f13b)})"
        else:
            fp = f_rub_passif(rubs, ref, "BALANCE")
            P.cell(row, 4).value = fp or 0
            if avec_n1:
                fp1 = f_rub_passif(rubs, ref, "BALANCE_N1")
                P.cell(row, 5).value = fp1 or 0
        for col in (4, 5):
            P.cell(row, col).number_format = FMT_MONTANT

    for ref, row in RES_ROW.items():
        fr = f_rub_cr(rubs, ref, "BALANCE")
        R.cell(row, 5).value = fr or 0
        if avec_n1:
            fr1 = f_rub_cr(rubs, ref, "BALANCE_N1")
            R.cell(row, 6).value = fr1 or 0
        for col in (5, 6):
            R.cell(row, col).number_format = FMT_MONTANT


# --------------------------------------------------------------------------
# TFT — postes ZA et FA à FQ, en formules Excel traçables
# --------------------------------------------------------------------------
# Formules transcrites de references/tft-formules-praticien.md (recoupées
# AUDCIF Titre IX §598-620). Les totaux ZB à ZH restent les formules du
# gabarit. Conditions : N-1 pour ZA/FB-FE, colonnes de mouvement pour FF-FQ.

TFT_ROW = {
    "ZA": 11, "FA": 14, "FB": 15, "FC": 16, "FD": 17, "FE": 18,
    "FF": 22, "FG": 23, "FH": 24, "FI": 25, "FJ": 26,
    "FK": 29, "FL": 30, "FM": 31, "FN": 32,
    "FO": 35, "FP": 36, "FQ": 37,
}


def _f(mode, tokens, feuille, exclude=()):
    return _corps(formule_tokens(list(tokens), mode, feuille, exclude=list(exclude)))


def construire_formules_tft(rubs, avec_n1, avec_mvt):
    """Renvoie ({ref: (formule_N, formule_N1|None)}, incomplets)."""
    B, B1 = "BALANCE", "BALANCE_N1"
    CR = "'Compte de Résultat'"
    tft, incomplet = {}, []

    def fa_pour(bal, col):
        to_h86 = f"{CR}!{col}44-{_f('c', ['86'], bal)}"
        rp_h85 = f"{CR}!{col}46-{_f('d', ['85'], bal)}"
        return (f"={CR}!{col}32+{_f('d', ['654'], bal)}-{_f('c', ['754'], bal)}"
                f"+{CR}!{col}41+({to_h86})-({rp_h85})-{CR}!{col}48-{CR}!{col}49")

    tft["FA"] = (fa_pour(B, "E"), fa_pour(B1, "F") if avec_n1 else None)

    def tresorerie_nette(bal):
        nets = "+".join(f"({_corps(f_rub_net(rubs, r, bal))})"
                        for r in ("BQ", "BR", "BS"))
        dq_dr = "+".join(f"({_corps(f_rub_passif(rubs, r, bal))})"
                         for r in ("DQ", "DR"))
        return f"({nets})-{_f('c', ['4726'], bal)}-({dq_dr})"

    if avec_n1:
        tft["ZA"] = (f"={tresorerie_nette(B1)}", None)

        def dnet(ref):
            return (f"({_corps(f_rub_net(rubs, ref, B))})"
                    f"-({_corps(f_rub_net(rubs, ref, B1))})")

        # Les lignes FB, FC, FD du gabarit se libellent « − Variation ... » et
        # le total ZB est une somme simple : la cellule porte donc l'OPPOSÉ de
        # la variation (une hausse des stocks pèse en négatif sur le flux).
        # FE (« + Variation du passif circulant ») porte la variation telle
        # quelle. Corrige une inversion de signe de la v2 du moteur.
        tft["FB"] = (f"=-({dnet('BA')}-({_f('d', ['485'], B)}-{_f('d', ['485'], B1)})"
                     f"+({_f('d', ['47818'], B)}-{_f('d', ['47818'], B1)})"
                     f"-({_f('c', ['47918'], B)}-{_f('c', ['47918'], B1)}))", None)
        tft["FC"] = (f"=-({dnet('BB')})", None)

        tiers_fd = ["414", "4494", "458", "461", "467", "4752"]
        t2714 = f"+{_f('md', ['2714', '2766'], B)}" if avec_mvt else ""
        if not avec_mvt:
            incomplet.append("FD (terme mouvement 2714/2766 non disponible : calculé sans lui)")
        tft["FD"] = (f"=-({dnet('BH')}+{dnet('BI')}+{dnet('BJ')}"
                     f"-({_f('d', tiers_fd, B)}-{_f('d', tiers_fd, B1)})"
                     f"+({_f('d', ['47811'], B)}-{_f('d', ['47811'], B1)})"
                     f"-({_f('c', ['47911'], B)}-{_f('c', ['47911'], B1)})"
                     f"{t2714})", None)

        dp_refs = ("DH", "DI", "DJ", "DK", "DM", "DN")

        def dp(bal):
            return "+".join(f"({_corps(f_rub_passif(rubs, r, bal))})" for r in dp_refs)

        tiers_fe = ["404", "461", "465", "4726", "481", "482"]
        t4752 = (f"+{_f('md', ['4752'], B)}-{_f('mc', ['4752'], B)}"
                 if avec_mvt else "")
        if not avec_mvt:
            incomplet.append("FE (terme mouvement 4752 non disponible : calculé sans lui)")
        tft["FE"] = (f"=({dp(B)})-({dp(B1)})"
                     f"-({_f('c', tiers_fe, B)}-{_f('c', tiers_fe, B1)})"
                     f"+({_f('c', ['4793'], B)}-{_f('c', ['4793'], B1)})"
                     f"-({_f('d', ['4783'], B)}-{_f('d', ['4783'], B1)})"
                     f"{t4752}", None)
    else:
        for ref in ("ZA", "FB", "FC", "FD", "FE"):
            tft[ref] = (None, None)
            incomplet.append(f"{ref} (balance N-1 absente)")

    if avec_mvt:
        def ad(bal):
            return "+".join(f"({_corps(f_rub_net(rubs, r, bal))})"
                            for r in ("AE", "AF", "AG", "AH"))

        def ai(bal):
            return "+".join(f"({_corps(f_rub_net(rubs, r, bal))})"
                            for r in ("AJ", "AK", "AL", "AM", "AN", "AP"))

        ad_d = f"({ad(B)})-({ad(B1)})" if avec_n1 else f"({ad(B)})"
        ai_d = f"({ai(B)})-({ai(B1)})" if avec_n1 else f"({ai(B)})"

        fournisseurs_inc = ["4041", "4046", "4811", "48161", "48171", "48181", "4821"]
        tft["FF"] = (f"={ad_d}+{_f('md', ['251'], B)}-{_f('mc', ['251'], B)}"
                     f"+{_f('md', fournisseurs_inc + ['281'], B)}"
                     f"-{_f('mc', fournisseurs_inc, B)}"
                     f"+{_f('d', ['6541', '811'], B)}", None)
        fournisseurs_c = ["4042", "4047", "4812", "48162", "48172", "48182", "4822"]
        tft["FG"] = (f"={ai_d}+{_f('md', ['252'], B)}-{_f('mc', ['252'], B)}"
                     f"+{_f('md', fournisseurs_c + ['282', '283', '284'], B)}"
                     f"-{_f('mc', ['17', '19842'] + fournisseurs_c, B)}"
                     f"-{_f('mc', ['106', '154'], B)}"
                     f"+{_f('d', ['6542', '812'], B)}", None)
        tft["FH"] = (f"={_f('md', ['26', '27'], B, exclude=['2714', '2766'])}"
                     f"+{_f('md', ['4813'], B)}-{_f('mc', ['4813'], B)}"
                     f"-{_f('mc', ['106', '154'], B)}"
                     f"+{_f('d', ['4782'], B)}-{_f('c', ['4792'], B)}", None)
        tft["FI"] = (f"={_f('c', ['754', '821', '822'], B)}"
                     f"-{_f('md', ['414', '485'], B, exclude=['4856'])}"
                     f"+{_f('mc', ['414', '485'], B, exclude=['4856'])}", None)
        tft["FJ"] = (f"={_f('c', ['826'], B)}"
                     f"+{_f('mc', ['27'], B, exclude=['2714', '2766'])}"
                     f"-{_f('md', ['4856'], B)}+{_f('mc', ['4856'], B)}", None)
        cap = ["101", "102", "1051"]
        cap_n1 = _f("c", cap, B1) if avec_n1 else "0"
        tft["FK"] = (f"={_f('c', cap, B)}-({cap_n1})"
                     f"-{_f('d', ['109', '4613', '467', '4581'], B)}"
                     f"-{_f('md', ['11', '12', '130', '131'], B)}"
                     f"+{_f('mc', ['103', '104', '11', '12', '139', '4619', '465'], B)}", None)
        f14_n1 = _f("c", ["14"], B1) if avec_n1 else "0"
        tft["FL"] = (f"={_f('c', ['14'], B)}-({f14_n1})"
                     f"+{_f('c', ['799'], B)}-{_f('d', ['4494', '4582'], B)}", None)
        tft["FM"] = (f"={_f('md', ['4619'], B)}+{_f('md', ['103', '104'], B)}", None)
        tft["FN"] = (f"={_f('md', ['465'], B)}", None)
        tft["FO"] = (f"={_f('mc', ['161', '162', '1661', '1662'], B)}"
                     f"+{_f('md', ['4713'], B)}-{_f('d', ['4784'], B)}", None)
        tft["FP"] = (f"={_f('mc', ['163', '164', '165', '166', '167', '168', '181', '182', '183'], B, exclude=['1661', '1662'])}"
                     f"-{_f('d', ['4784'], B)}", None)
        tft["FQ"] = (f"={_f('md', ['16', '17', '181', '182', '183'], B)}"
                     f"-{_f('c', ['4794'], B)}", None)
    else:
        for ref in ("FF", "FG", "FH", "FI", "FJ", "FK", "FL", "FM", "FN", "FO", "FP", "FQ"):
            tft[ref] = (None, None)
            incomplet.append(ref)

    return tft, sorted(set(incomplet))


def zh_depuis_bilan(bal, rubs):
    """Trésorerie nette de clôture recalculée en valeur, pour CONTROLES."""
    def net(ref):
        return (somme(bal, rubs[ref].brut, "brut_actif")
                - somme(bal, rubs[ref].amort, "amort"))

    def sc_tok(tok):
        e = Expr(include=[tok], sens="crediteur")
        return somme(bal, e, "?")

    dq = somme(bal, rubs["DQ"].brut, "passif")
    dr = somme(bal, rubs["DR"].brut, "passif")
    return net("BQ") + net("BR") + net("BS") - sc_tok("4726") - (dq + dr)


def injecter_tft(wb, tft):
    T = wb["TFT"]
    for ref, row in TFT_ROW.items():
        fN, fN1 = tft.get(ref, (None, None))
        if fN is not None:
            T.cell(row, 8).value = fN
            T.cell(row, 8).number_format = FMT_MONTANT
        if fN1 is not None:
            T.cell(row, 9).value = fN1
            T.cell(row, 9).number_format = FMT_MONTANT


# --------------------------------------------------------------------------
# Feuilles GARDE / BALANCE / CONTROLES / ANOMALIES
# --------------------------------------------------------------------------

def _affectations(l, rubs):
    """Liste des refs de rubriques qui captent ce compte (audit)."""
    refs = []
    for ref, r in rubs.items():
        if r.formule:
            continue
        if compte_dans_expr(l["compte"], r.brut) or compte_dans_expr(l["compte"], r.amort):
            refs.append(ref)
    return ", ".join(refs)


def ecrire_balance(wb, nom, bal, rubs):
    if nom in wb.sheetnames:
        del wb[nom]
    b = wb.create_sheet(nom)
    entetes = ["Compte", "Intitulé", "Préfixe 2", "Préfixe 3", "Préfixe 4",
               "Solde final débit", "Solde final crédit",
               "Mouvement débit", "Mouvement crédit", "Poste(s) d'affectation"]
    b.append(entetes)
    style_entetes(b, 1, 1, len(entetes))
    for l in bal:
        c = l["compte"]
        b.append([c, l["libelle"], c[:2], c[:3], c[:4],
                  round(l["sd"], 2), round(l["sc"], 2),
                  round(l.get("md", 0.0), 2), round(l.get("mc", 0.0), 2),
                  _affectations(l, rubs)])
    style_zone_donnees(b, 2, b.max_row, 1, len(entetes), cols_montant=(6, 7, 8, 9))
    largeurs(b, {"A": 12, "B": 42, "C": 9, "D": 9, "E": 9, "F": 15, "G": 15,
                 "H": 15, "I": 15, "J": 22})
    b.freeze_panes = "A2"
    return b


def ecrire_garde(wb, entite, identifiant, exercice, duree, avec_n1, avec_mvt):
    if "GARDE" in wb.sheetnames:
        del wb["GARDE"]
    g = wb.create_sheet("GARDE", 0)
    g.sheet_view.showGridLines = False
    g.merge_cells("B2:F3")
    g["B2"] = "LIASSE FISCALE OHADA — SYSCOHADA RÉVISÉ"
    g["B2"].font = F_TITRE
    g["B2"].fill = R_TITRE
    g["B2"].alignment = AL_CENTRE
    g.merge_cells("B4:F4")
    g["B4"] = "Système normal — États financiers annuels et notes annexes"
    g["B4"].font = F_SOUS_TITRE
    g["B4"].alignment = AL_CENTRE
    lignes = [
        ("Désignation de l'entité", entite or "—"),
        ("Numéro d'identification", identifiant or "—"),
        ("Exercice clos le", exercice or "—"),
        ("Durée de l'exercice (mois)", duree or "12"),
        ("Balance N-1 fournie", "Oui" if avec_n1 else "Non"),
        ("Colonnes de mouvement fournies", "Oui" if avec_mvt else "Non"),
    ]
    r = 6
    for lab, v in lignes:
        g[f"B{r}"] = lab
        g[f"B{r}"].font = F_GRAS
        g[f"D{r}"] = v
        r += 1
    r += 1
    g[f"B{r}"] = "Composition de la liasse"
    g[f"B{r}"].font = F_SOUS_TITRE
    r += 1
    sommaire = [
        "Fiches R1 à R4 — identification (à compléter)",
        "Bilan — Actif (feuille ACTIF) et Passif (feuille PASSIF)",
        "Compte de résultat",
        "Tableau des flux de trésorerie (TFT)",
        "Notes annexes 1 à 36 (une note par feuille)",
        "Feuilles d'audit : BALANCE, BALANCE_N1, CONTROLES, ANOMALIES",
    ]
    for s in sommaire:
        g[f"B{r}"] = "• " + s
        r += 1
    r += 1
    g[f"B{r}"] = ("Chaque poste et chaque ligne de note calculée porte une formule "
                  "Excel pointant vers la feuille BALANCE : "
                  "tout chiffre est retraçable jusqu'au compte qui l'alimente.")
    g[f"B{r}"].font = F_NORMAL
    g[f"B{r}"].alignment = AL_GAUCHE
    g.merge_cells(f"B{r}:F{r+2}")
    largeurs(g, {"A": 3, "B": 34, "C": 12, "D": 26, "E": 14, "F": 14})


def ajouter_controles(wb, bal, zh_val, avec_n1):
    if "CONTROLES" in wb.sheetnames:
        del wb["CONTROLES"]
    ctl = wb.create_sheet("CONTROLES")
    n = len(bal)
    ctl.append(["Contrôle", "Valeur", "Attendu"])
    style_entetes(ctl, 1, 1, 3)
    lignes = [
        ("Total solde débit balance", f"=SUM(BALANCE!F2:F{n+1})", ""),
        ("Total solde crédit balance", f"=SUM(BALANCE!G2:G{n+1})", ""),
        ("Écart balance (doit être 0)", "=B2-B3", 0),
        ("Total général actif net (BZ)", "=ACTIF!F40", ""),
        ("Total général passif (DZ)", "=PASSIF!D37", ""),
        ("Écart bilan actif - passif (doit être 0)", "=B5-B6", 0),
        ("Résultat net (compte de résultat, XI)", "='Compte de Résultat'!E50", ""),
        ("Trésorerie nette 31/12 recalculée depuis le bilan N", round(zh_val, 2), ""),
        ("Trésorerie nette 31/12 (TFT, poste ZH)", "='TFT'!H41", ""),
        ("Écart bilan - TFT trésorerie de clôture (0 si TFT complet)", "=B9-B10", 0),
        ("— Recoupements notes annexes / états —", "", ""),
        ("NOTE 4 total net vs immobilisations financières nettes (AR+AS)",
         "='NOTE 4'!B23-(ACTIF!F25+ACTIF!F26)", 0),
        ("NOTE 5 actif HAO net vs poste BA net", "='NOTE 5'!C19-ACTIF!F28", 0),
        ("NOTE 6 stocks nets vs poste BB net", "='NOTE 6'!C24-ACTIF!F29", 0),
        ("NOTE 7 clients nets vs poste BI net", "='NOTE 7'!B24-ACTIF!F32", 0),
        ("NOTE 8 autres créances nettes vs poste BJ net", "='NOTE 8'!B28-ACTIF!F33", 0),
        ("NOTE 9 titres nets vs poste BQ net", "='NOTE 9'!B21-ACTIF!F35", 0),
        ("NOTE 10 valeurs à encaisser nettes vs poste BR net", "='NOTE 10'!B19-ACTIF!F36", 0),
        ("NOTE 11 disponibilités nettes vs poste BS net", "='NOTE 11'!C27-ACTIF!F37", 0),
        ("NOTE 15A total vs postes CL+CM", "='NOTE 15A'!C29-(PASSIF!D18+PASSIF!D19)", 0),
        ("NOTE 16A total vs postes DA+DB+DC",
         "='NOTE 16A'!B17+'NOTE 16A'!B24+'NOTE 16A'!B39-PASSIF!D24", 0),
        ("NOTE 17 fournisseurs vs poste DJ", "='NOTE 17'!B15-PASSIF!D28", 0),
        ("NOTE 18 total vs poste DK", "='NOTE 18'!B26-PASSIF!D29", 0),
        ("NOTE 19 autres dettes vs poste DM", "='NOTE 19'!B33-PASSIF!D30", 0),
        ("NOTE 20 total vs postes DQ+DR", "='NOTE 20'!B24-(PASSIF!D33+PASSIF!D34)", 0),
        ("NOTE 21 CA vs poste XB", "='NOTE 21'!B34-'Compte de Résultat'!E16", 0),
        ("NOTE 22 total achats vs RA+RC+RE",
         "='NOTE 22'!B14+'NOTE 22'!B19+'NOTE 22'!B35-"
         "('Compte de Résultat'!E10+'Compte de Résultat'!E22+'Compte de Résultat'!E24)", 0),
        ("NOTE 23 transports vs poste RG", "='NOTE 23'!B15-'Compte de Résultat'!E26", 0),
        ("NOTE 24 services extérieurs vs poste RH", "='NOTE 24'!B24-'Compte de Résultat'!E27", 0),
        ("NOTE 25 impôts et taxes vs poste RI", "='NOTE 25'!B16-'Compte de Résultat'!E28", 0),
        ("NOTE 26 autres charges vs poste RJ", "='NOTE 26'!B19-'Compte de Résultat'!E29", 0),
        ("NOTE 27A personnel vs poste RK", "='NOTE 27A'!B18-'Compte de Résultat'!E31", 0),
    ]
    for lab, f, att in lignes:
        ctl.append([lab, f, att])
    style_zone_donnees(ctl, 2, ctl.max_row, 1, 3, cols_montant=(2,))
    largeurs(ctl, {"A": 62, "B": 20, "C": 10})
    return ctl


def ajouter_anomalies(wb, anomalies):
    if "ANOMALIES" in wb.sheetnames:
        del wb["ANOMALIES"]
    an = wb.create_sheet("ANOMALIES")
    an.append(["Gravité", "Compte", "Intitulé", "Problème", "Solution proposée"])
    style_entetes(an, 1, 1, 5)
    ordre = {"BLOQUANT": 0, "A_TRAITER": 1, "A_VERIFIER": 2, "MINEUR": 3, "INFO": 4}
    for x in sorted(anomalies, key=lambda z: ordre.get(z["gravite"], 9)):
        an.append([x["gravite"], x["compte"], x["libelle"], x["probleme"], x["solution"]])
    style_zone_donnees(an, 2, max(an.max_row, 2), 1, 5)
    largeurs(an, {"A": 12, "B": 12, "C": 26, "D": 62, "E": 62})
    an.freeze_panes = "A2"


# --------------------------------------------------------------------------
# Programme principal
# --------------------------------------------------------------------------

def main():
    ici = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser()
    ap.add_argument("balance_N")
    ap.add_argument("balance_N1", nargs="?")
    ap.add_argument("--gabarit", default=os.path.join(ici, "..", "assets", "gabarit-liasse.xlsx"))
    ap.add_argument("--correspondance", default=os.path.join(ici, "..", "references", "correspondance.tsv"))
    ap.add_argument("--sortie", default="liasse.xlsx")
    ap.add_argument("--entite", default="")
    ap.add_argument("--identifiant", default="")
    ap.add_argument("--exercice", default="")
    ap.add_argument("--duree", default="12")
    args = ap.parse_args()

    rubs = charger_maquette(args.correspondance)
    bal, idx = lire_balance(args.balance_N)
    print(f"Balance N : {len(bal)} comptes. Colonnes repérées : {idx}")

    anomalies = detecter_anomalies(bal, rubs)
    avec_mvt = any((l.get("md") or l.get("mc")) for l in bal)

    bal1 = None
    if args.balance_N1:
        bal1, _ = lire_balance(args.balance_N1)
        print(f"Balance N-1 : {len(bal1)} comptes.")
    avec_n1 = bal1 is not None

    import formules as _fm
    _fm.set_lignes_max(max(len(bal), len(bal1 or [])) + 20)

    tft, tft_incomplet = construire_formules_tft(rubs, avec_n1, avec_mvt)
    for ref in tft_incomplet:
        anomalies.append({
            "gravite": "INFO", "compte": "", "libelle": f"TFT — {ref}",
            "probleme": "Poste du TFT non calculé (colonne de mouvement de "
                        "l'exercice absente de la balance, ou balance N-1 non fournie).",
            "solution": "Fournir une balance avec colonnes 'mouvement débit'/'mouvement "
                        "crédit' (et la balance N-1) pour compléter ce poste. Voir "
                        "references/tft-formules-praticien.md.",
        })
    if not avec_n1:
        anomalies.append({
            "gravite": "INFO", "compte": "", "libelle": "Notes 3A/3C/28",
            "probleme": "Balance N-1 absente : les colonnes d'ouverture des notes de "
                        "mouvements portent la clôture N faute d'ouverture connue.",
            "solution": "Fournir la balance N-1 pour des notes de mouvements complètes.",
        })

    wb = openpyxl.load_workbook(args.gabarit)
    injecter(wb, rubs, avec_n1)
    injecter_tft(wb, tft)

    ecrire_balance(wb, "BALANCE", bal, rubs)
    if avec_n1:
        ecrire_balance(wb, "BALANCE_N1", bal1, rubs)

    notes_sn.injecter_notes(wb, avec_n1, avec_mvt,
                            entite=args.entite, identifiant=args.identifiant,
                            exercice=args.exercice, duree=args.duree)

    zh_val = zh_depuis_bilan(bal, rubs)
    ajouter_controles(wb, bal, zh_val, avec_n1)
    ajouter_anomalies(wb, anomalies)
    ecrire_garde(wb, args.entite, args.identifiant, args.exercice, args.duree,
                 avec_n1, avec_mvt)

    retirer_tirets(wb)
    wb.save(args.sortie)
    print(f"Liasse écrite : {args.sortie}")
    print("Ouvrir dans Excel/LibreOffice (ou convertir via soffice --headless "
          "--convert-to xlsx) pour recalculer les formules.")

    bloquants = [a for a in anomalies if a["gravite"] in ("BLOQUANT", "A_TRAITER")]
    print(f"Anomalies : {len(anomalies)} dont {len(bloquants)} à traiter avant remise.")
    for a in anomalies[:12]:
        print(f"  [{a['gravite']}] {a['compte']} {a['probleme']}")


if __name__ == "__main__":
    main()
